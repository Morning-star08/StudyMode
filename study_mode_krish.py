"""
Study Mode - Krish Baral
Updated UI and workflow wrapper for IOE entrance preparation.

Run with pythonw.exe on Windows for a fully silent background launch.
"""

import ctypes
import json
import os
import random
import subprocess
import threading
import time
from ctypes import wintypes
from datetime import date, datetime, timedelta

try:
    import winsound
except ImportError:
    winsound = None

os.environ.setdefault("PYGAME_HIDE_SUPPORT_PROMPT", "1")

import keyboard
import pystray
import tkinter as tk
from PIL import Image, ImageDraw, ImageTk
from plyer import notification

import _study_mode_base as base
from ioe_study_plan import (
    JOURNEY_END_DATE,
    JOURNEY_START_DATE,
    JOURNEY_TOTAL_DAYS,
    current_day_entry as canonical_current_day_entry,
    derive_subject_progress,
    english_quick_rules as canonical_english_quick_rules,
    formula_sheet_for_subject as canonical_formula_sheet_for_subject,
    journey_day_number as canonical_journey_day_number,
    load_study_plan,
    milestone_for_day,
    subject_context as canonical_subject_context,
    subject_for_day_number,
    weekly_target_for_day,
)


LOFI_URL = base.LOFI_URL
BRAVE = base.BRAVE
SUBJECT_URLS = base.SUBJECT_URLS
CURRICULUM = base.CURRICULUM
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
STREAK_FILE = os.path.join(BASE_DIR, "study_streak.json")
EXCEL_FILE = os.path.join(BASE_DIR, "study_tracker.xlsx")
NOTES_FILE = os.path.join(BASE_DIR, "study_notes.json")
POMODORO_WORK = base.POMODORO_WORK
POMODORO_BREAK = base.POMODORO_BREAK
MIN_SESSION = base.MIN_SESSION
LATE_NIGHT_H = base.LATE_NIGHT_H
IOE_DATE = JOURNEY_END_DATE
TOTAL_TARGET_HOURS = 400
MCQ_PRACTICE_URL = "https://www.youtube.com/results?search_query=IOE+entrance+exam+MCQ+2024"
MOCK_EXAM_SECONDS = 2 * 60 * 60
MOCK_EXAM_MARKS = {
    "MATHEMATICS": 50,
    "PHYSICS": 40,
    "CHEMISTRY": 30,
    "ENGLISH": 20,
}
SOUND_MODES = ["off", "soft", "strong"]
ALARM_SCHEDULE = {
    "morning": (7, 45),
    "evening": (19, 30),
}
SESSION_GOALS = [
    "Understand the concept",
    "Solve practice MCQs",
    "Revise notes",
    "Watch video + notes",
]
MOTIVATIONAL_QUOTES = [
    "The expert was once a beginner.",
    "IOE waits for no one.",
    "One topic a day beats zero.",
    "Small sessions build big results.",
    "Today's focus becomes tomorrow's confidence.",
    "You do not need perfect. You need consistent.",
    "One solved problem teaches more than one saved video.",
    "Six months of discipline can change your direction.",
    "Every honest session counts.",
    "Momentum is built in ordinary days.",
    "A calm hour today saves panic later.",
    "Study the chapter in front of you.",
    "Progress loves repetition.",
    "Hard topics become easy after enough returns.",
    "Write it once. Solve it twice. Remember it longer.",
    "Your future seat is earned in these quiet sessions.",
    "Focus now so exam day feels familiar.",
    "Confidence is collected, not wished for.",
    "A strong streak starts with one clean session.",
    "Solve first. Doubt later.",
    "Show up even when motivation is late.",
    "Krish, one more focused hour matters.",
    "Revision is how memory becomes speed.",
    "Keep stacking the ordinary wins.",
]

BG = base.BG
CYAN = base.CYAN
GREEN = base.GREEN
GOLD = base.GOLD
RED = base.RED
DIM = base.DIM
WHITE = base.WHITE
GREY = base.GREY

FONT_FACE = "Courier New"
POPUP_HEADER_PT = 18
POPUP_BODY_PT = 12
POPUP_TOPIC_PT = 20
POPUP_TIMER_PT = 28
POPUP_BG = "#070b0d"
POPUP_PANEL = "#0d1418"
POPUP_TEXT = WHITE
POPUP_MUTED = "#9ac9bc"
POPUP_DIM = "#4f7165"
BUTTON_BG = "#0f171a"
BUTTON_ACTIVE = "#162327"
WINDOWS_TOAST_SOUND = "ms-winsoundevent:Notification.Reminder"
HABIT_SOUND_PATTERNS = {
    "goal_pick": [(740, 55), (988, 75)],
    "session_start": [(523, 60), (659, 70), (784, 110)],
    "pause": [(420, 70), (360, 90)],
    "resume": [(494, 55), (659, 85)],
    "break_start": [(880, 70), (1175, 90)],
    "break_end": [(587, 60), (740, 65), (932, 95)],
    "success": [(523, 65), (659, 70), (784, 75), (1047, 140)],
    "steady": [(523, 70), (659, 95)],
    "gentle_end": [(392, 100), (440, 120)],
    "alarm": [(784, 65), (988, 75), (1175, 85)],
}

SUBJECT_HEX = {
    "MATHEMATICS": CYAN,
    "PHYSICS": "#58b7ff",
    "CHEMISTRY": "#8cf86d",
    "ENGLISH": GOLD,
    "MOCK EXAM": "#ff7a59",
    "General": CYAN,
}
SUBJECT_MARKS = {
    "MATHEMATICS": 50,
    "PHYSICS": 40,
    "CHEMISTRY": 30,
    "ENGLISH": 20,
}
GRADE_MESSAGES = {
    "S": "Outstanding week. Keep this pace and IOE gets closer fast.",
    "A": "Strong consistency. One more study day next week pushes this to S.",
    "B": "Good recovery week. Build just one more study day to level up.",
    "C": "A restart week is still a win. Show up tomorrow and rebuild momentum.",
}

STUDY_PLAN_DATA = load_study_plan()
JOURNEY_LAST_STUDY_DAY = JOURNEY_END_DATE
BUFFER_TOPIC = "Buffer Review & Catch-up"

DAILY_LESSON_LIBRARY = {
    "MATHEMATICS": [
        {"topic": "Set & Logic", "days": 2, "focus": "sets, intervals, and logic rules", "match": ["Set", "Logic"]},
        {"topic": "Functions", "days": 2, "focus": "function types and relations", "match": ["Functions"]},
        {"topic": "Matrices", "days": 3, "focus": "matrix operations and determinants", "match": ["Matrices"]},
        {"topic": "Complex Numbers", "days": 2, "focus": "complex operations and roots", "match": ["Complex"]},
        {"topic": "Sequence & Series", "days": 3, "focus": "AP, GP, and summation", "match": ["Sequence", "Series"]},
        {"topic": "Binomial Theorem", "days": 2, "focus": "expansion and coefficients", "match": ["Binomial"]},
        {"topic": "Trigonometry", "days": 4, "focus": "identities, equations, and triangles", "match": ["Trigonometric", "Triangles"]},
        {"topic": "Coordinate Geometry", "days": 5, "focus": "lines, circles, and conics", "match": ["Straight Lines", "Circles", "Conic", "3D Coordinates"]},
        {"topic": "Calculus - Limits", "days": 3, "focus": "limits and continuity", "match": ["Limits", "Continuity"]},
        {"topic": "Calculus - Derivatives", "days": 4, "focus": "derivatives and applications", "match": ["Derivatives", "Applications of Derivatives"]},
        {"topic": "Calculus - Integration", "days": 5, "focus": "integration and area", "match": ["Integration"]},
        {"topic": "Differential Equations", "days": 4, "focus": "solving first-order DEs", "match": ["Differential Equations"]},
        {"topic": "Vectors", "days": 3, "focus": "vector algebra and products", "match": ["Vectors", "Products of Vectors"]},
        {"topic": "Statistics & Probability", "days": 3, "focus": "data handling and probability models", "match": ["Statistics", "Probability"]},
        {"topic": "Practice MCQs", "days": 6, "focus": "timed mixed-question practice", "match": ["PRACTICE"]},
        {"topic": "Full Revision", "days": 5, "focus": "weak-topic revision and formula recall", "match": ["REVISION", "FINAL"]},
    ],
    "PHYSICS": [
        {"topic": "Vectors & Kinematics", "days": 3, "focus": "motion graphs and vector basics", "match": ["Vectors", "Kinematics"]},
        {"topic": "Newton's Laws", "days": 3, "focus": "forces, friction, and momentum", "match": ["Newton"]},
        {"topic": "Work Energy Power", "days": 2, "focus": "work-energy theorem and collisions", "match": ["Work", "Energy", "Power"]},
        {"topic": "Circular Motion & Gravitation", "days": 3, "focus": "satellite motion and centripetal force", "match": ["Circular Motion", "Gravitation"]},
        {"topic": "Rotational Dynamics", "days": 3, "focus": "torque, angular momentum, and MOI", "match": ["Rotational"]},
        {"topic": "Elasticity & Fluids", "days": 3, "focus": "moduli, pressure, and Bernoulli", "match": ["Elasticity", "Fluid"]},
        {"topic": "Heat & Thermodynamics", "days": 4, "focus": "heat transfer and thermodynamic laws", "match": ["Heat", "Thermodynamics"]},
        {"topic": "Optics", "days": 4, "focus": "mirrors, lenses, and wave optics", "match": ["Optics", "Light"]},
        {"topic": "Waves & Sound", "days": 3, "focus": "wave motion, standing waves, and Doppler", "match": ["Waves", "Sound"]},
        {"topic": "Electrostatics", "days": 3, "focus": "charge, field, and capacitance", "match": ["Electrostatics"]},
        {"topic": "DC Circuits", "days": 3, "focus": "Ohm, Kirchhoff, and circuit solving", "match": ["DC Circuits"]},
        {"topic": "Magnetism & EM Induction", "days": 4, "focus": "magnetic force, induction, and AC", "match": ["Magnetism", "EM Induction"]},
        {"topic": "Modern Physics", "days": 4, "focus": "photoelectric effect, atom, and nucleus", "match": ["Modern Physics", "Radioactivity"]},
        {"topic": "Practice MCQs", "days": 5, "focus": "timed mechanics to modern physics MCQs", "match": ["PRACTICE"]},
        {"topic": "Full Revision", "days": 4, "focus": "formula recall and full-paper review", "match": ["REVISION"]},
    ],
    "CHEMISTRY": [
        {"topic": "Chemical Arithmetic", "days": 2, "focus": "mole concept and stoichiometry", "match": ["Chemical Arithmetic"]},
        {"topic": "States of Matter", "days": 2, "focus": "gas laws and intermolecular forces", "match": ["States of Matter"]},
        {"topic": "Atomic Structure", "days": 3, "focus": "electronic configuration and periodic trends", "match": ["Atomic Structure", "Periodic Table"]},
        {"topic": "Oxidation & Equilibrium", "days": 3, "focus": "redox and equilibrium shifts", "match": ["Oxidation", "Equilibrium"]},
        {"topic": "Acid Base Salt", "days": 2, "focus": "pH, buffers, and ionic equilibrium", "match": ["Acid", "Base", "Salt"]},
        {"topic": "Electrochemistry", "days": 2, "focus": "cells and electrolysis", "match": ["Electrochemistry"]},
        {"topic": "Chemical Kinetics & Bonding", "days": 3, "focus": "rates, energetics, and bonding", "match": ["Energetics", "Kinetics", "Bonding"]},
        {"topic": "Non-metals", "days": 3, "focus": "important non-metals and compounds", "match": ["Non-metals"]},
        {"topic": "Metals & Metallurgy", "days": 2, "focus": "metallurgy principles and metal reactions", "match": ["Metals", "Metallurgy"]},
        {"topic": "Organic Fundamentals", "days": 3, "focus": "nomenclature, isomerism, and mechanisms", "match": ["Organic", "Fundamentals"]},
        {"topic": "Hydrocarbons", "days": 3, "focus": "alkanes, alkenes, and aromatic basics", "match": ["Hydrocarbons"]},
        {"topic": "Functional Groups", "days": 4, "focus": "alcohols, carbonyls, acids, and amines", "match": ["Functional Groups"]},
        {"topic": "Practice MCQs", "days": 4, "focus": "mixed chemistry entrance MCQs", "match": ["PRACTICE"]},
        {"topic": "Full Revision", "days": 3, "focus": "reaction charts and weak-area revision", "match": ["REVISION"]},
    ],
    "ENGLISH": [
        {"topic": "Tense & Concord", "days": 1, "focus": "tense usage and subject-verb agreement", "match": ["Tense", "Concord"]},
        {"topic": "Direct & Indirect Speech", "days": 1, "focus": "reported speech rules", "match": ["Direct", "Indirect Speech"]},
        {"topic": "Sentence Transformation", "days": 1, "focus": "sentence change and structure", "match": ["Sentence Transformation"]},
        {"topic": "Conditionals", "days": 1, "focus": "zero to third conditional patterns", "match": ["Conditionals"]},
        {"topic": "Active & Passive Voice", "days": 1, "focus": "voice conversion practice", "match": ["Active", "Passive Voice"]},
        {"topic": "Verbals & Parts of Speech", "days": 1, "focus": "verbals and grammar roles", "match": ["Verbals", "Parts of Speech"]},
        {"topic": "Prepositions & Punctuation", "days": 1, "focus": "usage patterns and punctuation", "match": ["Prepositions", "Punctuation"]},
        {"topic": "Vocabulary & Idioms", "days": 2, "focus": "word power and common idioms", "match": ["Vocabulary", "Idioms"]},
        {"topic": "Phonetics", "days": 1, "focus": "sound symbols and stress", "match": ["Phonetics"]},
        {"topic": "Reading Comprehension", "days": 2, "focus": "passage strategy and inference", "match": ["Reading Comprehension"]},
        {"topic": "Practice MCQs", "days": 2, "focus": "timed grammar and reading MCQs", "match": ["PRACTICE"]},
    ],
}


study_active = False
study_start_time = None
tray_icon = None
running = True
current_subject = None
current_session_goal = ""
current_session_topic = ""
current_session_quote = ""
current_session_recommended = ""
current_session_mode = "study"
current_session_day_number = 1
advance_after_session = False
excel_error_notified = False
_tk_root = None
current_resource_url = ""
break_window_snapshot = set()

WM_CLOSE = 0x0010
BREAK_WINDOW_PROCESS_ALLOWLIST = {
    "applicationframehost.exe",
    "cmd.exe",
    "conhost.exe",
    "dwm.exe",
    "explorer.exe",
    "powershell.exe",
    "python.exe",
    "pythonw.exe",
    "searchhost.exe",
    "shellexperiencehost.exe",
    "startmenuexperiencehost.exe",
    "taskmgr.exe",
    "textinputhost.exe",
    "widgets.exe",
    "wscript.exe",
}
BREAK_WINDOW_TITLE_ALLOWLIST = {
    "",
    "Program Manager",
    "Windows Input Experience",
}


def hide_console_window():
    if os.name != "nt":
        return
    try:
        hwnd = ctypes.windll.kernel32.GetConsoleWindow()
        if hwnd:
            ctypes.windll.user32.ShowWindow(hwnd, 0)
    except Exception:
        pass


def hex_to_rgb(value):
    value = value.lstrip("#")
    return tuple(int(value[i:i + 2], 16) for i in (0, 2, 4))


def xml_escape(text):
    return (
        str(text)
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
        .replace("'", "&apos;")
    )


def trim_text(text, max_chars):
    text = (text or "").strip()
    if len(text) <= max_chars:
        return text
    return text[: max_chars - 3].rstrip() + "..."


def subject_hex(subject):
    return SUBJECT_HEX.get(subject, CYAN)


def parse_iso_date(value):
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except Exception:
        return None


def recommended_subject_for_day(day=None):
    day = day or date.today()
    if day < JOURNEY_START_DATE:
        weekday_sequence = [
            "MATHEMATICS",
            "PHYSICS",
            "CHEMISTRY",
            "MATHEMATICS",
            "PHYSICS",
            "CHEMISTRY",
            "ENGLISH",
        ]
        return weekday_sequence[day.weekday()]
    return subject_for_day_number(canonical_journey_day_number(day))


def safe_open_url(url):
    try:
        if os.path.exists(BRAVE):
            subprocess.Popen([BRAVE, url])
        else:
            subprocess.Popen(["cmd", "/c", "start", "", url])
    except Exception:
        pass


def _run_powershell_capture(script):
    try:
        result = subprocess.run(
            ["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", script],
            capture_output=True,
            text=True,
            timeout=8,
            creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0),
        )
        if result.returncode != 0:
            return None
        return (result.stdout or "").strip()
    except Exception:
        return None


def _window_text(hwnd):
    length = ctypes.windll.user32.GetWindowTextLengthW(hwnd)
    buffer = ctypes.create_unicode_buffer(length + 1)
    ctypes.windll.user32.GetWindowTextW(hwnd, buffer, len(buffer))
    return buffer.value.strip()


def _process_name_for_pid(pid):
    if not pid:
        return ""
    try:
        output = subprocess.check_output(
            ["tasklist", "/fi", f"PID eq {pid}", "/fo", "csv", "/nh"],
            text=True,
            creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0),
        ).strip()
        if not output or output.startswith("INFO:"):
            return ""
        return output.split(",")[0].strip('"').lower()
    except Exception:
        return ""


def _list_closeable_windows():
    if os.name != "nt":
        return []

    windows = []

    @ctypes.WINFUNCTYPE(wintypes.BOOL, wintypes.HWND, wintypes.LPARAM)
    def enum_proc(hwnd, _lparam):
        try:
            if not ctypes.windll.user32.IsWindowVisible(hwnd):
                return True
            title = _window_text(hwnd)
            if title in BREAK_WINDOW_TITLE_ALLOWLIST:
                return True
            if ctypes.windll.user32.GetWindow(hwnd, 4):
                return True
            pid = wintypes.DWORD()
            ctypes.windll.user32.GetWindowThreadProcessId(hwnd, ctypes.byref(pid))
            process_name = _process_name_for_pid(pid.value)
            if process_name in BREAK_WINDOW_PROCESS_ALLOWLIST:
                return True
            windows.append({"hwnd": int(hwnd), "title": title, "pid": int(pid.value), "process_name": process_name})
        except Exception:
            pass
        return True

    try:
        ctypes.windll.user32.EnumWindows(enum_proc, 0)
    except Exception:
        return []
    return windows


def snapshot_break_windows():
    global break_window_snapshot
    break_window_snapshot = {item["hwnd"] for item in _list_closeable_windows()}


def close_windows_opened_during_break():
    if os.name != "nt" or not break_window_snapshot:
        return
    for item in _list_closeable_windows():
        if item["hwnd"] in break_window_snapshot:
            continue
        try:
            ctypes.windll.user32.PostMessageW(item["hwnd"], WM_CLOSE, 0, 0)
        except Exception:
            pass


def pause_study_browser_for_break():
    snapshot_break_windows()
    base.close_brave()


def restore_focus_after_break():
    close_windows_opened_during_break()
    base.close_brave()
    if current_resource_url:
        safe_open_url(LOFI_URL)
        time.sleep(1.5)
        safe_open_url(current_resource_url)


def load_streak():
    defaults = {
        "streak": 0,
        "last_date": "",
        "last_subject": "",
        "total_sessions": 0,
        "total_minutes": 0,
        "sessions": [],
        "day_progress": 0,
        "lesson_progress": {},
        "topic_progress": {},
        "alarm_history": {},
        "settings": {"sound": "soft"},
        "weak_topics": [],
        "mock_exams": [],
        "weekly_summary_last_shown": "",
    }
    data = {}
    try:
        if os.path.exists(STREAK_FILE):
            with open(STREAK_FILE, "r", encoding="utf-8") as file:
                data = json.load(file)
    except Exception:
        data = {}
    for key, value in defaults.items():
        if key not in data:
            data[key] = value.copy() if isinstance(value, dict) else list(value) if isinstance(value, list) else value
    return data


def save_streak(data):
    with open(STREAK_FILE, "w", encoding="utf-8") as file:
        json.dump(data, file, indent=2, ensure_ascii=False)


def get_sound_mode():
    data = load_streak()
    settings = data.setdefault("settings", {})
    mode = settings.get("sound", "soft")
    if mode not in SOUND_MODES:
        mode = "soft"
        settings["sound"] = mode
        save_streak(data)
    return mode


def cycle_sound_mode():
    data = load_streak()
    settings = data.setdefault("settings", {})
    current = settings.get("sound", "soft")
    if current not in SOUND_MODES:
        current = "soft"
    mode = SOUND_MODES[(SOUND_MODES.index(current) + 1) % len(SOUND_MODES)]
    settings["sound"] = mode
    save_streak(data)
    notification.notify(title="Study sound", message=f"Sound mode: {mode.upper()}", timeout=4)
    if mode != "off":
        play_habit_sound("goal_pick")
    return mode


def topic_key(subject, topic):
    return f"{subject}|{(topic or '').strip().lower()}"


def record_topic_difficulty(data, subject, topic, difficulty):
    if difficulty not in ("easy", "medium", "hard"):
        return
    weak_topics = data.setdefault("weak_topics", [])
    key = topic_key(subject, topic)
    existing = None
    for item in weak_topics:
        if topic_key(item.get("subject", ""), item.get("topic", "")) == key:
            existing = item
            break
    if existing is None:
        existing = {
            "subject": subject,
            "topic": topic,
            "hard_count": 0,
            "status": "active",
            "created": str(date.today()),
        }
        weak_topics.append(existing)
    existing["last_seen"] = str(date.today())
    existing["last_difficulty"] = difficulty
    existing["status"] = "active" if difficulty == "hard" else "watch" if difficulty == "medium" else "cleared"
    existing["hard_count"] = int(existing.get("hard_count", 0)) + (1 if difficulty == "hard" else 0)


def active_weak_topics(data=None, subject=None, limit=6):
    data = data or load_streak()
    items = []
    for item in data.get("weak_topics", []):
        if subject and item.get("subject") != subject:
            continue
        if item.get("status") not in ("active", "watch"):
            continue
        items.append(item)
    items.sort(key=lambda item: (item.get("last_seen", ""), int(item.get("hard_count", 0))), reverse=True)
    return items[:limit]


def is_weak_topic(subject, topic):
    key = topic_key(subject, topic)
    for item in active_weak_topics(subject=subject, limit=50):
        if topic_key(item.get("subject", ""), item.get("topic", "")) == key:
            return True
    return False


def latest_topic_sessions(data=None, subject=None):
    data = data or load_streak()
    latest = {}
    for session in data.get("sessions", []):
        session_subject = session.get("subject", "")
        session_topic = (session.get("topic", "") or "").strip()
        session_date = parse_iso_date(session.get("date", ""))
        if not session_date or not session_topic:
            continue
        if session_subject == "MOCK EXAM":
            continue
        if subject and session_subject != subject:
            continue
        key = (session_subject, session_topic)
        if key not in latest or session_date > latest[key]["date"]:
            latest[key] = {
                "subject": session_subject,
                "topic": session_topic,
                "date": session_date,
            }
    return latest


def spaced_review_candidates(subject, current_topic, today=None, data=None):
    data = data or load_streak()
    today = today or date.today()
    candidates = []
    for item in latest_topic_sessions(data=data, subject=subject).values():
        if item["topic"] == current_topic:
            continue
        gap = (today - item["date"]).days
        review_level = max([days for days in (3, 7, 14) if gap >= days], default=0)
        if review_level <= 0:
            continue
        candidates.append(
            {
                "subject": item["subject"],
                "topic": item["topic"],
                "gap": gap,
                "review_level": review_level,
            }
        )
    candidates.sort(key=lambda item: (item["review_level"], item["gap"]), reverse=True)
    return candidates


def highest_priority_review_topic(data=None, exclude_subject=None):
    data = data or load_streak()
    priority = {"MATHEMATICS": 0, "PHYSICS": 1, "CHEMISTRY": 2, "ENGLISH": 3}
    candidates = []
    for item in active_weak_topics(data=data, limit=20):
        if exclude_subject and item.get("subject") == exclude_subject:
            continue
        candidates.append(item)
    if candidates:
        candidates.sort(
            key=lambda item: (
                -priority.get(item.get("subject", "ENGLISH"), 9),
                int(item.get("hard_count", 0)),
                item.get("last_seen", ""),
            ),
            reverse=True,
        )
        return candidates[0]

    latest = list(latest_topic_sessions(data=data).values())
    latest = [item for item in latest if item["subject"] != exclude_subject]
    latest.sort(
        key=lambda item: (
            -priority.get(item["subject"], 9),
            item["date"],
        ),
        reverse=True,
    )
    return latest[0] if latest else None


def adaptive_routine(subject, entry, data=None, today=None):
    data = data or load_streak()
    today = today or date.today()
    current_topic = entry["topic"]
    weak_same_subject = [
        item
        for item in active_weak_topics(data=data, subject=subject, limit=10)
        if topic_key(subject, item.get("topic", "")) != topic_key(subject, current_topic)
    ]
    due_reviews = spaced_review_candidates(subject, current_topic, today=today, data=data)

    if subject == "ENGLISH":
        cross_topic = highest_priority_review_topic(data=data, exclude_subject="ENGLISH")
        if cross_topic:
            return {
                "mode": "light-english-plus-core-review",
                "headline": "Keep English light and use the extra attention for a core-subject comeback.",
                "review_target": f"{cross_topic['subject']}: {cross_topic['topic']}",
                "split": "About 70 minutes English, 20 minutes core review.",
                "morning_adjustment": "Add a 5-minute recall of one English rule before new material.",
                "evening_adjustment": f"Use 20 evening minutes to rewrite formulas or corrections from {cross_topic['subject']}: {cross_topic['topic']}.",
            }

    if weak_same_subject:
        target = weak_same_subject[0]
        return {
            "mode": "weak-topic-repair",
            "headline": "Do not study this subject in a straight line today. Repair the leak while moving forward.",
            "review_target": f"{target['subject']}: {target['topic']}",
            "split": "Roughly 70% today's planned topic, 30% weak-topic repair.",
            "morning_adjustment": f"Start with a 5-10 minute retrieval recap from {target['topic']}.",
            "evening_adjustment": f"Spend the last 20 minutes redoing 5-8 mistakes from {target['topic']}.",
        }

    if due_reviews:
        target = due_reviews[0]
        return {
            "mode": "spaced-review",
            "headline": "Bring back an older topic before memory decays. This is how progress sticks.",
            "review_target": f"{target['subject']}: {target['topic']} ({target['gap']} days ago)",
            "split": "Roughly 80% today's topic, 20% spaced review.",
            "morning_adjustment": f"Use the first 5 minutes to recall {target['topic']} from memory.",
            "evening_adjustment": f"Finish with 4-6 review MCQs from {target['topic']}.",
        }

    return {
        "mode": "forward-plus-recall",
        "headline": "Push the planned topic, but begin with retrieval so the learning compounds.",
        "review_target": "Yesterday's notes and one formula sheet line",
        "split": "Mainly new topic work with a short recall block.",
        "morning_adjustment": "Begin by recalling yesterday's key ideas without looking at notes.",
        "evening_adjustment": "End by writing one quick error-prevention note for tomorrow.",
    }


def record_mock_exam(minutes, scores):
    data = load_streak()
    total = sum(int(scores.get(subject, 0)) for subject in MOCK_EXAM_MARKS)
    max_total = sum(MOCK_EXAM_MARKS.values())
    data.setdefault("mock_exams", []).append(
        {
            "date": str(date.today()),
            "time": datetime.now().strftime("%H:%M"),
            "minutes": minutes,
            "scores": scores,
            "total": total,
            "max_total": max_total,
            "percent": round((total / max_total) * 100, 1) if max_total else 0,
        }
    )
    save_streak(data)
    return data


def load_notes():
    try:
        if os.path.exists(NOTES_FILE):
            with open(NOTES_FILE, "r", encoding="utf-8") as file:
                data = json.load(file)
                if isinstance(data, list):
                    return data
    except Exception:
        pass
    return []


def save_notes(notes):
    with open(NOTES_FILE, "w", encoding="utf-8") as file:
        json.dump(notes, file, indent=2, ensure_ascii=False)


def get_lesson_progress(subject):
    data = load_streak()
    return data.get("lesson_progress", {}).get(subject, 0)


def save_lesson_progress(subject, index):
    data = load_streak()
    lesson_progress = data.setdefault("lesson_progress", {})
    lesson_progress[subject] = max(0, min(index, len(DAILY_LESSON_LIBRARY.get(subject, []))))
    save_streak(data)


def lesson_cumulative_days(subject):
    cumulative = [0]
    total = 0
    for lesson in DAILY_LESSON_LIBRARY.get(subject, []):
        total += lesson["days"]
        cumulative.append(total)
    return cumulative


def subject_dates_in_journey(subject):
    dates = []
    current = JOURNEY_START_DATE
    while current <= JOURNEY_LAST_STUDY_DAY:
        if recommended_subject_for_day(current) == subject:
            dates.append(current)
        current += timedelta(days=1)
    return dates


def resolve_lesson_metadata(subject, lesson):
    for item in CURRICULUM.get(subject, []):
        item_text = f"{item.get('topic', '')} {item.get('detail', '')}".lower()
        for match in lesson.get("match", []):
            if match.lower() in item_text:
                return {
                    "detail": item.get("detail", lesson["focus"]),
                    "resource": item.get("resource", SUBJECT_URLS.get(subject, SUBJECT_URLS["General"])),
                    "week": item.get("week", "-"),
                }
    return {
        "detail": lesson["focus"],
        "resource": SUBJECT_URLS.get(subject, SUBJECT_URLS["General"]),
        "week": "-",
    }


def build_daily_schedule():
    calendar = []
    subject_timelines = {}
    for subject, lessons in DAILY_LESSON_LIBRARY.items():
        dates = subject_dates_in_journey(subject)
        timeline = []
        slot = 0
        for lesson_index, lesson in enumerate(lessons):
            if slot >= len(dates):
                break
            topic_start = slot
            finish_slot = min(len(dates) - 1, topic_start + lesson["days"] - 1)
            next_topic = lessons[lesson_index + 1]["topic"] if lesson_index + 1 < len(lessons) else BUFFER_TOPIC
            next_start = dates[topic_start + lesson["days"]] if topic_start + lesson["days"] < len(dates) else None
            for day_in_topic in range(1, lesson["days"] + 1):
                if slot >= len(dates):
                    break
                metadata = resolve_lesson_metadata(subject, lesson)
                timeline.append(
                    {
                        "date": dates[slot],
                        "subject": subject,
                        "topic": lesson["topic"],
                        "topic_index": lesson_index,
                        "day_of_topic": day_in_topic,
                        "topic_days": lesson["days"],
                        "focus": lesson["focus"],
                        "detail": metadata["detail"],
                        "resource": metadata["resource"],
                        "week": metadata["week"],
                        "finish_date": dates[finish_slot],
                        "next_topic": next_topic,
                        "next_topic_start": next_start,
                        "subject_occurrence": slot + 1,
                    }
                )
                slot += 1

        buffer_start = slot
        while slot < len(dates):
            timeline.append(
                {
                    "date": dates[slot],
                    "subject": subject,
                    "topic": BUFFER_TOPIC,
                    "topic_index": len(lessons),
                    "day_of_topic": slot - buffer_start + 1,
                    "topic_days": max(1, len(dates) - slot),
                    "focus": "mixed revision, backlog clearing, and mock practice",
                    "detail": "Use this day for catch-up, weak-topic drilling, or a full mixed test.",
                    "resource": SUBJECT_URLS.get(subject, SUBJECT_URLS["General"]),
                    "week": "-",
                    "finish_date": dates[-1],
                    "next_topic": "Exam Day",
                    "next_topic_start": None,
                    "subject_occurrence": slot + 1,
                }
            )
            slot += 1

        subject_timelines[subject] = timeline

    by_date = {}
    current = JOURNEY_START_DATE
    day_number = 1
    while current <= JOURNEY_LAST_STUDY_DAY:
        subject = recommended_subject_for_day(current)
        entry = next(item for item in subject_timelines[subject] if item["date"] == current)
        item = dict(entry)
        item["day_number"] = day_number
        calendar.append(item)
        by_date[current] = item
        current += timedelta(days=1)
        day_number += 1
    return calendar, by_date, subject_timelines


def daily_schedule_cache():
    if not hasattr(daily_schedule_cache, "_cache"):
        daily_schedule_cache._cache = build_daily_schedule()
    return daily_schedule_cache._cache


def journey_day_number(target=None):
    target = target or date.today()
    if target < JOURNEY_START_DATE:
        return 0
    if target > JOURNEY_LAST_STUDY_DAY:
        return JOURNEY_TOTAL_DAYS
    return (target - JOURNEY_START_DATE).days + 1


def scheduled_lesson_state(subject, target=None):
    target = target or date.today()
    _, _, subject_timelines = daily_schedule_cache()
    timeline = subject_timelines[subject]
    if target < timeline[0]["date"]:
        entry = dict(timeline[0])
        entry["subject_occurrence"] = 0
        entry["day_of_topic"] = 1
        return entry
    eligible = [item for item in timeline if item["date"] <= target]
    return dict(eligible[-1] if eligible else timeline[0])


def actual_lesson_state(subject, target=None):
    target = target or date.today()
    lessons = DAILY_LESSON_LIBRARY.get(subject, [])
    progress_index = get_lesson_progress(subject)
    scheduled = scheduled_lesson_state(subject, target)
    dates = subject_dates_in_journey(subject)
    cumulative = lesson_cumulative_days(subject)

    if progress_index >= len(lessons):
        finish_date = dates[-1] if dates else JOURNEY_LAST_STUDY_DAY
        return {
            "subject": subject,
            "topic": BUFFER_TOPIC,
            "topic_index": progress_index,
            "day_of_topic": 1,
            "topic_days": 1,
            "focus": "mixed revision, backlog clearing, and mock practice",
            "detail": "Use this day for catch-up, weak-topic drilling, or a full mixed test.",
            "resource": SUBJECT_URLS.get(subject, SUBJECT_URLS["General"]),
            "finish_date": finish_date,
            "next_topic": "Exam Day",
            "next_topic_start": None,
            "subject_occurrence": scheduled["subject_occurrence"],
        }

    lesson = lessons[progress_index]
    metadata = resolve_lesson_metadata(subject, lesson)
    slot_number = scheduled["subject_occurrence"]
    actual_day = slot_number - cumulative[progress_index]
    if actual_day < 1:
        actual_day = 1
    if actual_day > lesson["days"]:
        actual_day = lesson["days"]
    finish_slot = min(len(dates), cumulative[progress_index] + lesson["days"])
    finish_date = dates[finish_slot - 1] if finish_slot and finish_slot - 1 < len(dates) else JOURNEY_LAST_STUDY_DAY
    next_topic = lessons[progress_index + 1]["topic"] if progress_index + 1 < len(lessons) else BUFFER_TOPIC
    next_start = dates[finish_slot] if finish_slot < len(dates) else None
    return {
        "subject": subject,
        "topic": lesson["topic"],
        "topic_index": progress_index,
        "day_of_topic": actual_day,
        "topic_days": lesson["days"],
        "focus": lesson["focus"],
        "detail": metadata["detail"],
        "resource": metadata["resource"],
        "finish_date": finish_date,
        "next_topic": next_topic,
        "next_topic_start": next_start,
        "subject_occurrence": slot_number,
    }


def lesson_context(subject, target=None):
    target = target or date.today()
    scheduled = scheduled_lesson_state(subject, target)
    actual = actual_lesson_state(subject, target)
    cumulative = lesson_cumulative_days(subject)

    scheduled_progress = cumulative[min(scheduled["topic_index"], len(cumulative) - 1)] + scheduled["day_of_topic"] - 1
    if actual["topic"] == BUFFER_TOPIC:
        actual_progress = cumulative[-1]
    else:
        actual_progress = cumulative[min(actual["topic_index"], len(cumulative) - 1)] + actual["day_of_topic"] - 1

    behind_days = max(0, scheduled_progress - actual_progress)
    ahead_days = max(0, actual_progress - scheduled_progress)

    if get_lesson_progress(subject) != scheduled["topic_index"]:
        display = actual
    else:
        display = scheduled

    return {
        "scheduled": scheduled,
        "actual": actual,
        "display": display,
        "behind_days": behind_days,
        "ahead_days": ahead_days,
    }


def deadline_label(target_date):
    if not target_date:
        return "Journey end"
    delta = (target_date - date.today()).days
    if delta <= 0:
        return "Today"
    if delta == 1:
        return "Tomorrow"
    if delta == 2:
        return "In 2 days"
    return target_date.strftime("%A, %b %d")


def timer_deadline_label(target_date):
    if not target_date:
        return "finish at journey end"
    delta = (target_date - date.today()).days
    if delta <= 0:
        return "finish today"
    if delta == 1:
        return "finish in 1 day"
    return f"finish in {delta} days"


def catch_up_message(subject, behind_days):
    if behind_days <= 0:
        return ""
    return (
        f"YOU ARE {behind_days} DAYS BEHIND - suggested catch-up plan:\n"
        f"1. Stay on {subject} until this topic is finished.\n"
        "2. Add one extra 20-minute MCQ block tonight.\n"
        "3. Use the next buffer or Sunday revision slot for backlog."
    )


def lesson_tasks(entry):
    focus = trim_text(entry.get("focus", entry["topic"]), 48)
    day_of_topic = entry["day_of_topic"]
    total_days = entry["topic_days"]

    if day_of_topic <= 1:
        tasks = [
            "① Open a fresh notebook page for this topic",
            f"② Watch: {focus} (20 min)",
            "③ Write the core formula/definition list by hand",
            "④ Solve 5 starter questions before ending",
        ]
    elif day_of_topic == 2:
        tasks = [
            "① Re-read yesterday's notes",
            f"② Watch: {focus} (20 min)",
            "③ Solve 10 MCQs on this topic",
            "④ Write 5 key formulas by hand",
        ]
    else:
        tasks = [
            "① Review mistakes from the last session",
            f"② Solve 15 timed MCQs on {entry['topic']}",
            "③ Memorise formulas, exceptions, and shortcuts",
            "④ Summarise the chapter in one notebook page",
        ]

    if day_of_topic >= total_days:
        tasks[-1] = f"④ FINISH TODAY. Tomorrow starts {entry['next_topic']}"
    return tasks


def append_note_entry(session_date, session_time, subject, topic, goal, minutes, mood, notes, difficulty=""):
    entries = load_notes()
    entries.append(
        {
            "date": session_date,
            "time": session_time,
            "subject": subject,
            "topic": topic,
            "goal": goal,
            "minutes": minutes,
            "mood": mood,
            "difficulty": difficulty,
            "notes": notes,
        }
    )
    save_notes(entries)


base.load_streak = load_streak
base.save_streak = save_streak


def get_day_progress():
    data = load_streak()
    return max(0, min(int(data.get("day_progress", 0)), JOURNEY_TOTAL_DAYS - 1))


def save_day_progress(index):
    data = load_streak()
    data["day_progress"] = max(0, min(int(index), JOURNEY_TOTAL_DAYS - 1))
    save_streak(data)


def progress_state():
    return {"day_progress": get_day_progress()}


def get_lesson_progress(subject):
    progress = derive_subject_progress(STUDY_PLAN_DATA["plan"], get_day_progress())
    return progress["subject_days"].get(subject, 0)


def save_lesson_progress(subject, index):
    save_day_progress(index)


def daily_schedule_cache():
    if not hasattr(daily_schedule_cache, "_cache"):
        calendar = []
        by_date = {}
        subject_timelines = {}
        for entry in STUDY_PLAN_DATA["plan"]:
            item = dict(entry)
            item["day_of_topic"] = item["topic_day_index"]
            item["topic_days"] = item["topic_day_total"]
            item["detail"] = item["subtopic"]
            item["resource"] = item["resource_url"]
            calendar.append(item)
            by_date[item["date"]] = item
            subject_timelines.setdefault(item["subject"], []).append(item)
        daily_schedule_cache._cache = (calendar, by_date, subject_timelines)
    return daily_schedule_cache._cache


def journey_day_number(target=None):
    return canonical_journey_day_number(target)


def _decorate_subject_entry(subject, entry):
    subject_plan = [item for item in STUDY_PLAN_DATA["plan"] if item["subject"] == subject]
    position = next((idx for idx, item in enumerate(subject_plan) if item["day_number"] == entry["day_number"]), 0)
    finish_date = entry["date"] + timedelta(days=(entry["topic_day_total"] - entry["topic_day_index"]) * 7)
    next_topic = subject_plan[position + 1]["topic"] if position + 1 < len(subject_plan) else "Exam Day"
    next_topic_start = subject_plan[position + 1]["date"] if position + 1 < len(subject_plan) else None
    item = dict(entry)
    item.update(
        {
            "day_of_topic": item["topic_day_index"],
            "topic_days": item["topic_day_total"],
            "detail": item["subtopic"],
            "resource": item["resource_url"],
            "focus": item["formula_focus"],
            "finish_date": finish_date,
            "next_topic": next_topic,
            "next_topic_start": next_topic_start,
            "subject_occurrence": position + 1,
        }
    )
    return item


def lesson_context(subject, target=None):
    context = canonical_subject_context(subject, progress_state(), target)
    scheduled = _decorate_subject_entry(subject, context["scheduled"])
    actual = _decorate_subject_entry(subject, context["actual"])
    display = _decorate_subject_entry(subject, context["display"])
    return {
        "scheduled": scheduled,
        "actual": actual,
        "display": display,
        "behind_days": context["behind_days"],
        "ahead_days": context["ahead_days"],
        "expected_day": context["expected_day"],
        "day_progress": context["day_progress"],
    }


def catch_up_message(subject, behind_days):
    if behind_days <= 0:
        return ""
    return (
        f"YOU ARE {behind_days} DAYS BEHIND - suggested catch-up plan:\n"
        f"1. Stay on {subject} until this scheduled topic is closed.\n"
        "2. Cut notes to one clean page and spend the saved time on marked MCQs.\n"
        "3. Use Sunday review to clear backlog before touching fresh material."
    )


def lesson_tasks(entry):
    morning = entry.get("morning_plan", [])
    evening = entry.get("evening_plan", [])
    tasks = []
    if morning:
        tasks.append(f"1. Morning: {morning[1]}")
        tasks.append(f"2. Morning drill: {morning[2]}")
    if evening:
        tasks.append(f"3. Evening: {evening[1]}")
        tasks.append(f"4. Close: {evening[-1]}")
    return tasks


def sitting_finish_line(subject, entry, routine=None, now=None):
    now = now or datetime.now()
    is_morning = now.hour < 15
    session_name = "MORNING 50-MIN LOCK-IN" if is_morning else "EVENING 90-MIN LOCK-IN"
    mcq_target = int(entry.get("mcq_target_morning" if is_morning else "mcq_target_evening", 0))
    if mcq_target <= 0:
        mcq_target = 5 if subject == "ENGLISH" else 8
    day_of_topic = int(entry.get("day_of_topic", entry.get("topic_day_index", 1)))
    total_days = int(entry.get("topic_days", entry.get("topic_day_total", 1)))
    topic = entry.get("topic", subject)
    subtopic = entry.get("subtopic", entry.get("detail", topic))
    formula_focus = entry.get("formula_focus", entry.get("focus", "today's key formulas and rules"))
    checkpoint = entry.get("self_test_question", f"Can you explain {subtopic} without looking at notes?")
    routine = routine or {}

    if day_of_topic >= total_days:
        finish_line = f"Close {topic} today. No dragging this topic into the next study day."
        minimum_win = f"Finish the final subtopic, solve {mcq_target} MCQs, and answer the self-test aloud."
        stretch_win = f"Add 5 extra MCQs and write a one-page final summary for {topic}."
        pressure = "This is a completion sitting. Slow is okay. Leaving the topic half-open is not."
    else:
        finish_line = f"Complete today's slice of {topic}: {subtopic}."
        minimum_win = f"Understand {subtopic}, solve {mcq_target} MCQs, and write the key formulas before stopping."
        stretch_win = f"Preview the next slice for 10 minutes so tomorrow starts easier."
        pressure = "Do not chase perfect notes. The win is a finished slice, checked by questions."

    notebook_lines = [
        f"3 formulas/rules from: {formula_focus}",
        "1 mistake or trap you noticed while solving",
        "1 sentence explaining the idea in your own words",
    ]
    if routine and routine.get("review_target"):
        notebook_lines.append(f"Quick repair note for: {routine['review_target']}")

    return {
        "session_name": session_name,
        "finish_line": finish_line,
        "minimum_win": minimum_win,
        "stretch_win": stretch_win,
        "pressure": pressure,
        "mcq_target": mcq_target,
        "notebook_lines": notebook_lines,
        "checkpoint": checkpoint,
        "saved_goal": f"Finish: {subtopic} | {mcq_target} MCQs | {topic}",
    }


def week_bounds(today=None):
    today = today or date.today()
    start = today - timedelta(days=today.weekday())
    end = start + timedelta(days=6)
    return start, end


def compute_weekly_summary(data, today=None):
    today = today or date.today()
    start, end = week_bounds(today)
    current_entry = canonical_current_day_entry(today)
    weekly_target = weekly_target_for_day(current_entry)
    sessions = []
    for session in data.get("sessions", []):
        session_date = parse_iso_date(session.get("date", ""))
        if session_date and start <= session_date <= end:
            sessions.append(session)
    total_minutes = sum(int(session.get("minutes", 0)) for session in sessions)
    subjects = sorted({session.get("subject", "General") for session in sessions if session.get("subject")})
    studied_days = len({session.get("date", "") for session in sessions if session.get("date")})
    grade = "S" if studied_days >= 5 else "A" if studied_days == 4 else "B" if studied_days == 3 else "C"
    return {
        "start": start,
        "end": end,
        "total_minutes": total_minutes,
        "subjects": subjects,
        "studied_days": studied_days,
        "grade": grade,
        "message": GRADE_MESSAGES.get(grade, ""),
        "streak": data.get("streak", 0),
        "weekly_target": weekly_target,
    }


def compute_exam_projection(data):
    total_hours = round(data.get("total_minutes", 0) / 60, 1)
    remaining_hours = max(0.0, round(TOTAL_TARGET_HOURS - total_hours, 1))
    days_left = max(0, (IOE_DATE - date.today()).days)
    session_dates = [
        parse_iso_date(session.get("date", ""))
        for session in data.get("sessions", [])
        if session.get("date")
    ]
    session_dates = [session_date for session_date in session_dates if session_date]
    if session_dates:
        elapsed_days = max(1, (date.today() - min(session_dates)).days + 1)
    else:
        elapsed_days = 1
    avg_daily_hours = total_hours / elapsed_days if total_hours > 0 else 0.0
    projected_total = round(total_hours + avg_daily_hours * days_left, 1)
    return {
        "days_left": days_left,
        "total_hours": total_hours,
        "remaining_hours": remaining_hours,
        "avg_daily_hours": avg_daily_hours,
        "projected_total": projected_total,
    }


def notify_excel_issue(message):
    notification.notify(title="Study report", message=message, timeout=8)


def export_to_excel(data=None):
    global excel_error_notified
    data = data or load_streak()
    notes_data = load_notes()
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        if not excel_error_notified:
            excel_error_notified = True
            notify_excel_issue("Install openpyxl to generate study_tracker.xlsx.")
        return False

    BX = "050608"
    CY = "00FFE7"
    GR = "00C878"
    GO = "FFD700"
    RE = "FF4444"
    WH = "E8F0EE"
    DM = "0D1810"
    BL = "58B7FF"

    def style_cell(ws, row, col, value, bold=False, size=10, font_color=WH, bg=BX, align="center", wrap=False):
        cell = ws.cell(row, col, value=value)
        cell.font = Font(name=FONT_FACE, bold=bold, size=size, color=font_color)
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
        cell.fill = PatternFill("solid", start_color=bg)
        return cell

    try:
        workbook = Workbook()
        sessions = data.get("sessions", [])

        all_ws = workbook.active
        all_ws.title = "📘 All Sessions"
        all_ws.sheet_view.showGridLines = False
        all_ws.sheet_properties.tabColor = CY
        for col, width in enumerate([2, 14, 10, 10, 16, 28, 26, 14, 14, 12], start=1):
            all_ws.column_dimensions[get_column_letter(col)].width = width
        for row in range(1, 420):
            all_ws.row_dimensions[row].height = 24
            for col in range(1, 11):
                all_ws.cell(row, col).fill = PatternFill("solid", start_color=BX)

        all_ws.merge_cells("B2:J2")
        style_cell(all_ws, 2, 2, "KRISH BARAL - STUDY LOG", bold=True, size=16, font_color=CY, bg=BX)
        all_ws["B2"].alignment = Alignment(horizontal="center", vertical="center")

        headers = ["DATE", "TIME", "MINUTES", "SUBJECT", "TOPIC", "GOAL", "MOOD", "DIFFICULTY", "POMODOROS"]
        header_colors = [WH, WH, CY, BL, GR, GO, RE, GO, CY]
        for index, header in enumerate(headers, start=2):
            style_cell(all_ws, 4, index, header, bold=True, size=10, font_color=header_colors[index - 2], bg=DM)

        mood_colors = {"crushed": GR, "okay": GO, "struggled": RE}
        for row, session in enumerate(sessions, start=5):
            row_bg = "0A100C" if row % 2 else BX
            style_cell(all_ws, row, 2, session.get("date", ""), size=10, bg=row_bg)
            style_cell(all_ws, row, 3, session.get("time", ""), size=10, bg=row_bg)
            style_cell(all_ws, row, 4, session.get("minutes", 0), bold=True, size=12, font_color=CY, bg=row_bg)
            style_cell(all_ws, row, 5, session.get("subject", ""), bold=True, size=10, font_color=BL, bg=row_bg)
            style_cell(all_ws, row, 6, session.get("topic", ""), size=10, font_color=WH, bg=row_bg, align="left", wrap=True)
            style_cell(all_ws, row, 7, session.get("goal", ""), size=10, font_color=GO, bg=row_bg, align="left", wrap=True)
            style_cell(all_ws, row, 8, str(session.get("mood", "")).upper(), bold=True, size=10, font_color=mood_colors.get(session.get("mood", ""), WH), bg=row_bg)
            difficulty = str(session.get("difficulty", "")).upper()
            difficulty_color = RE if difficulty == "HARD" else GO if difficulty == "MEDIUM" else GR if difficulty == "EASY" else WH
            style_cell(all_ws, row, 9, difficulty, bold=True, size=10, font_color=difficulty_color, bg=row_bg)
            style_cell(all_ws, row, 10, max(0, int(session.get("minutes", 0)) // 25), bold=True, size=10, font_color=GR, bg=row_bg)
            all_ws.row_dimensions[row].height = 32

        total_row = 5 + len(sessions) + 1
        all_ws.merge_cells(start_row=total_row, start_column=2, end_row=total_row, end_column=3)
        style_cell(all_ws, total_row, 2, "TOTAL", bold=True, size=11, font_color=CY, bg=DM)
        style_cell(all_ws, total_row, 4, f"=SUM(D5:D{4 + len(sessions)})", bold=True, size=12, font_color=CY, bg=DM)

        stats_ws = workbook.create_sheet("📊 Stats")
        stats_ws.sheet_view.showGridLines = False
        stats_ws.sheet_properties.tabColor = GR
        for col, width in enumerate([2, 24, 18, 18, 18, 2], start=1):
            stats_ws.column_dimensions[get_column_letter(col)].width = width
        for row in range(1, 70):
            stats_ws.row_dimensions[row].height = 24
            for col in range(1, 7):
                stats_ws.cell(row, col).fill = PatternFill("solid", start_color=BX)

        stats_ws.merge_cells("B2:E2")
        style_cell(stats_ws, 2, 2, "STUDY STATS", bold=True, size=16, font_color=CY, bg=BX)
        stats_ws["B2"].alignment = Alignment(horizontal="center", vertical="center")

        projection = compute_exam_projection(data)
        stats = [
            ("STREAK", f"{data.get('streak', 0)} days", GO),
            ("SESSIONS", str(data.get("total_sessions", 0)), CY),
            ("TOTAL MINS", str(data.get("total_minutes", 0)), GR),
            ("TOTAL HOURS", f"{projection['total_hours']}h", BL),
            ("HOURS LEFT", f"{projection['remaining_hours']}h", GO),
            ("PROJECTED", f"{projection['projected_total']}h by exam", CY),
        ]
        row = 4
        for label, value, color in stats:
            style_cell(stats_ws, row, 2, label, bold=True, size=9, font_color="70857B", bg=DM)
            style_cell(stats_ws, row + 1, 2, value, bold=True, size=14, font_color=color, bg=DM)
            row += 3

        style_cell(stats_ws, 24, 2, "SUBJECT BREAKDOWN", bold=True, size=12, font_color=CY, bg=DM)
        stats_ws.merge_cells("B24:E24")
        stats_ws["B24"].alignment = Alignment(horizontal="center", vertical="center")
        subject_minutes = {}
        for session in sessions:
            subject = session.get("subject", "General")
            subject_minutes[subject] = subject_minutes.get(subject, 0) + int(session.get("minutes", 0))
        row = 26
        for subject, minutes in sorted(subject_minutes.items(), key=lambda item: -item[1]):
            style_cell(stats_ws, row, 2, subject, bold=True, size=10, font_color=subject_hex(subject).replace("#", ""), bg=BX)
            style_cell(stats_ws, row, 3, f"{minutes} min", bold=True, size=10, font_color=WH, bg=BX)
            style_cell(stats_ws, row, 4, f"{round(minutes / 60, 1)}h", size=10, font_color="90C7B8", bg=BX)
            row += 2

        topic_ws = workbook.create_sheet("📚 Topic Progress")
        topic_ws.sheet_view.showGridLines = False
        topic_ws.sheet_properties.tabColor = CY
        for col, width in enumerate([2, 18, 34, 50, 16], start=1):
            topic_ws.column_dimensions[get_column_letter(col)].width = width
        for row in range(1, 260):
            topic_ws.row_dimensions[row].height = 24
            for col in range(1, 6):
                topic_ws.cell(row, col).fill = PatternFill("solid", start_color=BX)

        topic_ws.merge_cells("B2:D2")
        style_cell(topic_ws, 2, 2, "IOE TOPIC PROGRESS - KRISH", bold=True, size=14, font_color=CY, bg=BX)
        topic_ws["B2"].alignment = Alignment(horizontal="center", vertical="center")
        row = 4
        subject_progress = derive_subject_progress(STUDY_PLAN_DATA["plan"], get_day_progress())
        for subject in ("MATHEMATICS", "PHYSICS", "CHEMISTRY", "ENGLISH"):
            color = subject_hex(subject).replace("#", "")
            subject_entries = [entry for entry in STUDY_PLAN_DATA["plan"] if entry["subject"] == subject]
            topic_names = []
            for entry in subject_entries:
                if entry["topic"] not in topic_names:
                    topic_names.append(entry["topic"])
            done_topics = subject_progress["subject_topics"].get(subject, [])
            topic_ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
            style_cell(topic_ws, row, 2, f"{subject} ({len(done_topics)}/{len(topic_names)} done)", bold=True, size=11, font_color=color, bg=DM, align="left")
            row += 1
            current_topic = next((entry["topic"] for entry in subject_entries if entry["day_number"] > get_day_progress()), topic_names[-1])
            for index, topic_name in enumerate(topic_names, start=1):
                topic_entry = next(entry for entry in subject_entries if entry["topic"] == topic_name)
                status = "DONE" if topic_name in done_topics else "CURRENT" if topic_name == current_topic else "UPCOMING"
                row_bg = "0A100C" if status == "DONE" else BX
                accent = GR if status == "DONE" else color if status == "CURRENT" else "3A544A"
                style_cell(topic_ws, row, 2, topic_entry["phase"], bold=True, size=10, font_color=accent, bg=row_bg)
                style_cell(topic_ws, row, 3, topic_name, size=10, font_color=WH, bg=row_bg, align="left", wrap=True)
                style_cell(topic_ws, row, 4, topic_entry["formula_focus"], size=9, font_color="90C7B8", bg=row_bg, align="left", wrap=True)
                style_cell(topic_ws, row, 5, status, bold=True, size=9, font_color=accent, bg=row_bg)
                topic_ws.row_dimensions[row].height = 34
                row += 1
            row += 1

        notes_ws = workbook.create_sheet("📝 Notes")
        notes_ws.sheet_view.showGridLines = False
        notes_ws.sheet_properties.tabColor = GO
        for col, width in enumerate([2, 14, 10, 16, 24, 26, 16, 62], start=1):
            notes_ws.column_dimensions[get_column_letter(col)].width = width
        for row in range(1, 320):
            notes_ws.row_dimensions[row].height = 24
            for col in range(1, 9):
                notes_ws.cell(row, col).fill = PatternFill("solid", start_color=BX)

        notes_ws.merge_cells("B2:H2")
        style_cell(notes_ws, 2, 2, "SESSION NOTES", bold=True, size=16, font_color=GO, bg=BX)
        notes_ws["B2"].alignment = Alignment(horizontal="center", vertical="center")
        note_headers = ["DATE", "TIME", "SUBJECT", "TOPIC", "GOAL", "DIFFICULTY", "NOTES"]
        for index, header in enumerate(note_headers, start=2):
            style_cell(notes_ws, 4, index, header, bold=True, size=10, font_color=GO if index >= 6 else CY, bg=DM)
        for row, note in enumerate(notes_data, start=5):
            row_bg = "0A100C" if row % 2 else BX
            style_cell(notes_ws, row, 2, note.get("date", ""), size=10, bg=row_bg)
            style_cell(notes_ws, row, 3, note.get("time", ""), size=10, bg=row_bg)
            style_cell(notes_ws, row, 4, note.get("subject", ""), bold=True, size=10, font_color=subject_hex(note.get("subject", "General")).replace("#", ""), bg=row_bg)
            style_cell(notes_ws, row, 5, note.get("topic", ""), size=10, font_color=WH, bg=row_bg, align="left", wrap=True)
            style_cell(notes_ws, row, 6, note.get("goal", ""), size=10, font_color=GO, bg=row_bg, align="left", wrap=True)
            style_cell(notes_ws, row, 7, str(note.get("difficulty", "")).upper(), bold=True, size=10, font_color=GO, bg=row_bg)
            style_cell(notes_ws, row, 8, note.get("notes", ""), size=10, font_color=WH, bg=row_bg, align="left", wrap=True)
            notes_ws.row_dimensions[row].height = 52

        schedule_ws = workbook.create_sheet("📅 Daily Schedule")
        schedule_ws.sheet_view.showGridLines = False
        schedule_ws.sheet_properties.tabColor = GO
        for col, width in enumerate([2, 14, 10, 16, 30, 16, 16], start=1):
            schedule_ws.column_dimensions[get_column_letter(col)].width = width
        for row in range(1, 260):
            schedule_ws.row_dimensions[row].height = 24
            for col in range(1, 8):
                schedule_ws.cell(row, col).fill = PatternFill("solid", start_color=BX)

        schedule_ws.merge_cells("B2:G2")
        style_cell(schedule_ws, 2, 2, "180-DAY DAILY SCHEDULE", bold=True, size=16, font_color=GO, bg=BX)
        schedule_ws["B2"].alignment = Alignment(horizontal="center", vertical="center")
        headers = ["DATE", "DAY#", "SUBJECT", "TOPIC", "DAY-OF-TOPIC", "STATUS"]
        for index, header in enumerate(headers, start=2):
            style_cell(schedule_ws, 4, index, header, bold=True, size=10, font_color=CY if index < 6 else GO, bg=DM)

        calendar, _, _ = daily_schedule_cache()
        today = date.today()
        for row, entry in enumerate(calendar, start=5):
            if entry["date"] < today:
                row_bg = "0A2414"
                status = "Done"
                status_color = GR
            elif entry["date"] == today:
                row_bg = "3A2C08"
                status = "Current"
                status_color = GO
            else:
                row_bg = BX
                status = "Upcoming"
                status_color = CY

            style_cell(schedule_ws, row, 2, str(entry["date"]), size=10, bg=row_bg)
            style_cell(schedule_ws, row, 3, entry["day_number"], bold=True, size=10, font_color=GO if entry["date"] == today else CY, bg=row_bg)
            style_cell(schedule_ws, row, 4, entry["subject"], bold=True, size=10, font_color=subject_hex(entry["subject"]).replace("#", ""), bg=row_bg)
            style_cell(schedule_ws, row, 5, entry["topic"], size=10, font_color=WH, bg=row_bg, align="left", wrap=True)
            style_cell(schedule_ws, row, 6, f"{entry['day_of_topic']} / {entry['topic_days']}", bold=True, size=10, font_color=WH, bg=row_bg)
            style_cell(schedule_ws, row, 7, status, bold=True, size=10, font_color=status_color, bg=row_bg)
            schedule_ws.row_dimensions[row].height = 30

        countdown_ws = workbook.create_sheet("🎯 IOE Countdown")
        weak_ws = workbook.create_sheet("Weak Topics")
        weak_ws.sheet_view.showGridLines = False
        weak_ws.sheet_properties.tabColor = GO
        for col, width in enumerate([2, 16, 18, 34, 16, 14, 14, 2], start=1):
            weak_ws.column_dimensions[get_column_letter(col)].width = width
        for row in range(1, 80):
            weak_ws.row_dimensions[row].height = 24
            for col in range(1, 9):
                weak_ws.cell(row, col).fill = PatternFill("solid", start_color=BX)

        weak_ws.merge_cells("B2:G2")
        style_cell(weak_ws, 2, 2, "WEAK TOPIC REVISION QUEUE", bold=True, size=16, font_color=GO, bg=BX)
        weak_ws["B2"].alignment = Alignment(horizontal="center", vertical="center")
        weak_headers = ["LAST SEEN", "SUBJECT", "TOPIC", "DIFFICULTY", "HARD HITS", "STATUS"]
        for index, header in enumerate(weak_headers, start=2):
            style_cell(weak_ws, 4, index, header, bold=True, size=10, font_color=GO if index >= 5 else CY, bg=DM)
        for row, item in enumerate(data.get("weak_topics", []), start=5):
            row_bg = "0A100C" if row % 2 else BX
            difficulty = str(item.get("last_difficulty", "")).upper()
            difficulty_color = RE if difficulty == "HARD" else GO if difficulty == "MEDIUM" else GR if difficulty == "EASY" else WH
            style_cell(weak_ws, row, 2, item.get("last_seen", ""), size=10, bg=row_bg)
            style_cell(weak_ws, row, 3, item.get("subject", ""), bold=True, size=10, font_color=subject_hex(item.get("subject", "General")).replace("#", ""), bg=row_bg)
            style_cell(weak_ws, row, 4, item.get("topic", ""), size=10, font_color=WH, bg=row_bg, align="left", wrap=True)
            style_cell(weak_ws, row, 5, difficulty, bold=True, size=10, font_color=difficulty_color, bg=row_bg)
            style_cell(weak_ws, row, 6, item.get("hard_count", 0), bold=True, size=10, font_color=RE, bg=row_bg)
            style_cell(weak_ws, row, 7, str(item.get("status", "")).upper(), bold=True, size=10, font_color=CY, bg=row_bg)
            weak_ws.row_dimensions[row].height = 32

        mock_ws = workbook.create_sheet("Mock Exams")
        mock_ws.sheet_view.showGridLines = False
        mock_ws.sheet_properties.tabColor = RE
        for col, width in enumerate([2, 14, 10, 10, 14, 12, 14, 12, 12, 12, 2], start=1):
            mock_ws.column_dimensions[get_column_letter(col)].width = width
        for row in range(1, 120):
            mock_ws.row_dimensions[row].height = 24
            for col in range(1, 12):
                mock_ws.cell(row, col).fill = PatternFill("solid", start_color=BX)

        mock_ws.merge_cells("B2:J2")
        style_cell(mock_ws, 2, 2, "IOE MOCK EXAM HISTORY", bold=True, size=16, font_color=RE, bg=BX)
        mock_ws["B2"].alignment = Alignment(horizontal="center", vertical="center")
        mock_headers = ["DATE", "TIME", "MIN", "MATH", "PHYSICS", "CHEM", "ENGLISH", "TOTAL", "PERCENT"]
        for index, header in enumerate(mock_headers, start=2):
            style_cell(mock_ws, 4, index, header, bold=True, size=10, font_color=CY if index < 9 else GO, bg=DM)
        for row, mock in enumerate(data.get("mock_exams", []), start=5):
            scores = mock.get("scores", {})
            row_bg = "0A100C" if row % 2 else BX
            style_cell(mock_ws, row, 2, mock.get("date", ""), size=10, bg=row_bg)
            style_cell(mock_ws, row, 3, mock.get("time", ""), size=10, bg=row_bg)
            style_cell(mock_ws, row, 4, mock.get("minutes", 0), bold=True, size=10, font_color=CY, bg=row_bg)
            style_cell(mock_ws, row, 5, scores.get("MATHEMATICS", 0), bold=True, size=10, font_color=subject_hex("MATHEMATICS").replace("#", ""), bg=row_bg)
            style_cell(mock_ws, row, 6, scores.get("PHYSICS", 0), bold=True, size=10, font_color=subject_hex("PHYSICS").replace("#", ""), bg=row_bg)
            style_cell(mock_ws, row, 7, scores.get("CHEMISTRY", 0), bold=True, size=10, font_color=subject_hex("CHEMISTRY").replace("#", ""), bg=row_bg)
            style_cell(mock_ws, row, 8, scores.get("ENGLISH", 0), bold=True, size=10, font_color=subject_hex("ENGLISH").replace("#", ""), bg=row_bg)
            style_cell(mock_ws, row, 9, f"{mock.get('total', 0)} / {mock.get('max_total', 140)}", bold=True, size=10, font_color=GO, bg=row_bg)
            style_cell(mock_ws, row, 10, f"{mock.get('percent', 0)}%", bold=True, size=10, font_color=GR if mock.get("percent", 0) >= 70 else GO, bg=row_bg)

        countdown_ws.sheet_view.showGridLines = False
        countdown_ws.sheet_properties.tabColor = RE
        for col, width in enumerate([2, 24, 22, 22, 2], start=1):
            countdown_ws.column_dimensions[get_column_letter(col)].width = width
        for row in range(1, 36):
            countdown_ws.row_dimensions[row].height = 26
            for col in range(1, 6):
                countdown_ws.cell(row, col).fill = PatternFill("solid", start_color=BX)

        countdown_ws.merge_cells("B2:D2")
        style_cell(countdown_ws, 2, 2, "IOE ENTRANCE EXAM", bold=True, size=18, font_color=CY, bg=BX)
        countdown_ws["B2"].alignment = Alignment(horizontal="center", vertical="center")
        days_left = projection["days_left"]
        day_color = CY if days_left > 60 else GO if days_left > 30 else RE
        countdown_ws.merge_cells("B4:D4")
        style_cell(countdown_ws, 4, 2, f"{days_left} DAYS REMAINING", bold=True, size=24, font_color=day_color, bg=BX)
        countdown_ws["B4"].alignment = Alignment(horizontal="center", vertical="center")
        countdown_rows = [
            ("Exam Date", str(IOE_DATE), GO),
            ("Hours Studied", f"{projection['total_hours']}h", CY),
            ("Hours Left", f"{projection['remaining_hours']}h", GO),
            ("Current Pace", f"{projection['avg_daily_hours']:.2f}h/day", GR),
            ("Projected Total", f"{projection['projected_total']}h", CY),
            ("Current Streak", f"{data.get('streak', 0)} days", GO),
        ]
        row = 8
        for label, value, color in countdown_rows:
            style_cell(countdown_ws, row, 2, label, bold=True, size=10, font_color="70857B", bg=DM)
            style_cell(countdown_ws, row, 3, value, bold=True, size=12, font_color=color, bg=DM)
            row += 2

        workbook.save(EXCEL_FILE)
        return True
    except PermissionError:
        notify_excel_issue("Close study_tracker.xlsx, then the report can save normally.")
        return False
    except Exception as exc:
        notify_excel_issue(f"Excel export failed: {exc}")
        return False


base.export_to_excel = export_to_excel


def update_streak(minutes, mood, subject, goal="", topic="", notes="", difficulty=""):
    data = load_streak()
    today = str(date.today())
    yesterday = str(date.today() - timedelta(days=1))
    last = data.get("last_date", "")
    if last == today:
        pass
    elif last == yesterday:
        data["streak"] = data.get("streak", 0) + 1
    else:
        data["streak"] = 1

    session_time = datetime.now().strftime("%H:%M")
    data["last_date"] = today
    data["last_subject"] = subject
    data["total_sessions"] = data.get("total_sessions", 0) + 1
    data["total_minutes"] = data.get("total_minutes", 0) + minutes
    data.setdefault("sessions", []).append(
        {
            "date": today,
            "time": session_time,
            "minutes": minutes,
            "subject": subject,
            "topic": topic,
            "goal": goal,
            "mood": mood,
            "difficulty": difficulty,
        }
    )
    record_topic_difficulty(data, subject, topic, difficulty)
    save_streak(data)
    append_note_entry(today, session_time, subject, topic, goal, minutes, mood, notes, difficulty)
    threading.Thread(target=lambda: export_to_excel(data), daemon=True).start()
    return data


def get_effective_study_seconds():
    focus_seconds = pomodoro_state.focus_seconds()
    if focus_seconds > 0:
        return focus_seconds
    if study_start_time:
        return max(0, int(time.time() - study_start_time))
    return 0


def show_windows_toast(title, message, sound_uri=WINDOWS_TOAST_SOUND):
    def _send():
        try:
            toast_xml = f"""
<toast duration='long'>
  <visual>
    <binding template='ToastGeneric'>
      <text>{xml_escape(title)}</text>
      <text>{xml_escape(message)}</text>
    </binding>
  </visual>
  <audio src='{xml_escape(sound_uri)}'/>
</toast>
""".strip()
            toast_json = json.dumps(toast_xml)
            script = f"""
[Windows.UI.Notifications.ToastNotificationManager, Windows.UI.Notifications, ContentType = WindowsRuntime] > $null
[Windows.Data.Xml.Dom.XmlDocument, Windows.Data.Xml.Dom.XmlDocument, ContentType = WindowsRuntime] > $null
$xml = New-Object Windows.Data.Xml.Dom.XmlDocument
$xml.LoadXml({toast_json})
$toast = [Windows.UI.Notifications.ToastNotification]::new($xml)
$notifier = [Windows.UI.Notifications.ToastNotificationManager]::CreateToastNotifier('StudyModeKrish')
$notifier.Show($toast)
"""
            subprocess.run(
                ["powershell", "-ExecutionPolicy", "Bypass", "-WindowStyle", "Hidden", "-Command", script],
                capture_output=True,
                timeout=10,
            )
        except Exception:
            notification.notify(title=title, message=message, timeout=8)

    threading.Thread(target=_send, daemon=True).start()


def add_corner_brackets(container, accent):
    width = max(container.winfo_width(), container.winfo_reqwidth())
    height = max(container.winfo_height(), container.winfo_reqheight())
    pieces = [
        (8, 8, 28, 3),
        (8, 8, 3, 28),
        (width - 36, 8, 28, 3),
        (width - 11, 8, 3, 28),
        (8, height - 11, 28, 3),
        (8, height - 36, 3, 28),
        (width - 36, height - 11, 28, 3),
        (width - 11, height - 36, 3, 28),
    ]
    for x, y, width, height in pieces:
        tk.Frame(container, bg=accent).place(x=x, y=y, width=width, height=height)


def make_readable_popup(title, width, height, accent, subtitle=None):
    root = tk.Toplevel(_tk_root)
    root.overrideredirect(True)
    root.attributes("-topmost", True)
    root.configure(bg=BG)
    sw = root.winfo_screenwidth()
    sh = root.winfo_screenheight()
    root.geometry(f"{width}x{height}+{(sw - width) // 2}+{(sh - height) // 2}")

    glow = tk.Frame(root, bg=accent)
    glow.place(x=0, y=0, width=width, height=height)
    shell = tk.Frame(root, bg=BG)
    shell.place(x=4, y=4, width=width - 8, height=height - 8)
    panel = tk.Frame(root, bg=POPUP_BG, highlightbackground=accent, highlightthickness=2)
    panel.place(x=10, y=10, width=width - 20, height=height - 20)
    panel.update_idletasks()
    add_corner_brackets(panel, accent)

    tk.Label(
        panel,
        text=title,
        font=(FONT_FACE, POPUP_HEADER_PT, "bold"),
        fg=accent,
        bg=POPUP_BG,
    ).pack(pady=(18, 6))
    if subtitle:
        tk.Label(
            panel,
            text=subtitle,
            font=(FONT_FACE, POPUP_BODY_PT),
            fg=POPUP_MUTED,
            bg=POPUP_BG,
            justify="center",
        ).pack(pady=(0, 10))
    tk.Frame(panel, bg=accent, height=2).pack(fill="x", padx=28, pady=(0, 14))
    root.bind("<Escape>", lambda event: root.destroy())
    root.grab_set()
    root.focus_force()
    return root, panel


def play_habit_sound(name):
    pattern = HABIT_SOUND_PATTERNS.get(name)
    if not pattern:
        return
    mode = get_sound_mode()
    if mode == "off":
        return
    if mode == "soft":
        pattern = [(frequency, max(35, int(duration * 0.65))) for frequency, duration in pattern[:2]]

    def _play():
        try:
            if winsound is not None:
                for frequency, duration in pattern:
                    winsound.Beep(int(frequency), int(duration))
                    time.sleep(0.03)
        except Exception:
            try:
                for _frequency, _duration in pattern:
                    notification.notify(title="Study Mode", message="", timeout=1)
                    break
            except Exception:
                pass

    threading.Thread(target=_play, daemon=True).start()


def popup_section(parent, title, accent):
    section = tk.Frame(parent, bg=POPUP_PANEL, highlightbackground=accent, highlightthickness=1)
    section.pack(fill="x", padx=28, pady=(0, 16))
    tk.Label(
        section,
        text=title,
        font=(FONT_FACE, 12, "bold"),
        fg=accent,
        bg=POPUP_PANEL,
        anchor="w",
    ).pack(fill="x", padx=16, pady=(12, 6))
    return section


def popup_button(parent, text, fg, command):
    return tk.Button(
        parent,
        text=text,
        font=(FONT_FACE, 13, "bold"),
        bg=BUTTON_BG,
        fg=fg,
        activebackground=BUTTON_ACTIVE,
        activeforeground=fg,
        relief="flat",
        bd=0,
        cursor="hand2",
        command=command,
        highlightbackground=fg,
        highlightthickness=2,
        padx=14,
        pady=10,
    )


def show_session_goal_popup(subject, topic):
    accent = subject_hex(subject)
    result = {"goal": None}
    root, panel = make_readable_popup(
        "SESSION GOAL",
        760,
        390,
        accent,
        subtitle=f"{subject} | {trim_text(topic, 60)}",
    )

    tk.Label(
        panel,
        text="What is your goal for this session?",
        font=(FONT_FACE, 16, "bold"),
        fg=POPUP_TEXT,
        bg=POPUP_BG,
    ).pack(pady=(10, 14))

    tk.Label(
        panel,
        text="Pick the feeling you want from this block. Keep it simple and start fast.",
        font=(FONT_FACE, 11),
        fg=POPUP_MUTED,
        bg=POPUP_BG,
        wraplength=660,
        justify="center",
    ).pack(pady=(0, 12))

    buttons = tk.Frame(panel, bg=POPUP_BG)
    buttons.pack(fill="both", expand=True, padx=34, pady=(0, 12))
    for col in range(2):
        buttons.grid_columnconfigure(col, weight=1)

    def choose_goal(chosen):
        play_habit_sound("goal_pick")
        result["goal"] = chosen
        root.after(120, root.destroy)

    for index, goal in enumerate(SESSION_GOALS):
        button = popup_button(buttons, goal, accent, lambda chosen=goal: choose_goal(chosen))
        button.configure(font=(FONT_FACE, 12, "bold"), pady=14)
        button.grid(
            row=index // 2,
            column=index % 2,
            sticky="nsew",
            padx=10,
            pady=10,
        )

    tk.Label(
        panel,
        text="Press ESC if you want to cancel starting this session.",
        font=(FONT_FACE, POPUP_BODY_PT),
        fg=POPUP_DIM,
        bg=POPUP_BG,
    ).pack(pady=(0, 14))
    root.wait_window()
    return result["goal"]


def show_notes_popup(minutes, subject, topic, goal, callback):
    accent = subject_hex(subject)
    completed = {"done": False}
    root, panel = make_readable_popup("SESSION NOTES", 820, 520, accent)

    tk.Label(
        panel,
        text=f"{minutes} minutes | {subject} | Goal: {goal or 'Not set'}",
        font=(FONT_FACE, 14, "bold"),
        fg=POPUP_TEXT,
        bg=POPUP_BG,
    ).pack(pady=(10, 6))
    tk.Label(
        panel,
        text=trim_text(topic or subject, 80),
        font=(FONT_FACE, POPUP_BODY_PT),
        fg=POPUP_MUTED,
        bg=POPUP_BG,
    ).pack(pady=(0, 10))

    section = popup_section(panel, "WHAT DID YOU LEARN TODAY? (OPTIONAL)", accent)
    text_box = tk.Text(
        section,
        height=10,
        font=(FONT_FACE, POPUP_BODY_PT),
        bg=BUTTON_BG,
        fg=POPUP_TEXT,
        insertbackground=CYAN,
        wrap="word",
        relief="flat",
        padx=12,
        pady=12,
    )
    text_box.pack(fill="both", expand=True, padx=16, pady=(0, 16))
    text_box.focus_set()

    button_row = tk.Frame(panel, bg=POPUP_BG)
    button_row.pack(fill="x", padx=28, pady=(0, 20))

    def finish(text):
        completed["done"] = True
        root.destroy()
        callback(text.strip())

    popup_button(button_row, "SAVE SESSION", accent, lambda: finish(text_box.get("1.0", "end"))).pack(side="left", expand=True, fill="x", padx=(0, 8))
    popup_button(button_row, "SKIP NOTE", CYAN, lambda: finish("")).pack(side="left", expand=True, fill="x", padx=(8, 0))
    root.wait_window()
    if not completed["done"]:
        callback("")


def show_topic_difficulty_popup(subject, topic, callback):
    accent = subject_hex(subject)
    completed = {"done": False}
    root, panel = make_readable_popup(
        "TOPIC CHECK",
        760,
        360,
        accent,
        subtitle=f"{subject} | {trim_text(topic, 60)}",
    )
    tk.Label(
        panel,
        text="How clear did this topic feel?",
        font=(FONT_FACE, 16, "bold"),
        fg=POPUP_TEXT,
        bg=POPUP_BG,
    ).pack(pady=(14, 8))
    tk.Label(
        panel,
        text="Hard topics get pulled into weekly revision automatically.",
        font=(FONT_FACE, 11),
        fg=POPUP_MUTED,
        bg=POPUP_BG,
        wraplength=660,
        justify="center",
    ).pack(pady=(0, 20))

    choices = tk.Frame(panel, bg=POPUP_BG)
    choices.pack(fill="x", padx=28, pady=(0, 18))
    for col in range(3):
        choices.grid_columnconfigure(col, weight=1)

    def choose(difficulty):
        completed["done"] = True
        play_habit_sound("goal_pick")
        root.destroy()
        callback(difficulty)

    for col, (label, color, difficulty) in enumerate(
        [
            ("EASY", GREEN, "easy"),
            ("MEDIUM", GOLD, "medium"),
            ("HARD", RED, "hard"),
        ]
    ):
        button = popup_button(choices, label, color, lambda value=difficulty: choose(value))
        button.configure(font=(FONT_FACE, 12, "bold"), pady=14)
        button.grid(row=0, column=col, sticky="nsew", padx=6)

    root.wait_window()
    if not completed["done"]:
        callback("medium")


def show_rating_popup(minutes, subject, goal, topic, callback):
    accent = subject_hex(subject)
    completed = {"done": False}
    root, panel = make_readable_popup("SESSION COMPLETE", 760, 380, accent)

    tk.Label(
        panel,
        text=f"{minutes} minutes in {subject}",
        font=(FONT_FACE, 16, "bold"),
        fg=POPUP_TEXT,
        bg=POPUP_BG,
    ).pack(pady=(10, 4))
    tk.Label(
        panel,
        text=f"Goal: {goal or 'Not set'}",
        font=(FONT_FACE, POPUP_BODY_PT),
        fg=POPUP_MUTED,
        bg=POPUP_BG,
    ).pack(pady=(0, 4))
    tk.Label(
        panel,
        text=trim_text(topic or subject, 78),
        font=(FONT_FACE, POPUP_BODY_PT),
        fg=POPUP_MUTED,
        bg=POPUP_BG,
    ).pack(pady=(0, 12))
    tk.Label(
        panel,
        text="How did this session feel?",
        font=(FONT_FACE, 14, "bold"),
        fg=POPUP_TEXT,
        bg=POPUP_BG,
    ).pack(pady=(0, 12))

    tk.Label(
        panel,
        text="Pick the closest feeling and I will save the session notes next.",
        font=(FONT_FACE, 11),
        fg=POPUP_MUTED,
        bg=POPUP_BG,
    ).pack(pady=(0, 10))

    choices = tk.Frame(panel, bg=POPUP_BG)
    choices.pack(fill="x", padx=28, pady=(0, 14))
    choices.grid_columnconfigure(0, weight=1)
    choices.grid_columnconfigure(1, weight=1)

    def choose(mood):
        completed["done"] = True
        play_habit_sound("goal_pick")
        root.destroy()
        show_topic_difficulty_popup(
            subject,
            topic,
            lambda difficulty: show_notes_popup(
                minutes,
                subject,
                topic,
                goal,
                lambda notes: callback(mood, notes, difficulty),
            ),
        )

    crushed_btn = popup_button(choices, "CRUSHED IT", CYAN, lambda: choose("crushed"))
    crushed_btn.configure(font=(FONT_FACE, 12, "bold"), pady=12)
    crushed_btn.grid(row=0, column=0, sticky="nsew", padx=(0, 8), pady=(0, 10))

    okay_btn = popup_button(choices, "IT WAS OKAY", GOLD, lambda: choose("okay"))
    okay_btn.configure(font=(FONT_FACE, 12, "bold"), pady=12)
    okay_btn.grid(row=0, column=1, sticky="nsew", padx=(8, 0), pady=(0, 10))

    struggled_btn = popup_button(choices, "STRUGGLED", RED, lambda: choose("struggled"))
    struggled_btn.configure(font=(FONT_FACE, 12, "bold"), pady=12)
    struggled_btn.grid(row=1, column=0, columnspan=2, sticky="nsew", pady=(0, 0))
    root.wait_window()
    if not completed["done"]:
        show_topic_difficulty_popup(
            subject,
            topic,
            lambda difficulty: show_notes_popup(
                minutes,
                subject,
                topic,
                goal,
                lambda notes: callback("okay", notes, difficulty),
            ),
        )


def show_alarm_popup(title, message, accent):
    root, panel = make_readable_popup(title, 700, 300, accent)
    tk.Label(
        panel,
        text=message,
        font=(FONT_FACE, 16, "bold"),
        fg=POPUP_TEXT,
        bg=POPUP_BG,
        wraplength=620,
        justify="center",
    ).pack(padx=30, pady=(40, 28))
    popup_button(panel, "CLOSE", accent, root.destroy).pack(fill="x", padx=28, pady=(0, 24))
    root.wait_window()


def show_weekly_progress_popup():
    data = load_streak()
    summary = compute_weekly_summary(data)
    weak_items = active_weak_topics(data, limit=4)
    accent = {"S": CYAN, "A": GREEN, "B": GOLD, "C": RED}[summary["grade"]]
    subjects = ", ".join(summary["subjects"]) if summary["subjects"] else "No subjects covered yet"
    root, panel = make_readable_popup(
        "WEEKLY PROGRESS",
        840,
        620,
        accent,
        subtitle=f"Week of {summary['start']} to {summary['end']}",
    )

    stats = popup_section(panel, "THIS WEEK", accent)
    lines = [
        f"Total minutes studied this week: {summary['total_minutes']}",
        f"Subjects covered: {subjects}",
        f"Current streak: {summary['streak']} day(s)",
        f"Study days this week: {summary['studied_days']}",
        f"Grade: {summary['grade']}",
        f"Planned MCQ target: {summary['weekly_target']['weekly_mcq_target']}",
    ]
    for line in lines:
        tk.Label(
            stats,
            text=line,
            font=(FONT_FACE, POPUP_BODY_PT),
            fg=POPUP_TEXT,
            bg=POPUP_PANEL,
            anchor="w",
            justify="left",
        ).pack(fill="x", padx=16, pady=4)

    message_section = popup_section(panel, "MESSAGE", accent)
    tk.Label(
        message_section,
        text=summary["message"],
        font=(FONT_FACE, 14, "bold"),
        fg=accent,
        bg=POPUP_PANEL,
        wraplength=740,
        justify="left",
    ).pack(fill="x", padx=16, pady=(0, 18))

    target_section = popup_section(panel, "WEEKLY TARGET", accent)
    target_lines = [
        f"Topics to close: {', '.join(summary['weekly_target']['topics_to_finish'][:4])}",
        f"Self-test: {summary['weekly_target']['weekly_self_test']}",
        f"Confidence target: {summary['weekly_target']['score_target']}",
    ]
    for line in target_lines:
        tk.Label(
            target_section,
            text=line,
            font=(FONT_FACE, POPUP_BODY_PT),
            fg=POPUP_TEXT,
            bg=POPUP_PANEL,
            anchor="w",
            justify="left",
            wraplength=740,
        ).pack(fill="x", padx=16, pady=3)

    weak_section = popup_section(panel, "WEAK TOPICS TO REVISIT", GOLD if weak_items else accent)
    weak_lines = [
        f"{item.get('subject', '')}: {item.get('topic', '')} ({item.get('last_difficulty', '').upper()})"
        for item in weak_items
    ] or ["No weak topics marked yet. The app will build this list after sessions."]
    for line in weak_lines:
        tk.Label(
            weak_section,
            text=line,
            font=(FONT_FACE, POPUP_BODY_PT),
            fg=POPUP_TEXT,
            bg=POPUP_PANEL,
            anchor="w",
            justify="left",
            wraplength=740,
        ).pack(fill="x", padx=16, pady=3)

    popup_button(panel, "CLOSE", accent, root.destroy).pack(fill="x", padx=28, pady=(8, 24))
    root.wait_window()


def maybe_show_weekly_summary():
    today = date.today()
    if today.weekday() != 6:
        return
    data = load_streak()
    today_key = str(today)
    if data.get("weekly_summary_last_shown") == today_key:
        return
    data["weekly_summary_last_shown"] = today_key
    save_streak(data)
    show_weekly_progress_popup()


def show_catchup_plan_popup(subject, behind_days, topic):
    accent = subject_hex(subject)
    root, panel = make_readable_popup("3-DAY CATCH-UP PLAN", 820, 500, accent)
    tk.Label(
        panel,
        text=f"{subject} is {behind_days} day(s) behind. Keep the plan small and finishable.",
        font=(FONT_FACE, POPUP_BODY_PT),
        fg=POPUP_MUTED,
        bg=POPUP_BG,
        wraplength=720,
        justify="center",
    ).pack(pady=(10, 18))

    section = popup_section(panel, "NEXT 3 STUDY BLOCKS", accent)
    lines = [
        f"DAY 1: Finish the core notes for {topic}. Stop after one clean summary page.",
        "DAY 2: Solve 20 MCQs. Mark every mistake with one-line corrections.",
        "DAY 3: Redo the mistakes, write formulas by hand, then mark the topic done if it feels clear.",
        "RULE: No perfect notes. The catch-up win is finishing the loop.",
    ]
    for line in lines:
        tk.Label(
            section,
            text=line,
            font=(FONT_FACE, POPUP_BODY_PT),
            fg=POPUP_TEXT if not line.startswith("RULE") else GOLD,
            bg=POPUP_PANEL,
            anchor="w",
            justify="left",
            wraplength=720,
        ).pack(fill="x", padx=16, pady=5)

    popup_button(panel, "I CAN DO THIS", accent, root.destroy).pack(fill="x", padx=28, pady=(12, 24))
    root.wait_window()


def show_weak_topics_popup():
    data = load_streak()
    items = active_weak_topics(data, limit=10)
    root, panel = make_readable_popup("WEAK TOPICS", 820, 520, GOLD)
    tk.Label(
        panel,
        text="These are the topics Krish marked Medium or Hard recently.",
        font=(FONT_FACE, POPUP_BODY_PT),
        fg=POPUP_MUTED,
        bg=POPUP_BG,
        wraplength=720,
        justify="center",
    ).pack(pady=(10, 16))

    section = popup_section(panel, "REVISION QUEUE", GOLD)
    lines = [
        f"{item.get('subject', '')}: {item.get('topic', '')} | {item.get('last_difficulty', '').upper()} | hard hits: {item.get('hard_count', 0)}"
        for item in items
    ] or ["No weak topics yet. They appear here after session wrap-up."]
    for line in lines:
        tk.Label(
            section,
            text=line,
            font=(FONT_FACE, POPUP_BODY_PT),
            fg=POPUP_TEXT,
            bg=POPUP_PANEL,
            anchor="w",
            justify="left",
            wraplength=720,
        ).pack(fill="x", padx=16, pady=4)

    popup_button(panel, "CLOSE", GOLD, root.destroy).pack(fill="x", padx=28, pady=(10, 24))
    root.wait_window()


def show_exam_reference_popup():
    references = STUDY_PLAN_DATA["references"]
    root, panel = make_readable_popup("EXAM REFERENCE", 980, 760, CYAN)

    top = popup_section(panel, "FAST RECALL", CYAN)
    fast_lines = [
        "Math formula focus:",
        *[f"- {item}" for item in canonical_formula_sheet_for_subject("MATHEMATICS")[:4]],
        "Physics formula focus:",
        *[f"- {item}" for item in canonical_formula_sheet_for_subject("PHYSICS")[:4]],
    ]
    for line in fast_lines:
        tk.Label(
            top,
            text=line,
            font=(FONT_FACE, POPUP_BODY_PT),
            fg=CYAN if line.endswith(":") else POPUP_TEXT,
            bg=POPUP_PANEL,
            anchor="w",
            justify="left",
            wraplength=860,
        ).pack(fill="x", padx=16, pady=2)

    english_section = popup_section(panel, "ENGLISH MINIMAL STRATEGY", GOLD)
    english_lines = list(references["english_strategy"]) + [f"Rule {idx + 1}: {rule}" for idx, rule in enumerate(canonical_english_quick_rules())]
    for line in english_lines:
        tk.Label(
            english_section,
            text=line,
            font=(FONT_FACE, POPUP_BODY_PT),
            fg=POPUP_TEXT,
            bg=POPUP_PANEL,
            anchor="w",
            justify="left",
            wraplength=860,
        ).pack(fill="x", padx=16, pady=2)

    topics_section = popup_section(panel, "TOP REPEATED IOE TOPICS", GREEN)
    for topic in references["top_20_topics"]:
        tk.Label(
            topics_section,
            text=f"- {topic}",
            font=(FONT_FACE, POPUP_BODY_PT),
            fg=POPUP_TEXT,
            bg=POPUP_PANEL,
            anchor="w",
            justify="left",
            wraplength=860,
        ).pack(fill="x", padx=16, pady=2)

    last_week = popup_section(panel, "LAST 7 DAYS", RED)
    for line in references["last_7_days"]:
        tk.Label(
            last_week,
            text=line,
            font=(FONT_FACE, POPUP_BODY_PT),
            fg=POPUP_TEXT,
            bg=POPUP_PANEL,
            anchor="w",
            justify="left",
            wraplength=860,
        ).pack(fill="x", padx=16, pady=2)

    popup_button(panel, "CLOSE", CYAN, root.destroy).pack(fill="x", padx=28, pady=(12, 24))
    root.wait_window()


def show_subject_selector(callback):
    today = date.today()
    recommended = recommended_subject_for_day(today)
    last = load_streak().get("last_subject", "")
    root, panel = make_readable_popup(
        "SELECT YOUR SUBJECT",
        980,
        720,
        subject_hex(recommended),
        subtitle=f"Today's recommended subject: {recommended} | {today.strftime('%A')}",
    )

    banner = tk.Frame(panel, bg=POPUP_PANEL, highlightbackground=subject_hex(recommended), highlightthickness=2)
    banner.pack(fill="x", padx=28, pady=(4, 18))
    tk.Label(
        banner,
        text=f"TODAY'S RECOMMENDED SUBJECT: {recommended}",
        font=(FONT_FACE, 16, "bold"),
        fg=subject_hex(recommended),
        bg=POPUP_PANEL,
    ).pack(pady=(12, 4))
    tk.Label(
        banner,
        text="Mon/Thu = Mathematics | Tue/Fri = Physics | Wed/Sat = Chemistry | Sun = English",
        font=(FONT_FACE, POPUP_BODY_PT),
        fg=POPUP_MUTED,
        bg=POPUP_PANEL,
    ).pack(pady=(0, 12))

    grid = tk.Frame(panel, bg=POPUP_BG)
    grid.pack(fill="both", expand=True, padx=28, pady=(0, 18))
    subjects = [
        ("MATHEMATICS", "Calculus, Algebra, Trigonometry", SUBJECT_MARKS["MATHEMATICS"]),
        ("PHYSICS", "Mechanics, Waves, Electricity", SUBJECT_MARKS["PHYSICS"]),
        ("CHEMISTRY", "Organic, Inorganic, Physical", SUBJECT_MARKS["CHEMISTRY"]),
        ("ENGLISH", "Grammar, Reading, Writing", SUBJECT_MARKS["ENGLISH"]),
    ]

    def pick_subject(name):
        root.destroy()
        callback(name)

    for index, (name, subtitle, marks) in enumerate(subjects):
        accent = subject_hex(name)
        card = tk.Frame(
            grid,
            bg=POPUP_PANEL,
            highlightbackground=accent,
            highlightthickness=4 if name == recommended else 2,
        )
        card.grid(row=index // 2, column=index % 2, sticky="nsew", padx=10, pady=10)
        grid.grid_columnconfigure(index % 2, weight=1)
        grid.grid_rowconfigure(index // 2, weight=1)
        card.configure(cursor="hand2")

        def bind_click(widget, chosen=name):
            widget.configure(cursor="hand2")
            widget.bind("<Button-1>", lambda event, subject_name=chosen: pick_subject(subject_name))

        flags = []
        if name == recommended:
            flags.append("RECOMMENDED TODAY")
        if name == last:
            flags.append("LAST SESSION")
        flag_label = tk.Label(
            card,
            text=" | ".join(flags) if flags else " ",
            font=(FONT_FACE, 11, "bold"),
            fg=accent,
            bg=POPUP_PANEL,
        )
        flag_label.pack(anchor="e", padx=16, pady=(12, 4))

        title_label = tk.Label(card, text=name, font=(FONT_FACE, 18, "bold"), fg=accent, bg=POPUP_PANEL)
        title_label.pack(pady=(6, 6))

        marks_label = tk.Label(
            card,
            text=f"{marks} marks",
            font=(FONT_FACE, 12, "bold"),
            fg=POPUP_TEXT,
            bg=POPUP_PANEL,
        )
        marks_label.pack()

        subtitle_label = tk.Label(
            card,
            text=subtitle,
            font=(FONT_FACE, POPUP_BODY_PT),
            fg=POPUP_MUTED,
            bg=POPUP_PANEL,
        )
        subtitle_label.pack(pady=(8, 8))

        context = lesson_context(name)
        next_topic = context["display"]["topic"]
        topic_label = tk.Label(
            card,
            text=f"Next topic: {trim_text(next_topic, 36)} | Day {context['display']['day_of_topic']} of {context['display']['topic_days']}",
            font=(FONT_FACE, POPUP_BODY_PT),
            fg=POPUP_TEXT,
            bg=POPUP_PANEL,
            wraplength=360,
            justify="center",
        )
        topic_label.pack(padx=18, pady=(0, 10))

        click_label = tk.Label(
            card,
            text="CLICK ANYWHERE ON THIS CARD TO START",
            font=(FONT_FACE, 11, "bold"),
            fg=accent,
            bg=POPUP_PANEL,
        )
        click_label.pack(padx=18, pady=(0, 16))

        for widget in (card, flag_label, title_label, marks_label, subtitle_label, topic_label, click_label):
            bind_click(widget)

    tk.Label(
        panel,
        text=f"Press ESC to use the recommended subject: {recommended}",
        font=(FONT_FACE, POPUP_BODY_PT),
        fg=POPUP_DIM,
        bg=POPUP_BG,
    ).pack(pady=(0, 18))
    root.bind("<Escape>", lambda event: (root.destroy(), callback(recommended)))
    root.wait_window()


def show_today_plan(subject, on_start):
    global current_session_goal, current_session_quote, current_session_recommended, current_session_topic
    recommended = recommended_subject_for_day()
    accent = subject_hex(subject)
    quote = random.choice(MOTIVATIONAL_QUOTES)
    current_session_quote = quote
    current_session_recommended = recommended

    context = lesson_context(subject)
    scheduled = context["scheduled"]
    display = context["display"]
    behind_days = context["behind_days"]
    ahead_days = context["ahead_days"]
    day_number = display["day_number"]
    overall_pct = int(((context["day_progress"] + 1) / max(1, JOURNEY_TOTAL_DAYS)) * 100)
    subject_plan = [entry for entry in STUDY_PLAN_DATA["plan"] if entry["subject"] == subject]
    subject_pct = int((display["subject_occurrence"] / max(1, len(subject_plan))) * 100)
    weekly_target = weekly_target_for_day(display)
    milestone = milestone_for_day(display)
    routine = adaptive_routine(subject, display)
    sitting = sitting_finish_line(subject, display, routine)
    tasks = lesson_tasks(display)
    if is_weak_topic(subject, display["topic"]):
        tasks.append("5. Weak-topic drill: redo 5 missed questions before moving on.")
    detail = display.get("detail", display.get("focus", display["topic"]))
    resource = display.get("resource", SUBJECT_URLS.get(subject, SUBJECT_URLS["General"]))
    morning_plan = list(display.get("morning_plan", []))
    evening_plan = list(display.get("evening_plan", []))
    if morning_plan:
        morning_plan.insert(1, f"Adaptive focus: {routine['morning_adjustment']}")
    if evening_plan:
        evening_plan.insert(2, f"Adaptive focus: {routine['evening_adjustment']}")

    popup_width = min(1040, max(860, _tk_root.winfo_screenwidth() - 120))
    popup_height = min(860, max(640, _tk_root.winfo_screenheight() - 120))

    root, panel = make_readable_popup(
        "TODAY'S PLAN",
        popup_width,
        popup_height,
        accent,
        subtitle=f"Recommended today: {recommended} | Journey day {day_number}/{JOURNEY_TOTAL_DAYS}",
    )

    button_row = tk.Frame(panel, bg=POPUP_BG)
    button_row.pack(side="bottom", fill="x", padx=28, pady=(0, 20))

    helper_label = tk.Label(
        panel,
        text="Scroll for the full lesson plan. The session buttons stay pinned at the bottom.",
        font=(FONT_FACE, 11),
        fg=POPUP_MUTED,
        bg=POPUP_BG,
        wraplength=popup_width - 120,
        justify="center",
    )
    helper_label.pack(side="bottom", padx=28, pady=(0, 10))

    content_wrap = tk.Frame(panel, bg=POPUP_BG)
    content_wrap.pack(fill="both", expand=True, padx=0, pady=(0, 10))

    canvas = tk.Canvas(
        content_wrap,
        bg=POPUP_BG,
        highlightthickness=0,
        bd=0,
        relief="flat",
    )
    scrollbar = tk.Scrollbar(content_wrap, orient="vertical", command=canvas.yview)
    scroll_frame = tk.Frame(canvas, bg=POPUP_BG)
    scroll_window = canvas.create_window((0, 0), window=scroll_frame, anchor="nw")
    canvas.configure(yscrollcommand=scrollbar.set)
    canvas.pack(side="left", fill="both", expand=True, padx=(0, 6))
    scrollbar.pack(side="right", fill="y", padx=(0, 18))

    def sync_scroll_region(event=None):
        canvas.configure(scrollregion=canvas.bbox("all"))
        canvas.itemconfigure(scroll_window, width=canvas.winfo_width())

    scroll_frame.bind("<Configure>", sync_scroll_region)
    canvas.bind("<Configure>", sync_scroll_region)

    def on_mousewheel(event):
        delta = 0
        if getattr(event, "delta", 0):
            delta = -1 * int(event.delta / 120)
        elif getattr(event, "num", None) == 4:
            delta = -1
        elif getattr(event, "num", None) == 5:
            delta = 1
        if delta:
            canvas.yview_scroll(delta, "units")

    for widget in (root, panel, canvas, scroll_frame):
        widget.bind("<MouseWheel>", on_mousewheel, add="+")
        widget.bind("<Button-4>", on_mousewheel, add="+")
        widget.bind("<Button-5>", on_mousewheel, add="+")

    badge_row = tk.Frame(scroll_frame, bg=POPUP_BG)
    badge_row.pack(fill="x", padx=28, pady=(8, 18))
    badges = [
        (subject, "#071014", accent),
        (f"DAY {day_number}/{JOURNEY_TOTAL_DAYS}", "#071014", GOLD),
        (f"DAY {display['day_of_topic']} OF {display['topic_days']} ON THIS TOPIC", GREEN, POPUP_PANEL),
    ]
    if subject == recommended:
        badges.append(("ON ROTATION", "#071014", GREEN))
    else:
        badges.append((f"REC: {recommended}", subject_hex(recommended), POPUP_PANEL))
    for text, fg, bg in badges:
        tk.Label(
            badge_row,
            text=text,
            font=(FONT_FACE, 13, "bold"),
            fg=fg,
            bg=bg,
            padx=16,
            pady=8,
            highlightbackground=accent if bg == POPUP_PANEL else bg,
            highlightthickness=2,
        ).pack(side="left", padx=(0, 12))

    finish_section = popup_section(scroll_frame, "THIS SITTING'S FINISH LINE", accent)
    finish_header = tk.Frame(finish_section, bg=POPUP_PANEL)
    finish_header.pack(fill="x", padx=16, pady=(0, 8))
    tk.Label(
        finish_header,
        text=sitting["session_name"],
        font=(FONT_FACE, 15, "bold"),
        fg=GOLD,
        bg=POPUP_PANEL,
        anchor="w",
        justify="left",
    ).pack(fill="x")
    tk.Label(
        finish_header,
        text=sitting["finish_line"],
        font=(FONT_FACE, POPUP_TOPIC_PT, "bold"),
        fg=accent,
        bg=POPUP_PANEL,
        wraplength=popup_width - 140,
        anchor="w",
        justify="left",
    ).pack(fill="x", pady=(4, 0))
    finish_lines = [
        f"Minimum win before stopping: {sitting['minimum_win']}",
        f"MCQ target for this sitting: {sitting['mcq_target']}",
        f"Checkpoint: {sitting['checkpoint']}",
        f"Stretch win if you still have energy: {sitting['stretch_win']}",
        f"Mentor pressure: {sitting['pressure']}",
    ]
    for line in finish_lines:
        tk.Label(
            finish_section,
            text=line,
            font=(FONT_FACE, 13, "bold" if line.startswith("Minimum") or line.startswith("Mentor") else "normal"),
            fg=GOLD if line.startswith("Minimum") else RED if line.startswith("Mentor") else POPUP_TEXT,
            bg=POPUP_PANEL,
            anchor="w",
            justify="left",
            wraplength=popup_width - 140,
        ).pack(fill="x", padx=16, pady=3)
    tk.Label(
        finish_section,
        text="Notebook proof to write before you call the session done:",
        font=(FONT_FACE, 13, "bold"),
        fg=accent,
        bg=POPUP_PANEL,
        anchor="w",
        justify="left",
        wraplength=popup_width - 140,
    ).pack(fill="x", padx=16, pady=(8, 3))
    for line in sitting["notebook_lines"]:
        tk.Label(
            finish_section,
            text=f"- {line}",
            font=(FONT_FACE, POPUP_BODY_PT),
            fg=POPUP_TEXT,
            bg=POPUP_PANEL,
            anchor="w",
            justify="left",
            wraplength=popup_width - 140,
        ).pack(fill="x", padx=16, pady=2)

    schedule_section = popup_section(scroll_frame, "DAILY LESSON SCHEDULE", accent)
    schedule_lines = [
        f"DAY {day_number} OF {JOURNEY_TOTAL_DAYS}",
        f"SUBJECT: {subject}",
        f"PHASE: {display['phase']} | {display['month_label']} | WEEK {display['week_number']}",
        f"TOPIC: {display['topic']}",
        f"SUBTOPIC: {display['subtopic']}",
        f"DAY {display['day_of_topic']} OF {display['topic_days']} ON THIS TOPIC",
        f"DIFFICULTY BAND: {display['difficulty_band']}",
        "TODAY'S TASK:",
    ] + tasks + [
        f"FINISH THIS TOPIC BY: {deadline_label(display.get('finish_date'))}",
        f"NEXT TOPIC: {display.get('next_topic', BUFFER_TOPIC)}"
        + (
            f" (starts {deadline_label(display.get('next_topic_start'))})"
            if display.get("next_topic_start") and display.get("next_topic_start") >= date.today()
            else ""
        ),
    ]
    for line in schedule_lines:
        tk.Label(
            schedule_section,
            text=line,
            font=(FONT_FACE, 13 if line.startswith("①") or line.startswith("②") or line.startswith("③") or line.startswith("④") else POPUP_BODY_PT, "bold" if ":" in line and not line.startswith("①") else "normal"),
            fg=accent if line.startswith("DAY ") or line.startswith("SUBJECT") or line.startswith("TOPIC") else POPUP_TEXT,
            bg=POPUP_PANEL,
            anchor="w",
            justify="left",
            wraplength=popup_width - 140,
        ).pack(fill="x", padx=16, pady=2)

    quote_section = popup_section(scroll_frame, "TODAY'S QUOTE", accent)
    tk.Label(
        quote_section,
        text=f"\"{quote}\"",
        font=(FONT_FACE, 14, "bold"),
        fg=GOLD,
        bg=POPUP_PANEL,
        wraplength=popup_width - 140,
        justify="center",
    ).pack(fill="x", padx=16, pady=(0, 14))

    detail_section = popup_section(scroll_frame, "FOCUS FOR THIS LESSON", accent)
    tk.Label(
        detail_section,
        text=detail,
        font=(FONT_FACE, POPUP_BODY_PT),
        fg=POPUP_TEXT,
        bg=POPUP_PANEL,
        justify="left",
        wraplength=popup_width - 140,
        anchor="w",
    ).pack(fill="x", padx=16, pady=(0, 14))

    adaptive_section = popup_section(scroll_frame, "ADAPTIVE ROUTINE", accent)
    adaptive_lines = [
        routine["headline"],
        f"Review target: {routine['review_target']}",
        f"Focus split: {routine['split']}",
    ]
    for line in adaptive_lines:
        tk.Label(
            adaptive_section,
            text=line,
            font=(FONT_FACE, POPUP_BODY_PT),
            fg=POPUP_TEXT,
            bg=POPUP_PANEL,
            justify="left",
            wraplength=popup_width - 140,
            anchor="w",
        ).pack(fill="x", padx=16, pady=3)

    morning_section = popup_section(scroll_frame, "MORNING SESSION (50 MIN)", accent)
    for line in morning_plan:
        tk.Label(
            morning_section,
            text=line,
            font=(FONT_FACE, POPUP_BODY_PT),
            fg=POPUP_TEXT,
            bg=POPUP_PANEL,
            justify="left",
            wraplength=popup_width - 140,
            anchor="w",
        ).pack(fill="x", padx=16, pady=3)

    evening_section = popup_section(scroll_frame, "EVENING SESSION (90 MIN)", accent)
    for line in evening_plan:
        tk.Label(
            evening_section,
            text=line,
            font=(FONT_FACE, POPUP_BODY_PT),
            fg=POPUP_TEXT,
            bg=POPUP_PANEL,
            justify="left",
            wraplength=popup_width - 140,
            anchor="w",
        ).pack(fill="x", padx=16, pady=3)

    self_test_section = popup_section(scroll_frame, "SELF-TEST AND FORMULA FOCUS", accent)
    for line in [
        f"Self-test: {display['self_test_question']}",
        f"Formula focus: {display['formula_focus']}",
        f"MCQ targets today: morning {display['mcq_target_morning']} | evening {display['mcq_target_evening']}",
    ]:
        tk.Label(
            self_test_section,
            text=line,
            font=(FONT_FACE, POPUP_BODY_PT),
            fg=POPUP_TEXT,
            bg=POPUP_PANEL,
            justify="left",
            wraplength=popup_width - 140,
            anchor="w",
        ).pack(fill="x", padx=16, pady=3)

    weekly_section = popup_section(scroll_frame, "WEEKLY TARGET", accent)
    weekly_lines = [
        f"Topics to finish this week: {', '.join(weekly_target['topics_to_finish'][:4])}",
        f"Weekly MCQ target: {weekly_target['weekly_mcq_target']}",
        f"Weekly self-test: {weekly_target['weekly_self_test']}",
        f"Confidence score target: {weekly_target['score_target']}",
    ]
    for line in weekly_lines:
        tk.Label(
            weekly_section,
            text=line,
            font=(FONT_FACE, POPUP_BODY_PT),
            fg=POPUP_TEXT,
            bg=POPUP_PANEL,
            justify="left",
            wraplength=popup_width - 140,
            anchor="w",
        ).pack(fill="x", padx=16, pady=3)

    milestone_section = popup_section(scroll_frame, "MONTHLY CHECKPOINT", accent)
    milestone_lines = [
        milestone["title"],
        f"Mini mock scope: {milestone['mini_mock_scope']}",
        f"On-track score: {milestone['score_target']}",
        f"If behind: {milestone['catch_up_plan']}",
    ]
    for line in milestone_lines:
        tk.Label(
            milestone_section,
            text=line,
            font=(FONT_FACE, 13 if line == milestone["title"] else POPUP_BODY_PT, "bold" if line == milestone["title"] else "normal"),
            fg=accent if line == milestone["title"] else POPUP_TEXT,
            bg=POPUP_PANEL,
            justify="left",
            wraplength=popup_width - 140,
            anchor="w",
        ).pack(fill="x", padx=16, pady=3)

    progress_section = popup_section(scroll_frame, "PROGRESS SIGNAL", accent)
    progress_message = (
        f"Overall journey progress: {overall_pct}%.\n"
        f"{subject} roadmap progress: {subject_pct}%."
    )
    if behind_days > 0:
        progress_message += f"\n{catch_up_message(subject, behind_days)}"
    elif ahead_days > 0:
        progress_message += f"\nYOU ARE {ahead_days} DAYS AHEAD - keep the momentum."
    elif scheduled["topic"] != display["topic"]:
        progress_message += f"\nSchedule expects: {scheduled['topic']}."
    tk.Label(
        progress_section,
        text=progress_message,
        font=(FONT_FACE, POPUP_BODY_PT),
        fg=POPUP_TEXT,
        bg=POPUP_PANEL,
        justify="left",
        wraplength=popup_width - 140,
        anchor="w",
    ).pack(fill="x", padx=16, pady=(0, 10))
    bar = tk.Frame(progress_section, bg="#193128", height=22)
    bar.pack(fill="x", padx=16, pady=(0, 16))
    tk.Frame(bar, bg=GREEN if overall_pct < 70 else GOLD if overall_pct < 90 else RED).place(relheight=1, relwidth=max(0.0, overall_pct / 100))
    if behind_days > 0:
        popup_button(
            progress_section,
            "MAKE 3-DAY CATCH-UP PLAN",
            GOLD,
            lambda: show_catchup_plan_popup(subject, behind_days, display["topic"]),
        ).pack(fill="x", padx=16, pady=(0, 16))

    tk.Label(
        scroll_frame,
        text="Open the resource for a refresher, then start the session and choose a goal.",
        font=(FONT_FACE, POPUP_BODY_PT),
        fg=POPUP_MUTED,
        bg=POPUP_BG,
        wraplength=popup_width - 120,
        justify="center",
    ).pack(padx=28, pady=(0, 14))

    def open_resource():
        safe_open_url(resource)

    def start_with_goal(mark_next=False):
        global current_session_day_number, current_session_goal, current_session_topic, advance_after_session
        goal = show_session_goal_popup(subject, display["topic"])
        if not goal:
            return

        current_session_day_number = display["day_number"]
        current_session_goal = f"{goal} | {sitting['saved_goal']}"
        current_session_topic = display["topic"]
        advance_after_session = mark_next
        root.destroy()
        threading.Thread(target=on_start, daemon=True).start()

    popup_button(button_row, "OPEN RESOURCE", accent, open_resource).pack(side="left", expand=True, fill="x", padx=(0, 10))
    popup_button(button_row, "START SESSION", CYAN, lambda: start_with_goal(False)).pack(side="left", expand=True, fill="x", padx=10)
    popup_button(button_row, "DONE AND NEXT", GREEN, lambda: start_with_goal(True)).pack(side="left", expand=True, fill="x", padx=(10, 0))
    root.wait_window()


def show_late_night_warning(hour, callback):
    root, panel = make_readable_popup("LATE NIGHT WARNING", 760, 360, GOLD)
    tk.Label(
        panel,
        text=f"It is already after {hour}:00.",
        font=(FONT_FACE, 16, "bold"),
        fg=POPUP_TEXT,
        bg=POPUP_BG,
    ).pack(pady=(34, 8))
    tk.Label(
        panel,
        text="Sleep helps memory, retention, and pace. You can still continue if you really want this session.",
        font=(FONT_FACE, POPUP_BODY_PT),
        fg=POPUP_MUTED,
        bg=POPUP_BG,
        wraplength=680,
        justify="center",
    ).pack(pady=(0, 24))

    buttons = tk.Frame(panel, bg=POPUP_BG)
    buttons.pack(fill="x", padx=28, pady=(0, 20))
    popup_button(buttons, "YES, START ANYWAY", CYAN, lambda: (root.destroy(), callback(True))).pack(side="left", expand=True, fill="x", padx=(0, 8))
    popup_button(buttons, "NO, I'LL REST", RED, lambda: (root.destroy(), callback(False))).pack(side="left", expand=True, fill="x", padx=(8, 0))
    root.wait_window()


def show_min_session_warning(elapsed_min, callback):
    root, panel = make_readable_popup("STOPPING EARLY?", 760, 380, GOLD)
    tk.Label(
        panel,
        text=f"You have only studied for {elapsed_min} minutes.",
        font=(FONT_FACE, 16, "bold"),
        fg=POPUP_TEXT,
        bg=POPUP_BG,
    ).pack(pady=(24, 10))
    tk.Label(
        panel,
        text="A slightly longer session usually helps retention more. Do you want to keep going or stop anyway?",
        font=(FONT_FACE, POPUP_BODY_PT),
        fg=POPUP_MUTED,
        bg=POPUP_BG,
        wraplength=680,
        justify="center",
    ).pack(pady=(0, 16))

    progress = min(1.0, elapsed_min / max(1, MIN_SESSION // 60))
    bar = tk.Frame(panel, bg="#193128", height=24)
    bar.pack(fill="x", padx=34, pady=(0, 22))
    tk.Frame(bar, bg=RED if progress < 0.5 else GOLD if progress < 1 else GREEN).place(relheight=1, relwidth=progress)

    buttons = tk.Frame(panel, bg=POPUP_BG)
    buttons.pack(fill="x", padx=28, pady=(0, 22))
    popup_button(buttons, "KEEP GOING", CYAN, lambda: (root.destroy(), callback(False))).pack(side="left", expand=True, fill="x", padx=(0, 8))
    popup_button(buttons, "STOP ANYWAY", RED, lambda: (root.destroy(), callback(True))).pack(side="left", expand=True, fill="x", padx=(8, 0))
    root.wait_window()


def show_ioe_countdown():
    data = load_streak()
    projection = compute_exam_projection(data)
    days_left = projection["days_left"]
    accent = CYAN if days_left > 60 else GOLD if days_left > 30 else RED
    root, panel = make_readable_popup("IOE ENTRANCE COUNTDOWN", 780, 500, accent)

    tk.Label(
        panel,
        text=str(days_left),
        font=(FONT_FACE, 42, "bold"),
        fg=accent,
        bg=POPUP_BG,
    ).pack(pady=(12, 0))
    tk.Label(
        panel,
        text="days remaining",
        font=(FONT_FACE, 14, "bold"),
        fg=POPUP_TEXT,
        bg=POPUP_BG,
    ).pack(pady=(0, 14))

    countdown_section = popup_section(panel, "PACE CHECK", accent)
    lines = [
        f"Studied so far: {projection['total_hours']} hours",
        f"You need {projection['remaining_hours']} hours more.",
        f"At current pace, you'll have {projection['projected_total']} hours.",
        f"Current streak: {data.get('streak', 0)} days",
    ]
    for line in lines:
        tk.Label(
            countdown_section,
            text=line,
            font=(FONT_FACE, 13, "bold"),
            fg=POPUP_TEXT,
            bg=POPUP_PANEL,
            anchor="w",
        ).pack(fill="x", padx=16, pady=5)

    tk.Label(
        panel,
        text=f"Exam date: {IOE_DATE} | Target: {TOTAL_TARGET_HOURS} total study hours",
        font=(FONT_FACE, POPUP_BODY_PT),
        fg=POPUP_MUTED,
        bg=POPUP_BG,
    ).pack(pady=(4, 18))
    popup_button(panel, "CLOSE", accent, root.destroy).pack(fill="x", padx=28, pady=(0, 24))
    root.wait_window()


def trigger_daily_alarm(kind):
    recommended = recommended_subject_for_day()
    accent = subject_hex(recommended)
    if kind == "morning":
        title = "MORNING SESSION"
        message = f"Time to study {recommended}"
    else:
        title = "EVENING SESSION"
        message = "Don't break the streak"
    show_windows_toast(title, message, WINDOWS_TOAST_SOUND)
    play_habit_sound("alarm")
    base.run_on_main_thread(lambda: show_alarm_popup(title, message, accent))


def alarm_scheduler_loop():
    while running:
        now = datetime.now()
        today_key = str(now.date())
        data = load_streak()
        history = data.setdefault("alarm_history", {})
        triggered = False
        for kind, (hour, minute) in ALARM_SCHEDULE.items():
            if now.hour == hour and now.minute == minute and history.get(kind) != today_key:
                history[kind] = today_key
                save_streak(data)
                trigger_daily_alarm(kind)
                triggered = True
        time.sleep(65 if triggered else 20)


def ensure_alarm_scheduler():
    if getattr(ensure_alarm_scheduler, "_started", False):
        return
    ensure_alarm_scheduler._started = True
    threading.Thread(target=alarm_scheduler_loop, daemon=True).start()


class PomodoroState:
    def __init__(self, work_secs, break_secs):
        self.default_work_secs = work_secs
        self.default_break_secs = break_secs
        self.work_secs = work_secs
        self.break_secs = break_secs
        self.breaks_enabled = True
        self.lock = threading.Lock()
        self.stop_event = threading.Event()
        self.running = False
        self.phase = "idle"
        self.session = 0
        self.phase_started_at = 0.0
        self.paused = False
        self.pause_started_at = None
        self.paused_accumulated = 0.0
        self.focus_accumulated = 0

    def start(self, work_secs=None, break_secs=None, breaks_enabled=True):
        with self.lock:
            self.work_secs = work_secs or self.default_work_secs
            self.break_secs = self.default_break_secs if break_secs is None else break_secs
            self.breaks_enabled = breaks_enabled
            self.stop_event = threading.Event()
            self.running = True
            self.phase = "work"
            self.session = 1
            self.phase_started_at = time.monotonic()
            self.paused = False
            self.pause_started_at = None
            self.paused_accumulated = 0.0
            self.focus_accumulated = 0

    def stop(self):
        with self.lock:
            self.work_secs = self.default_work_secs
            self.break_secs = self.default_break_secs
            self.breaks_enabled = True
            self.running = False
            self.phase = "idle"
            self.paused = False
            self.pause_started_at = None
            self.stop_event.set()

    def toggle_pause(self):
        with self.lock:
            if not self.running:
                return False
            now = time.monotonic()
            if self.paused:
                if self.pause_started_at is not None:
                    self.paused_accumulated += now - self.pause_started_at
                self.pause_started_at = None
                self.paused = False
            else:
                self.paused = True
                self.pause_started_at = now
            return self.paused

    def _phase_elapsed_locked(self, now=None):
        if not self.running or self.phase == "idle":
            return 0
        now = now or time.monotonic()
        ref = self.pause_started_at if (self.paused and self.pause_started_at is not None) else now
        return max(0, int(ref - self.phase_started_at - self.paused_accumulated))

    def snapshot(self):
        with self.lock:
            phase = self.phase if self.phase != "idle" else "work"
            phase_total = self.break_secs if phase == "break" else self.work_secs
            elapsed = self._phase_elapsed_locked()
            remaining = max(0, phase_total - elapsed)
            focus_seconds = self.focus_accumulated + (elapsed if phase == "work" and self.running else 0)
            progress = 0.0 if phase_total <= 0 else min(1.0, elapsed / phase_total)
            return {
                "running": self.running,
                "phase": phase,
                "session": max(1, self.session),
                "paused": self.paused,
                "elapsed": elapsed,
                "remaining": remaining,
                "phase_total": phase_total,
                "progress": progress,
                "focus_seconds": focus_seconds,
            }

    def focus_seconds(self):
        return self.snapshot()["focus_seconds"]

    def advance_if_needed(self):
        with self.lock:
            if not self.running or self.paused or self.phase == "idle":
                return None
            now = time.monotonic()
            elapsed = self._phase_elapsed_locked(now)
            phase_total = self.break_secs if self.phase == "break" else self.work_secs
            if elapsed < phase_total:
                return None

            if self.phase == "work":
                finished_session = self.session
                self.focus_accumulated += self.work_secs
                if not self.breaks_enabled:
                    self.running = False
                    self.phase = "idle"
                    self.stop_event.set()
                    return {
                        "new_phase": "complete",
                        "finished_session": finished_session,
                        "next_session": self.session,
                    }
                self.phase = "break"
                next_session = self.session
            else:
                finished_session = self.session
                self.phase = "work"
                self.session += 1
                next_session = self.session

            self.phase_started_at = now
            self.paused_accumulated = 0.0
            self.pause_started_at = None
            self.paused = False
            return {
                "new_phase": self.phase,
                "finished_session": finished_session,
                "next_session": next_session,
            }


pomodoro_state = PomodoroState(POMODORO_WORK, POMODORO_BREAK)


class StudyTimer:
    W, H = 380, 100
    MINI_W, MINI_H = 238, 52

    def __init__(self):
        self.root = None
        self.canvas = None
        self._tk_img = None
        self._running = False
        self.subject = "General"
        self.pause_btn = None
        self.stop_btn = None
        self.min_btn = None
        self._drag_x = 0
        self._drag_y = 0
        self._dragging = False
        self._press_x = 0
        self._press_y = 0
        self.minimized = False
        self.auto_minimized_once = False
        self._window_x = None
        self._window_y = None

    def start(self, subject):
        self.close()
        self.subject = subject
        self._running = True
        self.minimized = False
        self.auto_minimized_once = False
        self._run_window()

    def _run_window(self):
        self.root = tk.Toplevel(_tk_root)
        self.root.overrideredirect(True)
        self.root.attributes("-topmost", True)
        self.root.attributes("-alpha", 0.96)
        self.root.configure(bg=BG)
        sw = self.root.winfo_screenwidth()
        self._window_x = sw - self.W - 12
        self._window_y = 8
        self._apply_geometry()

        self.canvas = tk.Canvas(self.root, bg=BG, highlightthickness=0)
        self.canvas.pack(fill="both", expand=True)
        self.canvas.bind("<ButtonPress-1>", self._drag_start)
        self.canvas.bind("<B1-Motion>", self._drag_move)
        self.canvas.bind("<ButtonRelease-1>", self._drag_release)
        self.canvas.bind("<Double-Button-1>", lambda _event: self._toggle_minimize() if self.minimized else None)

        self.min_btn = tk.Button(
            self.root,
            text="MIN",
            font=(FONT_FACE, 10, "bold"),
            bg=BUTTON_BG,
            fg=CYAN,
            activebackground=BUTTON_ACTIVE,
            activeforeground=CYAN,
            relief="flat",
            bd=0,
            cursor="hand2",
            command=self._toggle_minimize,
            highlightbackground=CYAN,
            highlightthickness=2,
        )

        self.pause_btn = tk.Button(
            self.root,
            text="PAUSE",
            font=(FONT_FACE, 12, "bold"),
            bg=BUTTON_BG,
            fg=GOLD,
            activebackground=BUTTON_ACTIVE,
            activeforeground=GOLD,
            relief="flat",
            bd=0,
            cursor="hand2",
            command=self._toggle_pause,
            highlightbackground=GOLD,
            highlightthickness=2,
        )

        self.stop_btn = tk.Button(
            self.root,
            text="STOP",
            font=(FONT_FACE, 12, "bold"),
            bg=BUTTON_BG,
            fg=RED,
            activebackground=BUTTON_ACTIVE,
            activeforeground=RED,
            relief="flat",
            bd=0,
            cursor="hand2",
            command=self._end_session,
            highlightbackground=RED,
            highlightthickness=2,
        )
        self._layout_controls()
        self._update()
        self.root.after(10000, self._auto_minimize)

    def _apply_geometry(self):
        if not self.root:
            return
        width = self.MINI_W if self.minimized else self.W
        height = self.MINI_H if self.minimized else self.H
        x = self._window_x if self._window_x is not None else self.root.winfo_x()
        y = self._window_y if self._window_y is not None else self.root.winfo_y()
        self.root.geometry(f"{width}x{height}+{x}+{y}")
        if self.canvas:
            self.canvas.configure(width=width, height=height)

    def _layout_controls(self):
        if not self.root or not self.min_btn or not self.pause_btn or not self.stop_btn:
            return
        if self.minimized:
            self.min_btn.configure(text="OPEN")
            self.min_btn.place(x=self.MINI_W - 60, y=10, width=48, height=28)
            self.pause_btn.place_forget()
            self.stop_btn.place_forget()
        else:
            self.min_btn.configure(text="MIN")
            self.min_btn.place(x=236, y=16, width=56, height=28)
            self.pause_btn.place(x=296, y=16, width=72, height=28)
            self.stop_btn.place(x=296, y=54, width=72, height=28)

    def _drag_start(self, event):
        self._drag_x = event.x_root - self.root.winfo_x()
        self._drag_y = event.y_root - self.root.winfo_y()
        self._press_x = event.x
        self._press_y = event.y
        self._dragging = False

    def _drag_move(self, event):
        if abs(event.x - self._press_x) > 8 or abs(event.y - self._press_y) > 8:
            self._dragging = True
        if self._dragging:
            self._window_x = event.x_root - self._drag_x
            self._window_y = event.y_root - self._drag_y
            self._apply_geometry()

    def _drag_release(self, _event):
        self._dragging = False

    def _toggle_minimize(self):
        self.minimized = not self.minimized
        play_habit_sound("goal_pick")
        self._apply_geometry()
        self._layout_controls()
        self._draw_once()

    def _auto_minimize(self):
        if not self._running or not self.root or self.minimized or self.auto_minimized_once:
            return
        self.auto_minimized_once = True
        self.minimized = True
        self._apply_geometry()
        self._layout_controls()
        self._draw_once()

    def _toggle_pause(self):
        paused_now = pomodoro_state.toggle_pause()
        play_habit_sound("pause" if paused_now else "resume")
        self._refresh_buttons()
        self._draw_once()

    def _end_session(self):
        self._running = False
        threading.Thread(target=deactivate_study_mode, daemon=True).start()

    def close(self):
        self._running = False
        try:
            if self.root:
                self.root.destroy()
        except Exception:
            pass
        self.root = None
        self.canvas = None
        self.min_btn = None
        self.pause_btn = None
        self.stop_btn = None

    def _refresh_buttons(self):
        if not self.pause_btn:
            return
        paused = pomodoro_state.snapshot()["paused"]
        pause_fg = GREEN if paused else GOLD
        self.pause_btn.configure(
            text="RESUME" if paused else "PAUSE",
            fg=pause_fg,
            activeforeground=pause_fg,
            highlightbackground=pause_fg,
        )

    def _render(self, state):
        width = self.MINI_W if self.minimized else self.W
        height = self.MINI_H if self.minimized else self.H
        subject_color = hex_to_rgb(subject_hex(self.subject))
        recommended = recommended_subject_for_day()
        recommended_color = hex_to_rgb(subject_hex(recommended))
        if self.subject in DAILY_LESSON_LIBRARY:
            context = lesson_context(self.subject)
            display = context["display"]
        else:
            display = {
                "topic": current_session_topic or self.subject,
                "finish_date": date.today(),
            }
        paused = state["paused"]
        phase = state["phase"]
        remaining = state["remaining"]
        progress = state["progress"]
        focus_seconds = state["focus_seconds"]
        session = state["session"]
        phase_color = (255, 170, 70) if phase == "break" else subject_color
        border_color = RED if paused else "#%02x%02x%02x" % phase_color

        image = Image.new("RGB", (width, height), hex_to_rgb(BG))
        draw = ImageDraw.Draw(image)
        for x in range(0, width, 24):
            draw.line([(x, 0), (x, height)], fill=(10, 16, 18), width=1)
        for y in range(0, height, 24):
            draw.line([(0, y), (width, y)], fill=(10, 16, 18), width=1)
        for pad in range(5, 0, -1):
            draw.rounded_rectangle([pad, pad, width - pad, height - pad], radius=12, outline=hex_to_rgb(border_color))
        draw.rounded_rectangle([0, 0, width - 1, height - 1], radius=12, outline=hex_to_rgb(border_color), width=1)
        draw.rounded_rectangle([0, 0, 7, height - 1], radius=4, fill=subject_color)

        if self.minimized:
            compact_timer = base.format_clock(remaining)
            compact_phase = "MOCK" if current_session_mode == "mock" else "BREAK" if phase == "break" else "PAUSED" if paused else "FOCUS"
            draw.text((16, 8), trim_text(self.subject, 11), fill=subject_color, font=base.get_font(10, True))
            draw.text((92, 8), f"D{journey_day_number()}/{JOURNEY_TOTAL_DAYS}", fill=(255, 215, 0), font=base.get_font(9, True))
            draw.text((16, 27), compact_timer, fill=phase_color if not paused else (255, 90, 90), font=base.get_font(18, True))
            draw.text((118, 30), compact_phase, fill=(210, 230, 225), font=base.get_font(9, True))
            return image

        draw.text((16, 8), trim_text(self.subject, 14), fill=subject_color, font=base.get_font(10, True))
        draw.text((102, 8), f"DAY {journey_day_number()}/{JOURNEY_TOTAL_DAYS}", fill=GOLD, font=base.get_font(9, True))
        draw.text((212, 8), f"REC: {trim_text(recommended, 8)}", fill=recommended_color, font=base.get_font(9, True))

        status_text = "PAUSED" if paused else "BREAK" if phase == "break" else "FOCUS"
        if current_session_mode == "mock" and not paused:
            status_text = "MOCK"
        status_color = (255, 80, 80) if paused else (255, 170, 70) if phase == "break" else (0, 220, 100)
        draw.text((16, 26), status_text, fill=status_color, font=base.get_font(9, True))
        draw.text((102, 26), trim_text(timer_deadline_label(display.get("finish_date")), 24), fill=(120, 200, 170), font=base.get_font(9))
        rhythm = "MOCK EXAM | 120 MIN" if current_session_mode == "mock" else f"POMODORO {session} | 25/5"
        draw.text((16, 38), rhythm, fill=(110, 190, 170), font=base.get_font(8, True))

        timer = base.format_clock(remaining)
        timer_font = base.get_font(POPUP_TIMER_PT, True)
        timer_bbox = draw.textbbox((0, 0), timer, font=timer_font)
        timer_x = max(14, (294 - (timer_bbox[2] - timer_bbox[0])) // 2)
        draw.text((timer_x, 48), timer, fill=phase_color if not paused else (255, 90, 90), font=timer_font)

        topic = current_session_topic or display["topic"]
        draw.text((16, 80), trim_text(topic, 34), fill=(210, 230, 225), font=base.get_font(11, True))

        bar_x1, bar_y1, bar_x2, bar_y2 = 16, height - 8, 292, height - 4
        draw.rounded_rectangle([bar_x1, bar_y1, bar_x2, bar_y2], radius=2, fill=(18, 35, 22))
        if progress > 0:
            draw.rounded_rectangle(
                [bar_x1, bar_y1, bar_x1 + int((bar_x2 - bar_x1) * progress), bar_y2],
                radius=2,
                fill=phase_color,
            )
        return image

    def _draw_once(self):
        if not self.root or not self.canvas:
            return
        self.canvas.delete("all")
        image = self._render(pomodoro_state.snapshot())
        self._tk_img = ImageTk.PhotoImage(image)
        self.canvas.create_image(0, 0, image=self._tk_img, anchor="nw")

    def _update(self):
        if not self._running or not self.root:
            return
        try:
            self._refresh_buttons()
            self._draw_once()
            self.root.after(250, self._update)
        except Exception:
            pass


study_timer = StudyTimer()


def pomodoro_loop():
    while study_active and running and pomodoro_state.running:
        if pomodoro_state.stop_event.wait(1):
            return
        transition = pomodoro_state.advance_if_needed()
        if not transition:
            continue
        if transition["new_phase"] == "complete":
            play_habit_sound("success")
            notification.notify(
                title="Mock Exam Complete",
                message="Time is up. Log your subject scores now.",
                timeout=10,
            )
            threading.Thread(target=deactivate_study_mode, daemon=True).start()
            return
        if transition["new_phase"] == "break":
            pause_study_browser_for_break()
            play_habit_sound("break_start")
            base.play_voice("break")
            notification.notify(
                title="Break Time",
                message=f"Pomodoro {transition['finished_session']} complete. Take 5 minutes.",
                timeout=8,
            )
        else:
            restore_focus_after_break()
            play_habit_sound("break_end")
            base.play_voice("resume")
            notification.notify(
                title="Back to Work",
                message=f"Break finished. Pomodoro {transition['next_session']} starts now.",
                timeout=5,
            )


def make_tray_icon(active=False):
    image = Image.new("RGB", (64, 64), (10, 12, 14))
    draw = ImageDraw.Draw(image)
    glow = (0, 200, 140) if active else (80, 80, 80)
    draw.ellipse([8, 8, 56, 56], fill=glow)
    draw.ellipse([16, 16, 48, 48], fill=(5, 6, 8))
    draw.rectangle([30, 18, 34, 44], fill=glow)
    draw.rectangle([22, 26, 42, 30], fill=glow)
    return image


def update_tray():
    if not tray_icon:
        return
    tray_icon.icon = make_tray_icon(study_active)
    tray_icon.title = "Study Mode - ACTIVE" if study_active else "Study Mode - Press F9"


def open_excel_report():
    if not os.path.exists(EXCEL_FILE):
        if not export_to_excel(load_streak()):
            return
    try:
        subprocess.Popen(["cmd", "/c", "start", "", EXCEL_FILE])
    except Exception:
        notify_excel_issue("Could not open study_tracker.xlsx.")


def show_mock_start_popup(callback):
    accent = subject_hex("MOCK EXAM")
    root, panel = make_readable_popup(
        "MOCK EXAM MODE",
        780,
        430,
        accent,
        subtitle="2-hour full IOE practice block",
    )
    tk.Label(
        panel,
        text="Run this like the real exam: no pomodoro breaks, no pausing unless absolutely needed.",
        font=(FONT_FACE, 14, "bold"),
        fg=POPUP_TEXT,
        bg=POPUP_BG,
        wraplength=680,
        justify="center",
    ).pack(pady=(18, 16))

    section = popup_section(panel, "TARGET MARKS", accent)
    for line in ["Mathematics 50 | Physics 40 | Chemistry 30 | English 20", "Total: 140 marks | Time: 120 minutes"]:
        tk.Label(
            section,
            text=line,
            font=(FONT_FACE, POPUP_BODY_PT, "bold"),
            fg=POPUP_TEXT,
            bg=POPUP_PANEL,
            justify="center",
        ).pack(fill="x", padx=16, pady=4)

    buttons = tk.Frame(panel, bg=POPUP_BG)
    buttons.pack(fill="x", padx=28, pady=(10, 24))

    def start():
        root.destroy()
        callback()

    popup_button(buttons, "START 2H MOCK", accent, start).pack(side="left", expand=True, fill="x", padx=(0, 8))
    popup_button(buttons, "OPEN MCQS", CYAN, lambda: safe_open_url(MCQ_PRACTICE_URL)).pack(side="left", expand=True, fill="x", padx=8)
    popup_button(buttons, "CANCEL", RED, root.destroy).pack(side="left", expand=True, fill="x", padx=(8, 0))
    root.wait_window()


def show_mock_score_popup(minutes, callback):
    accent = subject_hex("MOCK EXAM")
    completed = {"done": False}
    root, panel = make_readable_popup(
        "MOCK SCORE",
        780,
        560,
        accent,
        subtitle=f"{minutes} minutes logged",
    )
    tk.Label(
        panel,
        text="Enter the marks you scored. Blank means 0.",
        font=(FONT_FACE, 14, "bold"),
        fg=POPUP_TEXT,
        bg=POPUP_BG,
    ).pack(pady=(12, 14))

    form = popup_section(panel, "SUBJECT SCORES", accent)
    entries = {}
    for subject, max_marks in MOCK_EXAM_MARKS.items():
        row = tk.Frame(form, bg=POPUP_PANEL)
        row.pack(fill="x", padx=16, pady=5)
        tk.Label(
            row,
            text=f"{subject} / {max_marks}",
            font=(FONT_FACE, POPUP_BODY_PT, "bold"),
            fg=subject_hex(subject),
            bg=POPUP_PANEL,
            anchor="w",
        ).pack(side="left", fill="x", expand=True)
        entry = tk.Entry(
            row,
            font=(FONT_FACE, POPUP_BODY_PT, "bold"),
            bg=BUTTON_BG,
            fg=POPUP_TEXT,
            insertbackground=CYAN,
            relief="flat",
            justify="center",
            width=8,
        )
        entry.insert(0, "0")
        entry.pack(side="right", padx=(14, 0), ipady=6)
        entries[subject] = entry

    error_label = tk.Label(panel, text="", font=(FONT_FACE, 11, "bold"), fg=RED, bg=POPUP_BG)
    error_label.pack(pady=(0, 8))

    def save_scores():
        scores = {}
        for subject, entry in entries.items():
            raw = entry.get().strip() or "0"
            try:
                score = int(float(raw))
            except ValueError:
                error_label.configure(text=f"Use numbers only for {subject}.")
                return
            max_marks = MOCK_EXAM_MARKS[subject]
            if score < 0 or score > max_marks:
                error_label.configure(text=f"{subject} must be between 0 and {max_marks}.")
                return
            scores[subject] = score
        completed["done"] = True
        root.destroy()
        callback(scores)

    popup_button(panel, "SAVE MOCK SCORE", accent, save_scores).pack(fill="x", padx=28, pady=(0, 24))
    root.wait_window()
    if not completed["done"]:
        callback({subject: 0 for subject in MOCK_EXAM_MARKS})


def start_recommended_session():
    if study_active:
        notification.notify(title="Study Mode", message="A session is already running.", timeout=4)
        return
    do_activate(recommended_subject_for_day())


def start_mock_exam():
    if study_active:
        notification.notify(title="Mock Exam", message="Stop the current session before starting a mock.", timeout=5)
        return
    base.run_on_main_thread(lambda: show_mock_start_popup(do_activate_mock_final))


def do_activate_mock_final():
    global advance_after_session, break_window_snapshot, current_resource_url, study_active, study_start_time, current_subject, current_session_goal, current_session_topic, current_session_mode

    current_session_mode = "mock"
    current_subject = "MOCK EXAM"
    current_session_goal = "Full IOE mock exam"
    current_session_topic = "Full IOE Mock Exam"
    current_resource_url = MCQ_PRACTICE_URL
    break_window_snapshot = set()
    advance_after_session = False
    study_active = True
    study_start_time = time.time()

    safe_open_url(MCQ_PRACTICE_URL)
    play_habit_sound("session_start")
    pomodoro_state.start(work_secs=MOCK_EXAM_SECONDS, break_secs=0, breaks_enabled=False)
    threading.Thread(target=pomodoro_loop, daemon=True).start()
    base.run_on_main_thread(lambda: study_timer.start("MOCK EXAM"))
    update_tray()
    notification.notify(title="Mock Exam Started", message="120 minutes. Treat it like the real exam.", timeout=6)


def do_activate_final(subject):
    global break_window_snapshot, current_resource_url, current_session_day_number, study_active, study_start_time, current_subject, current_session_topic, current_session_mode

    current_session_mode = "study"
    current_subject = subject
    study_active = True
    study_start_time = time.time()
    display = lesson_context(subject)["display"]
    sitting = sitting_finish_line(subject, display, adaptive_routine(subject, display))
    current_session_day_number = display["day_number"]
    current_session_topic = current_session_topic or display["topic"]
    subject_url = SUBJECT_URLS.get(subject, SUBJECT_URLS["General"])
    resource_url = display.get("resource", subject_url)
    current_resource_url = resource_url
    break_window_snapshot = set()

    base.close_edge()
    safe_open_url(LOFI_URL)
    time.sleep(2)
    safe_open_url(resource_url)

    threading.Thread(target=lambda: base.play_voice(f"start_{random.randint(0, 3)}"), daemon=True).start()
    play_habit_sound("session_start")
    pomodoro_state.start()
    threading.Thread(target=pomodoro_loop, daemon=True).start()
    base.run_on_main_thread(lambda: study_timer.start(subject))
    update_tray()

    notification.notify(
        title=f"Day {display['day_number']}/{JOURNEY_TOTAL_DAYS} - {subject}",
        message=f"Complete: {sitting['finish_line']}",
        timeout=5,
    )


def do_activate(subject):
    base.run_on_main_thread(lambda: show_today_plan(subject, lambda: do_activate_final(subject)))


def activate_study_mode():
    hour = datetime.now().hour
    if hour >= LATE_NIGHT_H:
        def after_warn(proceed):
            if proceed:
                base.run_on_main_thread(lambda: show_subject_selector(do_activate))

        base.run_on_main_thread(lambda: show_late_night_warning(hour, after_warn))
    else:
        base.run_on_main_thread(lambda: show_subject_selector(do_activate))


def do_stop():
    global advance_after_session, break_window_snapshot, current_resource_url, current_session_day_number, current_session_goal, current_session_topic, current_subject, current_session_mode, study_active, study_start_time

    study_active = False
    minutes = max(1, int(get_effective_study_seconds() / 60))
    study_start_time = None
    subject = current_subject or recommended_subject_for_day()
    goal = current_session_goal
    topic = current_session_topic or subject
    mode = current_session_mode

    pomodoro_state.stop()
    base.close_brave()
    base.run_on_main_thread(study_timer.close)
    update_tray()

    if mode == "mock":
        def after_scores(scores):
            global advance_after_session, break_window_snapshot, current_resource_url, current_session_day_number, current_session_goal, current_session_topic, current_subject, current_session_mode

            mock_data = record_mock_exam(minutes, scores)
            total = mock_data.get("mock_exams", [{}])[-1].get("total", 0)
            max_total = mock_data.get("mock_exams", [{}])[-1].get("max_total", 140)
            update_streak(minutes, "mock", "MOCK EXAM", goal, topic, "", "")
            play_habit_sound("success" if total >= int(max_total * 0.7) else "steady")
            notification.notify(
                title=f"Mock saved: {total}/{max_total}",
                message=f"Today's win: You completed a serious IOE practice block.",
                timeout=10,
            )
            current_session_goal = ""
            current_session_topic = ""
            current_session_day_number = 1
            current_resource_url = ""
            break_window_snapshot = set()
            current_subject = None
            current_session_mode = "study"
            advance_after_session = False

        base.run_on_main_thread(lambda: show_mock_score_popup(minutes, after_scores))
        return

    def after_wrapup(mood, notes, difficulty):
        global advance_after_session, break_window_snapshot, current_resource_url, current_session_day_number, current_session_goal, current_session_topic, current_subject, current_session_mode

        data = update_streak(minutes, mood, subject, goal, topic, notes, difficulty)
        if advance_after_session:
            save_day_progress(max(get_day_progress(), current_session_day_number))
        messages = {
            "crushed": f"That's {data['streak']} days straight. Outstanding.",
            "okay": f"Consistency beats intensity. {data['streak']} day streak.",
            "struggled": f"Hard days still count. {data['streak']} day streak.",
        }
        win_line = f"Today's win: You protected {minutes} minute(s) for {subject}."

        def play_end():
            if mood == "crushed":
                play_habit_sound("success")
                time.sleep(1.8)
                base.play_voice("end_crushed")
            elif mood == "okay":
                play_habit_sound("steady")
                time.sleep(0.8)
                base.play_voice("end_okay")
            else:
                play_habit_sound("gentle_end")
                time.sleep(0.6)
                base.play_voice("end_struggled")

        threading.Thread(target=play_end, daemon=True).start()
        notification.notify(title=f"{minutes} min - {subject}", message=f"{messages.get(mood, '')}\n{win_line}", timeout=10)
        current_session_goal = ""
        current_session_topic = ""
        current_session_day_number = 1
        current_resource_url = ""
        break_window_snapshot = set()
        current_subject = None
        current_session_mode = "study"
        advance_after_session = False

    base.run_on_main_thread(lambda: show_rating_popup(minutes, subject, goal, topic, after_wrapup))


def deactivate_study_mode():
    if not study_active:
        return
    elapsed = int(get_effective_study_seconds() / 60)
    if elapsed < (MIN_SESSION // 60):
        def after_warn(stop_now):
            if stop_now:
                do_stop()

        base.run_on_main_thread(lambda: show_min_session_warning(elapsed, after_warn))
    else:
        do_stop()


def shutdown_app():
    global running
    running = False
    pomodoro_state.stop()
    try:
        if tray_icon:
            tray_icon.stop()
    except Exception:
        pass
    try:
        _tk_root.quit()
        _tk_root.destroy()
    except Exception:
        pass
    os._exit(0)


def main():
    global _tk_root, tray_icon

    hide_console_window()
    keyboard.unhook_all_hotkeys()

    _tk_root = tk.Tk()
    _tk_root.withdraw()
    _tk_root.overrideredirect(True)
    _tk_root.attributes("-alpha", 0.0)

    base._tk_root = _tk_root
    _tk_root.after(100, base._process_tk_queue)
    ensure_alarm_scheduler()
    _tk_root.after(900, maybe_show_weekly_summary)

    keyboard.add_hotkey("f8", lambda: start_recommended_session() if not study_active else None)
    keyboard.add_hotkey("f9", lambda: deactivate_study_mode() if study_active else activate_study_mode())

    menu = pystray.Menu(
        pystray.MenuItem("Start Today's Recommended", lambda icon, item: start_recommended_session()),
        pystray.MenuItem("Start Study Mode", lambda icon, item: activate_study_mode() if not study_active else None),
        pystray.MenuItem("Start 2h Mock Exam", lambda icon, item: start_mock_exam()),
        pystray.MenuItem("Stop Study Mode", lambda icon, item: deactivate_study_mode() if study_active else None),
        pystray.Menu.SEPARATOR,
        pystray.MenuItem("Weak Topics", lambda icon, item: base.run_on_main_thread(show_weak_topics_popup)),
        pystray.MenuItem("Exam Reference", lambda icon, item: base.run_on_main_thread(show_exam_reference_popup)),
        pystray.MenuItem("Sound: Toggle Off/Soft/Strong", lambda icon, item: cycle_sound_mode()),
        pystray.MenuItem("IOE Countdown", lambda icon, item: base.run_on_main_thread(show_ioe_countdown)),
        pystray.MenuItem("Practice MCQs", lambda icon, item: safe_open_url(MCQ_PRACTICE_URL)),
        pystray.MenuItem("Open Excel Report", lambda icon, item: open_excel_report()),
        pystray.Menu.SEPARATOR,
        pystray.MenuItem("Exit", lambda icon, item: shutdown_app()),
    )
    tray_icon = pystray.Icon("StudyModeKrish", make_tray_icon(False), "Study Mode - F8 recommended, F9 menu", menu)
    threading.Thread(target=tray_icon.run, daemon=True).start()
    _tk_root.mainloop()


if __name__ == "__main__":
    main()

