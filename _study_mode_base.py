"""
Study Mode — Krish Baral | Full System + IOE Curriculum Guide
=============================================================
F9          → Start / Stop study mode
Features    → Subject selector, Today's Topic popup, IOE 6-month curriculum,
              Topic progress tracker, Late night warning, Min session warning,
              IOE countdown, Pomodoro 25/5, Voice, Streak, Excel log
"""

import time, threading, os, json, subprocess, random, math, queue
os.environ.setdefault("PYGAME_HIDE_SUPPORT_PROMPT", "1")
from datetime import datetime, date
from plyer import notification
import pystray
from PIL import Image, ImageDraw, ImageFont, ImageFilter, ImageTk
import keyboard
import tkinter as tk
import pygame

# ══════════════════════════════════════════
#  SETTINGS
# ══════════════════════════════════════════
LOFI_URL       = "https://www.youtube.com/watch?v=jfKfPfyJRdk"
BRAVE          = r"C:\Program Files\BraveSoftware\Brave-Browser\Application\brave.exe"

SUBJECT_URLS = {
    "MATHEMATICS": "https://www.khanacademy.org/math",
    "PHYSICS":     "https://www.youtube.com/c/ProfessorLeonard/playlists",
    "CHEMISTRY":   "https://www.youtube.com/@TheOrganicChemistryTutor/playlists",
    "ENGLISH":     "https://www.khanacademy.org/humanities/grammar",
    "General":     "https://www.youtube.com/@3blue1brown/videos",
}
BASE_DIR       = os.path.dirname(os.path.abspath(__file__))
STREAK_FILE    = os.path.join(BASE_DIR, "study_streak.json")
EXCEL_FILE     = os.path.join(BASE_DIR, "study_tracker.xlsx")
VOICES_DIR     = os.path.join(BASE_DIR, "study_voices")
POMODORO_WORK  = 25 * 60
POMODORO_BREAK =  5 * 60
MIN_SESSION    = 10 * 60
LATE_NIGHT_H   = 23
IOE_DATE       = date(2026, 7, 1)

BG = "#050608"; CYAN = "#00ffe7"; GREEN = "#00c878"
GOLD = "#ffd700"; RED = "#ff4444"; DIM = "#0d1810"
WHITE = "#e8f0ee"; GREY = "#507060"

# ══════════════════════════════════════════
#  IOE 6-MONTH CURRICULUM
#  Month 1-2: Basics | Month 3-4: Core Topics
#  Month 5: Practice | Month 6: Revision
# ══════════════════════════════════════════
CURRICULUM = {
    "MATHEMATICS": [
        # Month 1 — Foundations
        {"week": 1, "topic": "Set & Real Numbers",        "detail": "Sets, intervals, absolute value, logic connectives", "resource": "https://www.khanacademy.org/math/algebra/x2f8bb11595b61c86:foundation-algebra"},
        {"week": 1, "topic": "Logic & Functions",          "detail": "Function types: injective, surjective, bijective", "resource": "https://www.khanacademy.org/math/algebra/x2f8bb11595b61c86:functions"},
        {"week": 2, "topic": "Matrices & Determinants",    "detail": "Types, properties, inverse of matrix", "resource": "https://www.khanacademy.org/math/precalculus/x9e81a4f98389efdf:matrices"},
        {"week": 2, "topic": "Complex Numbers",            "detail": "Complex number operations, polynomial equations", "resource": "https://www.khanacademy.org/math/precalculus/x9e81a4f98389efdf:complex"},
        {"week": 3, "topic": "Sequence & Series",          "detail": "AP, GP, permutation, combination", "resource": "https://www.khanacademy.org/math/precalculus/x9e81a4f98389efdf:series"},
        {"week": 3, "topic": "Binomial Theorem",           "detail": "Binomial expansion, exponential & logarithmic series", "resource": "https://www.khanacademy.org/math/precalculus"},
        # Month 2 — Trigonometry & Coordinate Geometry
        {"week": 4, "topic": "Trigonometric Equations",    "detail": "Trig equations, general values, inverse trig functions", "resource": "https://www.khanacademy.org/math/trigonometry"},
        {"week": 4, "topic": "Properties of Triangles",    "detail": "In-centre, ortho-centre, circum-centre, solution of triangles", "resource": "https://www.khanacademy.org/math/trigonometry"},
        {"week": 5, "topic": "Straight Lines & Pair",      "detail": "Straight lines, pair of lines equations", "resource": "https://www.khanacademy.org/math/geometry"},
        {"week": 5, "topic": "Circles",                    "detail": "Circle equations, tangent, normal", "resource": "https://www.khanacademy.org/math/geometry/hs-geo-circles"},
        {"week": 6, "topic": "Conic Sections",             "detail": "Parabola, Ellipse, Hyperbola — standard equations", "resource": "https://www.khanacademy.org/math/precalculus/x9e81a4f98389efdf:conics"},
        {"week": 6, "topic": "3D Coordinates & Planes",    "detail": "Coordinates in space, plane equations", "resource": "https://www.youtube.com/results?search_query=3D+coordinate+geometry+IOE"},
        # Month 3 — Calculus
        {"week": 7, "topic": "Limits & Continuity",        "detail": "Limits, indeterminate forms, L'Hospital's rule", "resource": "https://www.khanacademy.org/math/calculus-1/cs1-limits-and-continuity"},
        {"week": 8, "topic": "Derivatives",                "detail": "Rules of derivatives, geometrical meaning, higher order", "resource": "https://www.khanacademy.org/math/calculus-1/cs1-derivatives-definition-and-basic-rules"},
        {"week": 9, "topic": "Applications of Derivatives","detail": "Tangent/normal, rate of change, maxima & minima", "resource": "https://www.khanacademy.org/math/calculus-1/cs1-applications-of-derivatives"},
        {"week": 10,"topic": "Integration",                "detail": "Rules, standard integrals, definite integral, area", "resource": "https://www.khanacademy.org/math/calculus-1/cs1-integrals"},
        {"week": 11,"topic": "Differential Equations",     "detail": "Variable separable, homogeneous, linear, exact DEs", "resource": "https://www.khanacademy.org/math/differential-equations"},
        # Month 4 — Vectors, Stats, Probability
        {"week": 12,"topic": "Vectors in Plane & Space",   "detail": "Vector algebra, linear combination, dependence", "resource": "https://www.khanacademy.org/math/linear-algebra"},
        {"week": 13,"topic": "Products of Vectors",        "detail": "Scalar product, vector product, scalar triple product", "resource": "https://www.youtube.com/results?search_query=vector+products+IOE+entrance"},
        {"week": 14,"topic": "Statistics",                 "detail": "Measures of location, dispersion, correlation, regression", "resource": "https://www.khanacademy.org/math/statistics-probability"},
        {"week": 15,"topic": "Probability",                "detail": "Conditional, compound, Bayes theorem, binomial distribution", "resource": "https://www.khanacademy.org/math/statistics-probability/probability-library"},
        # Month 5 — Practice
        {"week": 16,"topic": "PRACTICE — Set, Algebra",    "detail": "Solve past IOE MCQs on Set, Algebra, Complex Numbers", "resource": "https://www.youtube.com/results?search_query=IOE+entrance+math+MCQ+practice"},
        {"week": 17,"topic": "PRACTICE — Trig & Coord Geo","detail": "Solve past IOE MCQs on Trigonometry & Coordinate Geometry", "resource": "https://www.youtube.com/results?search_query=IOE+entrance+coordinate+geometry+MCQ"},
        {"week": 18,"topic": "PRACTICE — Calculus",        "detail": "Solve past IOE MCQs on Limits, Derivatives, Integration", "resource": "https://www.youtube.com/results?search_query=IOE+entrance+calculus+MCQ"},
        {"week": 19,"topic": "PRACTICE — Full Math Paper", "detail": "Attempt a full 50-mark math paper under timed conditions", "resource": "https://www.youtube.com/results?search_query=IOE+entrance+full+math+paper+solution"},
        # Month 6 — Revision
        {"week": 20,"topic": "REVISION — Weak Topics",     "detail": "Identify your 5 weakest topics and drill them hard", "resource": "https://www.khanacademy.org/math"},
        {"week": 21,"topic": "REVISION — Formula Sheet",   "detail": "Write all key formulas by hand. Memorise them.", "resource": "https://www.youtube.com/results?search_query=IOE+math+formula+sheet"},
        {"week": 22,"topic": "FINAL — Full Mock Test",     "detail": "Full timed mock exam. Treat it like the real thing.", "resource": "https://www.youtube.com/results?search_query=IOE+entrance+mock+test+2024"},
    ],
    "PHYSICS": [
        # Month 1
        {"week": 1, "topic": "Physical Quantities & Vectors","detail": "Dimensions, resolution, polygon law, vector algebra", "resource": "https://www.youtube.com/results?search_query=IOE+physics+vectors+kinematics"},
        {"week": 1, "topic": "Kinematics & Projectile",    "detail": "Equations of motion, projectile motion, relative motion", "resource": "https://www.khanacademy.org/science/physics/one-dimensional-motion"},
        {"week": 2, "topic": "Newton's Laws & Friction",    "detail": "Conservation of momentum, Newton's laws applications", "resource": "https://www.khanacademy.org/science/physics/forces-newtons-laws"},
        {"week": 2, "topic": "Work, Energy & Power",        "detail": "Work-energy theorem, conservation of energy, collisions", "resource": "https://www.khanacademy.org/science/physics/work-and-energy"},
        {"week": 3, "topic": "Circular Motion & Gravitation","detail": "Centripetal force, banking, satellite motion, SHM", "resource": "https://www.khanacademy.org/science/physics/circular-motion-and-gravitation"},
        {"week": 4, "topic": "Rotational Dynamics",         "detail": "Moment of inertia, torque, angular momentum", "resource": "https://www.youtube.com/results?search_query=rotational+dynamics+IOE+entrance+physics"},
        # Month 2
        {"week": 5, "topic": "Elasticity & Fluid Mechanics","detail": "Hooke's law, moduli, surface tension, Bernoulli", "resource": "https://www.khanacademy.org/science/physics/fluids"},
        {"week": 6, "topic": "Heat & Thermodynamics",       "detail": "Specific heat, thermal expansion, conduction, radiation", "resource": "https://www.khanacademy.org/science/physics/thermodynamics"},
        {"week": 7, "topic": "Laws of Thermodynamics",      "detail": "First/second law, Carnot cycle, entropy", "resource": "https://www.khanacademy.org/science/physics/thermodynamics"},
        # Month 3
        {"week": 8, "topic": "Geometric Optics",            "detail": "Mirrors, refraction, lenses, lens maker formula", "resource": "https://www.khanacademy.org/science/physics/geometric-optics"},
        {"week": 9, "topic": "Wave Optics & Light",         "detail": "Interference, diffraction, polarization", "resource": "https://www.youtube.com/results?search_query=wave+optics+IOE+entrance"},
        {"week": 10,"topic": "Waves & Sound",               "detail": "Wave motion, sound in pipes and strings, Doppler effect", "resource": "https://www.khanacademy.org/science/physics/mechanical-waves-and-sound"},
        # Month 4
        {"week": 11,"topic": "Electrostatics",              "detail": "Coulomb's law, Gauss law, capacitors, energy", "resource": "https://www.khanacademy.org/science/physics/electric-charge-electric-force-and-voltage"},
        {"week": 12,"topic": "DC Circuits",                 "detail": "Ohm's law, Kirchhoff's law, Joule's law", "resource": "https://www.khanacademy.org/science/physics/circuits-topic"},
        {"week": 13,"topic": "Magnetism & EM Induction",    "detail": "Biot-Savart, Faraday's law, transformer, AC circuits", "resource": "https://www.khanacademy.org/science/physics/magnetic-forces-and-magnetic-fields"},
        # Month 5
        {"week": 14,"topic": "Modern Physics",              "detail": "Photoelectric effect, Bohr's theory, X-rays, semiconductors", "resource": "https://www.khanacademy.org/science/physics/quantum-physics"},
        {"week": 15,"topic": "Radioactivity & Nuclear",     "detail": "Fission, fusion, radioactive decay, mass-energy relation", "resource": "https://www.khanacademy.org/science/physics/quantum-physics"},
        {"week": 16,"topic": "PRACTICE — Mechanics MCQ",   "detail": "Solve past IOE MCQs on Mechanics chapters", "resource": "https://www.youtube.com/results?search_query=IOE+physics+MCQ+mechanics"},
        {"week": 17,"topic": "PRACTICE — EM & Modern MCQ", "detail": "Solve past IOE MCQs on Electromagnetism & Modern Physics", "resource": "https://www.youtube.com/results?search_query=IOE+physics+MCQ+electromagnetism"},
        # Month 6
        {"week": 18,"topic": "REVISION — Full Physics Mock","detail": "Attempt full 40-mark physics paper under timed conditions", "resource": "https://www.youtube.com/results?search_query=IOE+physics+full+paper+solution"},
    ],
    "CHEMISTRY": [
        # Month 1-2
        {"week": 1, "topic": "Chemical Arithmetic",         "detail": "Dalton's theory, stoichiometry, Avogadro hypothesis", "resource": "https://www.khanacademy.org/science/chemistry/chemical-reactions-stoichiome"},
        {"week": 2, "topic": "States of Matter",            "detail": "Gaseous, liquid and solid states — properties", "resource": "https://www.khanacademy.org/science/chemistry/states-of-matter-and-intermolecular-forces"},
        {"week": 3, "topic": "Atomic Structure & Periodic Table","detail": "Quantum numbers, electronic config, periodic trends", "resource": "https://www.khanacademy.org/science/chemistry/electronic-structure-of-atoms"},
        {"week": 4, "topic": "Oxidation, Reduction & Equilibrium","detail": "Redox reactions, equilibrium constants", "resource": "https://www.khanacademy.org/science/chemistry/chemical-equilibrium"},
        {"week": 5, "topic": "Acid, Base & Salt",           "detail": "Ionic equilibrium, pH, buffer solutions", "resource": "https://www.khanacademy.org/science/chemistry/acid-base-equilibrium"},
        {"week": 6, "topic": "Electrochemistry",            "detail": "Electrolysis, Faraday's law, galvanic cells", "resource": "https://www.khanacademy.org/science/chemistry/oxidation-reduction"},
        # Month 3
        {"week": 7, "topic": "Chemical Energetics & Kinetics","detail": "Enthalpy, reaction rates, activation energy", "resource": "https://www.khanacademy.org/science/chemistry/thermodynamics-chemistry"},
        {"week": 8, "topic": "Chemical Bonding",            "detail": "Ionic, covalent, VSEPR, molecular shape", "resource": "https://www.khanacademy.org/science/chemistry/chemical-bonds"},
        # Month 4
        {"week": 9, "topic": "Non-metals",                  "detail": "H, O, N, halogens, C, P, S and their compounds", "resource": "https://www.youtube.com/results?search_query=IOE+chemistry+non+metals"},
        {"week": 10,"topic": "Metals & Metallurgy",         "detail": "Metallurgical principles, alkali metals, coinage metals", "resource": "https://www.youtube.com/results?search_query=IOE+inorganic+chemistry+metals"},
        # Month 5
        {"week": 11,"topic": "Organic — Fundamentals",      "detail": "Nomenclature, isomerism, reaction mechanism basics", "resource": "https://www.khanacademy.org/science/organic-chemistry"},
        {"week": 12,"topic": "Hydrocarbons",                "detail": "Alkanes, alkenes, alkynes, aromatic hydrocarbons", "resource": "https://www.khanacademy.org/science/organic-chemistry/alkanes-cycloalkanes"},
        {"week": 13,"topic": "Functional Groups",           "detail": "Alcohols, aldehydes, ketones, carboxylic acids, amines", "resource": "https://www.khanacademy.org/science/organic-chemistry"},
        {"week": 14,"topic": "PRACTICE — Chem MCQ Full",   "detail": "Solve past IOE Chemistry MCQs across all chapters", "resource": "https://www.youtube.com/results?search_query=IOE+chemistry+MCQ+past+papers"},
        # Month 6
        {"week": 15,"topic": "REVISION — Reaction Chart",  "detail": "Make a chart of all key reactions. Memorise equations.", "resource": "https://www.youtube.com/results?search_query=IOE+chemistry+important+reactions"},
    ],
    "ENGLISH": [
        {"week": 1, "topic": "Tense & Concord",             "detail": "Sequence of tenses, subject-verb agreement", "resource": "https://www.khanacademy.org/humanities/grammar/parts-of-speech-the-verb"},
        {"week": 2, "topic": "Direct & Indirect Speech",    "detail": "Reporting verbs, tense back-shift rules", "resource": "https://www.khanacademy.org/humanities/grammar"},
        {"week": 3, "topic": "Sentence Transformation",     "detail": "Kinds of sentences, transformation practice", "resource": "https://www.youtube.com/results?search_query=sentence+transformation+IOE+english"},
        {"week": 4, "topic": "Conditionals & Structures",   "detail": "Zero, first, second, third conditionals", "resource": "https://www.khanacademy.org/humanities/grammar"},
        {"week": 5, "topic": "Active & Passive Voice",      "detail": "All tenses passive voice conversion", "resource": "https://www.khanacademy.org/humanities/grammar/parts-of-speech-the-verb/verb-varb/a/active-and-passive-voice-review-article"},
        {"week": 6, "topic": "Verbals & Parts of Speech",   "detail": "Infinitives, gerunds, participles, parts of speech", "resource": "https://www.khanacademy.org/humanities/grammar/parts-of-speech-the-verb"},
        {"week": 7, "topic": "Prepositions & Punctuation",  "detail": "Common preposition usage, punctuation rules", "resource": "https://www.khanacademy.org/humanities/grammar/punctuation"},
        {"week": 8, "topic": "Vocabulary & Idioms",         "detail": "Learn 20 new words/day, idiom practice", "resource": "https://www.youtube.com/results?search_query=IOE+english+vocabulary+idioms"},
        {"week": 9, "topic": "Phonetics",                   "detail": "Phonemes, vowels, consonants, syllable stress", "resource": "https://www.youtube.com/results?search_query=IOE+english+phonetics+stress"},
        {"week": 10,"topic": "Reading Comprehension",       "detail": "Practice passages, technical English reading", "resource": "https://www.youtube.com/results?search_query=IOE+english+reading+comprehension"},
        {"week": 11,"topic": "PRACTICE — English MCQ",     "detail": "Solve past IOE English MCQs, all grammar topics", "resource": "https://www.youtube.com/results?search_query=IOE+entrance+english+MCQ"},
    ],
}

# ══════════════════════════════════════════
#  STATE
# ══════════════════════════════════════════
study_active     = False
study_start_time = None
tray_icon        = None
running          = True
current_subject  = None

pygame.mixer.pre_init(44100, -16, 2, 512)
pygame.mixer.init()

# ══════════════════════════════════════════
#  AUDIO
# ══════════════════════════════════════════
def play_voice(name):
    try:
        path = os.path.join(VOICES_DIR, f"{name}.mp3")
        if not os.path.exists(path): return
        pygame.mixer.music.load(path)
        pygame.mixer.music.play()
    except: pass

def play_sound(effect):
    def _go():
        try:
            cmds = {
                "party":  "[console]::beep(523,80);[console]::beep(659,80);[console]::beep(784,80);[console]::beep(1047,150);[console]::beep(1568,200)",
                "chime":  "[console]::beep(880,100);[console]::beep(1108,100);[console]::beep(1320,200)",
                "gentle": "[console]::beep(440,150);[console]::beep(392,200)",
                "start":  "[console]::beep(523,60);[console]::beep(784,120)",
            }
            if effect in cmds:
                subprocess.run(["powershell","-ExecutionPolicy","Bypass","-Command",cmds[effect]],capture_output=True)
        except: pass
    threading.Thread(target=_go, daemon=True).start()

# ══════════════════════════════════════════
#  PILLOW HELPERS
# ══════════════════════════════════════════
def get_font(size, bold=False):
    paths = [
        f"C:/Windows/Fonts/cour{'bd' if bold else ''}.ttf",
        f"/usr/share/fonts/truetype/dejavu/DejaVuSansMono{'Bold' if bold else ''}.ttf",
    ]
    for p in paths:
        if os.path.exists(p): return ImageFont.truetype(p, size)
    return ImageFont.load_default()

def cx(draw, w, text, y, fnt, fill):
    bbox = draw.textbbox((0,0), text, font=fnt)
    tw = bbox[2]-bbox[0]
    draw.text(((w-tw)//2, y), text, fill=fill, font=fnt)

def make_base(W, H, seed=42):
    img = Image.new("RGB",(W,H),(5,6,8))
    draw = ImageDraw.Draw(img)
    for x in range(0,W,40): draw.line([(x,0),(x,H)],fill=(11,16,13),width=1)
    for y in range(0,H,40): draw.line([(0,y),(W,y)],fill=(11,16,13),width=1)
    random.seed(seed)
    for _ in range(35):
        x,y=random.randint(0,W),random.randint(0,H)
        br=random.randint(15,120)
        draw.ellipse([x,y,x+1,y+1],fill=(br,br,int(br*1.1)))
    border=Image.new("RGBA",(W,H),(0,0,0,0))
    bd=ImageDraw.Draw(border)
    for i in range(10,0,-1):
        a=int(160*(1-i/10))
        bd.rectangle([i,i,W-i,H-i],outline=(0,255,231,a),width=1)
    border=border.filter(ImageFilter.GaussianBlur(3))
    img=Image.alpha_composite(img.convert("RGBA"),border).convert("RGB")
    draw=ImageDraw.Draw(img)
    draw.rectangle([0,0,W-1,H-1],outline=(0,160,130),width=1)
    def corner(x1,y1,dx,dy):
        L=22
        draw.line([(x1,y1),(x1+L*dx,y1)],fill=(0,255,231),width=2)
        draw.line([(x1,y1),(x1,y1+L*dy)],fill=(0,255,231),width=2)
        draw.rectangle([x1-2,y1-2,x1+2,y1+2],fill=(0,255,231))
    corner(8,8,1,1); corner(W-8,8,-1,1)
    corner(8,H-8,1,-1); corner(W-8,H-8,-1,-1)
    return img, ImageDraw.Draw(img)

def format_clock(total_seconds, include_hours=False):
    total_seconds = max(0, int(total_seconds))
    hrs, rem = divmod(total_seconds, 3600)
    mins, secs = divmod(rem, 60)
    if include_hours or hrs:
        return f"{hrs:02d}:{mins:02d}:{secs:02d}"
    return f"{mins:02d}:{secs:02d}"

# ══════════════════════════════════════════
#  TK MAIN THREAD QUEUE
# ══════════════════════════════════════════
_tk_queue = queue.Queue()
_tk_root  = None

def run_on_main_thread(fn):
    _tk_queue.put(fn)

def _process_tk_queue():
    while not _tk_queue.empty():
        try: _tk_queue.get_nowait()()
        except Exception as e: print(f"TK queue error: {e}")
    if _tk_root:
        _tk_root.after(100, _process_tk_queue)

# ══════════════════════════════════════════
#  CURRICULUM HELPERS
# ══════════════════════════════════════════
def get_topic_progress(subject):
    data = load_streak()
    progress = data.get("topic_progress", {})
    return progress.get(subject, 0)

def save_topic_progress(subject, idx):
    data = load_streak()
    if "topic_progress" not in data:
        data["topic_progress"] = {}
    data["topic_progress"][subject] = idx
    save_streak(data)

def get_current_topic(subject):
    topics = CURRICULUM.get(subject, [])
    if not topics: return None
    idx = get_topic_progress(subject)
    idx = min(idx, len(topics) - 1)
    return topics[idx], idx, len(topics)

def wrap_text(draw, text, fnt, max_w):
    """Wrap text to fit within max_w pixels."""
    words = text.split()
    lines = []
    current = ""
    for word in words:
        test = (current + " " + word).strip()
        bb = draw.textbbox((0,0), test, font=fnt)
        if bb[2] - bb[0] <= max_w:
            current = test
        else:
            if current: lines.append(current)
            current = word
    if current: lines.append(current)
    return lines

# ══════════════════════════════════════════
#  POPUP: TODAY'S STUDY PLAN
# ══════════════════════════════════════════
def show_today_plan(subject, on_start):
    topic_data, idx, total = get_current_topic(subject)
    if not topic_data:
        on_start(); return

    topic    = topic_data["topic"]
    detail   = topic_data["detail"]
    resource = topic_data["resource"]
    week     = topic_data["week"]
    pct_done = int((idx / total) * 100)

    subject_colors = {
        "MATHEMATICS": (0, 255, 231),
        "PHYSICS":     (80, 180, 255),
        "CHEMISTRY":   (160, 255, 80),
        "ENGLISH":     (255, 190, 80),
    }
    col = subject_colors.get(subject, (0, 255, 231))
    r, g, b = col

    W, H = 520, 380
    img, draw = make_base(W, H, 99)

    # Header
    cx(draw, W, "TODAY'S STUDY PLAN", 14, get_font(13, True), col)
    draw.line([(30, 40), (W-30, 40)], fill=(0, 60, 45), width=1)

    # Subject badge
    draw.rounded_rectangle([20, 50, 120, 68], radius=4, fill=col)
    sb = draw.textbbox((0,0), subject[:4], font=get_font(9, True))
    draw.text((20+(100-(sb[2]-sb[0]))//2, 53), subject[:4], fill=(5,6,8), font=get_font(9, True))

    # Week badge
    week_txt = f"WEEK {week}"
    draw.rounded_rectangle([130, 50, 230, 68], radius=4, fill=(18,35,22), outline=col, width=1)
    wb = draw.textbbox((0,0), week_txt, font=get_font(9, True))
    draw.text((130+(100-(wb[2]-wb[0]))//2, 53), week_txt, fill=col, font=get_font(9, True))

    # Progress badge
    prog_txt = f"{idx+1} / {total} topics"
    draw.rounded_rectangle([240, 50, 380, 68], radius=4, fill=(18,35,22), outline=(0,80,60), width=1)
    pb = draw.textbbox((0,0), prog_txt, font=get_font(8))
    draw.text((240+(140-(pb[2]-pb[0]))//2, 54), prog_txt, fill=(0,160,100), font=get_font(8))

    # Topic name
    draw.text((20, 82), "TOPIC", fill=(0, 100, 75), font=get_font(8, True))
    cx(draw, W, topic, 98, get_font(14, True), col)

    draw.line([(30, 122), (W-30, 122)], fill=(0, 50, 38), width=1)

    # Detail
    draw.text((20, 132), "WHAT TO STUDY", fill=(0, 100, 75), font=get_font(8, True))
    detail_lines = wrap_text(draw, detail, get_font(10), W - 40)
    for i, line in enumerate(detail_lines[:3]):
        draw.text((20, 150 + i*18), line, fill=(200, 230, 210), font=get_font(10))

    draw.line([(30, 210), (W-30, 210)], fill=(0, 50, 38), width=1)

    # How to study tips
    draw.text((20, 220), "HOW TO STUDY THIS", fill=(0, 100, 75), font=get_font(8, True))
    tips = [
        "① Watch 1 YouTube video on this topic first (20 min)",
        "② Take notes — write key formulas by hand",
        "③ Solve 10 practice MCQs from past IOE papers",
    ]
    for i, tip in enumerate(tips):
        draw.text((20, 238 + i*20), tip, fill=(140, 190, 150), font=get_font(9))

    draw.line([(30, 302), (W-30, 302)], fill=(0, 50, 38), width=1)

    # Progress bar
    draw.text((20, 312), f"SYLLABUS PROGRESS  {pct_done}%", fill=(0, 80, 60), font=get_font(8))
    bar_x1, bar_y1, bar_x2, bar_y2 = 20, 328, W-20, 340
    draw.rounded_rectangle([bar_x1, bar_y1, bar_x2, bar_y2], radius=4, fill=(15,28,18))
    fill_w = int((bar_x2 - bar_x1) * pct_done / 100)
    if fill_w > 0:
        fill_col = (0,200,100) if pct_done < 70 else (255,200,60) if pct_done < 90 else (255,80,80)
        draw.rounded_rectangle([bar_x1, bar_y1, bar_x1+fill_w, bar_y2], radius=4, fill=fill_col)

    # Buttons
    root = tk.Toplevel(_tk_root)
    root.overrideredirect(True)
    root.attributes('-topmost', True)
    root.attributes('-alpha', 0.0)
    root.configure(bg=BG)
    sw = root.winfo_screenwidth(); sh = root.winfo_screenheight()
    root.geometry(f"{W}x{H}+{(sw-W)//2}+{(sh-H)//2}")
    canvas = tk.Canvas(root, width=W, height=H, bg=BG, highlightthickness=0)
    canvas.pack()
    tk_img = ImageTk.PhotoImage(img)
    canvas.create_image(0, 0, image=tk_img, anchor="nw")

    def open_resource():
        try: subprocess.Popen([BRAVE, resource])
        except: pass

    def mark_done_start():
        # Mark topic complete, advance to next
        new_idx = min(idx + 1, total - 1)
        save_topic_progress(subject, new_idx)
        root.destroy()
        on_start()

    def just_start():
        root.destroy()
        on_start()

    # Resource button
    btn_res = tk.Button(root, text="📖  OPEN RESOURCE", font=("Courier New", 9, "bold"),
        bg=DIM, fg="#%02x%02x%02x" % col, activebackground="#1a2a1a",
        activeforeground="#%02x%02x%02x" % col, relief="flat", bd=0, cursor="hand2",
        command=open_resource, highlightbackground="#%02x%02x%02x" % col, highlightthickness=1)
    btn_res.place(x=16, y=H-46, width=150, height=34)

    # Start session (same topic)
    btn_start = tk.Button(root, text="▶  START SESSION", font=("Courier New", 9, "bold"),
        bg=DIM, fg="#00ffe7", activebackground="#1a2a1a", activeforeground="#00ffe7",
        relief="flat", bd=0, cursor="hand2", command=just_start,
        highlightbackground="#00ffe7", highlightthickness=1)
    btn_start.place(x=174, y=H-46, width=150, height=34)

    # Mark done & next topic
    btn_done = tk.Button(root, text="✓  DONE → NEXT TOPIC", font=("Courier New", 9, "bold"),
        bg=DIM, fg="#00c878", activebackground="#1a2a1a", activeforeground="#00c878",
        relief="flat", bd=0, cursor="hand2", command=mark_done_start,
        highlightbackground="#00c878", highlightthickness=1)
    btn_done.place(x=332, y=H-46, width=172, height=34)

    alpha = [0.0]
    def fade():
        if alpha[0] < 1.0:
            alpha[0] = min(1.0, alpha[0] + 0.1)
            root.attributes('-alpha', alpha[0])
            root.after(25, fade)
    fade()
    root.wait_window()

# ══════════════════════════════════════════
#  POPUP BASE
# ══════════════════════════════════════════
class Popup:
    def __init__(self, W, H, centered=True):
        self.W=W; self.H=H
        self.root=tk.Toplevel(_tk_root)
        self.root.overrideredirect(True)
        self.root.attributes('-topmost',True)
        self.root.attributes('-alpha',0.0)
        self.root.configure(bg=BG)
        sw=self.root.winfo_screenwidth(); sh=self.root.winfo_screenheight()
        if centered: self.root.geometry(f"{W}x{H}+{(sw-W)//2}+{(sh-H)//2}")
        else:        self.root.geometry(f"{W}x{H}+{sw-W-16}+56")
        self.canvas=tk.Canvas(self.root,width=W,height=H,bg=BG,highlightthickness=0)
        self.canvas.pack()
        self.root.bind("<Escape>",lambda e: self.root.destroy())
        self._a=0.0

    def _fade(self):
        if self._a<1.0:
            self._a=min(1.0,self._a+0.1)
            self.root.attributes('-alpha',self._a)
            self.root.after(25,self._fade)

    def btn(self, text, x, y, w, h, fg, cmd):
        b=tk.Button(self.root,text=text,font=("Courier New",10,"bold"),
            bg=DIM,fg=fg,activebackground="#1a2a1a",activeforeground=fg,
            relief="flat",bd=0,cursor="hand2",command=cmd,
            highlightbackground=fg,highlightthickness=1)
        b.place(x=x,y=y,width=w,height=h)

# ══════════════════════════════════════════
#  FLOATING TIMER WIDGET
# ══════════════════════════════════════════
class StudyTimer:
    W, H = 320, 80

    def __init__(self):
        self.root        = None
        self.canvas      = None
        self._tk_img     = None
        self.paused      = False
        self.pause_start = None
        self.paused_secs = 0
        self._running    = False
        self.subject     = "General"
        self._pause_rect = (0,0,0,0)
        self._stop_rect  = (0,0,0,0)

    def start(self, subject):
        self.subject     = subject
        self.paused      = False
        self.paused_secs = 0
        self._running    = True
        self._run_window()

    def _run_window(self):
        self.root = tk.Toplevel(_tk_root)
        self.root.overrideredirect(True)
        self.root.attributes('-topmost', True)
        self.root.attributes('-alpha', 0.93)
        self.root.configure(bg='#050608')
        sw = self.root.winfo_screenwidth()
        self.root.geometry(f"{self.W}x{self.H}+{sw-self.W-12}+8")
        self.canvas = tk.Canvas(self.root, width=self.W, height=self.H,
                                bg='#050608', highlightthickness=0, cursor="hand2")
        self.canvas.pack()
        self._drag_x = 0; self._drag_y = 0
        self._dragging = False
        self.canvas.bind("<ButtonPress-1>",   self._drag_start)
        self.canvas.bind("<B1-Motion>",        self._drag_move)
        self.canvas.bind("<ButtonRelease-1>",  self._on_release)
        self._update()

    def _drag_start(self, e):
        self._drag_x   = e.x_root - self.root.winfo_x()
        self._drag_y   = e.y_root - self.root.winfo_y()
        self._dragging = False
        self._press_x  = e.x
        self._press_y  = e.y

    def _drag_move(self, e):
        if abs(e.x - self._press_x) > 4 or abs(e.y - self._press_y) > 4:
            self._dragging = True
        if self._dragging:
            self.root.geometry(f"+{e.x_root-self._drag_x}+{e.y_root-self._drag_y}")

    def _on_release(self, e):
        if self._dragging:
            self._dragging = False
            return
        x, y = e.x, e.y
        px1, py1, px2, py2 = self._pause_rect
        sx1, sy1, sx2, sy2 = self._stop_rect
        if px1 <= x <= px2 and py1 <= y <= py2:
            self._toggle_pause()
        elif sx1 <= x <= sx2 and sy1 <= y <= sy2:
            self._end_session()

    def _toggle_pause(self):
        if not self.paused:
            self.paused      = True
            self.pause_start = time.time()
        else:
            self.paused = False
            if self.pause_start:
                self.paused_secs += int(time.time() - self.pause_start)
                self.pause_start = None

    def _end_session(self):
        self._running = False
        try:
            if self.root: self.root.destroy(); self.root = None
        except: pass
        threading.Thread(target=deactivate_study_mode, daemon=True).start()

    def close(self):
        self._running = False
        try:
            if self.root: self.root.destroy(); self.root = None
        except: pass

    def _render(self, net_secs, pom_secs):
        W, H = self.W, self.H
        acol = {
            "MATHEMATICS": (0,255,231),
            "PHYSICS":     (80,180,255),
            "CHEMISTRY":   (160,255,80),
            "ENGLISH":     (255,190,80),
        }.get(self.subject, (0,255,231))
        bcol = (255,80,80) if self.paused else (0,255,231)

        img  = Image.new("RGB",(W,H),(5,6,8))
        draw = ImageDraw.Draw(img)

        for x in range(0,W,20): draw.line([(x,0),(x,H)],fill=(9,13,10),width=1)
        for y in range(0,H,20): draw.line([(0,y),(W,y)],fill=(9,13,10),width=1)

        for i in range(4,0,-1):
            draw.rounded_rectangle([i,i,W-i,H-i],radius=10, outline=bcol)
        draw.rounded_rectangle([0,0,W-1,H-1],radius=10,outline=bcol,width=1)

        draw.rounded_rectangle([0,0,5,H-1],radius=4,fill=acol)
        draw.text((12,7), self.subject[:4], fill=acol, font=get_font(8,True))

        dcol = (255,80,80) if self.paused else (0,220,100)
        draw.ellipse([12,H-19,20,H-11], fill=dcol)
        draw.text((24,H-20), "PAUSED" if self.paused else "LIVE", fill=dcol, font=get_font(7,True))

        # Current topic name on timer
        topic_data, idx, total = get_current_topic(self.subject)
        if topic_data:
            short = topic_data["topic"][:22]
            draw.text((12, 20), short, fill=(0,130,90), font=get_font(7))

        hrs  = net_secs // 3600
        mins = (net_secs % 3600) // 60
        secs = net_secs % 60
        timer = f"{hrs:02d}:{mins:02d}:{secs:02d}"
        fnt_big = get_font(18, True)
        bb  = draw.textbbox((0,0), timer, font=fnt_big)
        tw  = bb[2]-bb[0]
        tx  = max(10, ((W-72)-tw)//2 + 8)
        draw.text((tx, 30), timer, fill=bcol, font=fnt_big)

        bx1,by1,bx2,by2 = 12, H-9, W-12, H-4
        draw.rounded_rectangle([bx1,by1,bx2,by2],radius=2,fill=(18,35,22))
        pct = min(1.0, pom_secs / (25*60))
        if pct > 0:
            draw.rounded_rectangle([bx1,by1,bx1+int((bx2-bx1)*pct),by2],radius=2,fill=acol)

        pom_min = pom_secs // 60
        pl  = f"POMO {pom_min}/25"
        draw.text((W-78, H-20), pl, fill=(0,100,75), font=get_font(7))

        PBX1, PBY1, PBX2, PBY2 = W-68, 28, W-36, 54
        pc = (255,160,40) if not self.paused else (0,200,100)
        draw.rounded_rectangle([PBX1,PBY1,PBX2,PBY2],radius=5,fill=(10,18,12),outline=pc,width=2)
        pl2  = "||" if not self.paused else "▶"
        fnt9 = get_font(10, True)
        bb3  = draw.textbbox((0,0), pl2, font=fnt9)
        draw.text((PBX1+(PBX2-PBX1-(bb3[2]-bb3[0]))//2, PBY1+4), pl2, fill=pc, font=fnt9)
        self._pause_rect = (PBX1, PBY1, PBX2, PBY2)

        SBX1, SBY1, SBX2, SBY2 = W-32, 28, W-4, 54
        draw.rounded_rectangle([SBX1,SBY1,SBX2,SBY2],radius=5,fill=(10,18,12),outline=(220,50,50),width=2)
        bb4 = draw.textbbox((0,0), "■", font=fnt9)
        draw.text((SBX1+(SBX2-SBX1-(bb4[2]-bb4[0]))//2, SBY1+4), "■", fill=(220,50,50), font=fnt9)
        self._stop_rect = (SBX1, SBY1, SBX2, SBY2)

        return img

    def _update(self):
        if not self._running or not self.root: return
        try:
            if study_start_time:
                raw        = int(time.time() - study_start_time)
                paused_now = int(time.time()-self.pause_start) if (self.paused and self.pause_start) else 0
                net_secs   = max(0, raw - self.paused_secs - paused_now)
                pom_secs   = net_secs % (25*60)
            else:
                net_secs = 0; pom_secs = 0

            img = self._render(net_secs, pom_secs)
            self._tk_img = ImageTk.PhotoImage(img)
            self.canvas.create_image(0, 0, image=self._tk_img, anchor="nw")
            self.root.after(1000, self._update)
        except: pass


study_timer = StudyTimer()

# ══════════════════════════════════════════
#  POPUP: SUBJECT SELECTOR
# ══════════════════════════════════════════
def show_subject_selector(callback):
    W,H=500,440
    img,draw=make_base(W,H,42)
    cx(draw,W,"SELECT  SUBJECT",14,get_font(15,True),(0,255,231))
    draw.line([(30,40),(W-30,40)],fill=(0,60,45),width=1)
    cx(draw,W,"What are you studying today, sir?",52,get_font(11),(100,160,130))

    def sym_math(d,ccx,cy,col):
        for i,l in enumerate([28,20,28]):
            yy=cy-12+i*12; d.line([(ccx-l//2,yy),(ccx+l//2,yy)],fill=col,width=2)
        d.line([(ccx-14,cy-12),(ccx-14,cy+12)],fill=col,width=2)

    def sym_physics(d,ccx,cy,col):
        d.ellipse([ccx-5,cy-5,ccx+5,cy+5],fill=col)
        d.ellipse([ccx-20,cy-8,ccx+20,cy+8],outline=col,width=1)
        pts=[(ccx+20*math.cos(math.radians(a)),cy+12*math.sin(math.radians(a))) for a in range(0,360,10)]
        for j in range(len(pts)-1): d.line([pts[j],pts[j+1]],fill=col,width=1)

    def sym_chemistry(d,ccx,cy,col):
        d.line([(ccx-5,cy-16),(ccx+5,cy-16)],fill=col,width=2)
        d.line([(ccx-5,cy-16),(ccx-14,cy+14)],fill=col,width=2)
        d.line([(ccx+5,cy-16),(ccx+14,cy+14)],fill=col,width=2)
        d.line([(ccx-14,cy+14),(ccx+14,cy+14)],fill=col,width=2)
        d.ellipse([ccx-4,cy+2,ccx+4,cy+10],fill=col)

    def sym_english(d,ccx,cy,col):
        d.line([(ccx,cy-14),(ccx,cy+14)],fill=col,width=2)
        d.arc([ccx-22,cy-14,ccx-2,cy+14],200,340,fill=col,width=2)
        d.arc([ccx+2,cy-14,ccx+22,cy+14],200,340,fill=col,width=2)

    sym_drawers=[sym_math,sym_physics,sym_chemistry,sym_english]
    subjects=[
        ("MATHEMATICS","Calculus · Algebra · Trig",    (0,255,231)),
        ("PHYSICS",    "Mechanics · Waves · EM",        (80,180,255)),
        ("CHEMISTRY",  "Organic · Inorganic · Physical",(160,255,80)),
        ("ENGLISH",    "Grammar · Reading · Writing",   (255,190,80)),
    ]
    pos=[(16,76),(258,76),(16,242),(258,242)]
    CW,CH=220,150
    last=load_streak().get("last_subject","")

    for i,(name,sub,col) in enumerate(subjects):
        x,y=pos[i]; r,g,b=col; ccx=x+CW//2
        is_last=(name==last)
        draw.rectangle([x,y,x+CW,y+CH],fill=(12,20,14) if is_last else (8,14,10))
        if is_last:
            for glow in range(4,0,-1):
                draw.rectangle([x-glow,y-glow,x+CW+glow,y+CH+glow],outline=(r//2,g//2,b//2),width=1)
            draw.rectangle([x,y,x+CW,y+CH],outline=col,width=3)
        else:
            draw.rectangle([x,y,x+CW,y+CH],outline=col,width=2)
        draw.rectangle([x+2,y+2,x+CW-2,y+6],fill=col)
        sym_drawers[i](draw,ccx,y+36,col)
        nb=draw.textbbox((0,0),name,font=get_font(14,True)); ntw=nb[2]-nb[0]
        draw.text((ccx-ntw//2,y+68),name,fill=col,font=get_font(14,True))
        draw.line([(x+20,y+94),(x+CW-20,y+94)],fill=(r//4,g//4,b//4+8),width=1)
        sb2=draw.textbbox((0,0),sub,font=get_font(9)); stw=sb2[2]-sb2[0]
        draw.text((ccx-stw//2,y+102),sub,fill=(r//2,g//2+20,50),font=get_font(9))

        # Show current topic on card
        topics = CURRICULUM.get(name, [])
        tidx = get_topic_progress(name)
        tidx = min(tidx, len(topics)-1)
        if topics:
            short = topics[tidx]["topic"][:18]
            tb = draw.textbbox((0,0), short, font=get_font(7))
            draw.text((ccx-(tb[2]-tb[0])//2, y+120), short, fill=(r//3,g//3,b//3+20), font=get_font(7))

        if is_last:
            badge="LAST"
            tl=draw.textbbox((0,0),badge,font=get_font(7,True)); tlw=tl[2]-tl[0]
            draw.rectangle([x+CW-tlw-12,y+CH-20,x+CW-2,y+CH-2],fill=col)
            draw.text((x+CW-tlw-8,y+CH-19),badge,fill=(5,6,8),font=get_font(7,True))

    draw.line([(30,404),(W-30,404)],fill=(0,55,42),width=1)
    last_text = f"Last: {last}  ·  Click subject → see today's plan  ·  ESC to skip" if last else "Click a subject to see today's study plan  ·  ESC to skip"
    cx(draw,W,last_text,414,get_font(9),(0,120,90))

    root=tk.Toplevel(_tk_root)
    root.overrideredirect(True)
    root.attributes('-topmost',True)
    root.attributes('-alpha',0.0)
    root.configure(bg=BG)
    sw=root.winfo_screenwidth(); sh=root.winfo_screenheight()
    root.geometry(f"{W}x{H}+{(sw-W)//2}+{(sh-H)//2}")
    canvas=tk.Canvas(root,width=W,height=H,bg=BG,highlightthickness=0,cursor="hand2")
    canvas.pack()
    tk_img=ImageTk.PhotoImage(img)
    canvas.create_image(0,0,image=tk_img,anchor="nw")

    def on_click(event):
        for i,(name,sub,col) in enumerate(subjects):
            x,y=pos[i]
            if x<=event.x<=x+CW and y<=event.y<=y+CH:
                root.destroy(); callback(name); return

    canvas.bind("<Button-1>", on_click)
    root.bind("<Escape>", lambda e: [root.destroy(), callback("General")])

    def on_motion(event):
        for i,(name,sub,col) in enumerate(subjects):
            x,y=pos[i]
            if x<=event.x<=x+CW and y<=event.y<=y+CH:
                canvas.config(cursor="hand2"); return
        canvas.config(cursor="")
    canvas.bind("<Motion>", on_motion)

    alpha=[0.0]
    def fade():
        if alpha[0]<1.0:
            alpha[0]=min(1.0,alpha[0]+0.1)
            root.attributes('-alpha',alpha[0])
            root.after(25,fade)
    fade()
    root.wait_window()

# ══════════════════════════════════════════
#  POPUP: LATE NIGHT WARNING
# ══════════════════════════════════════════
def show_late_night_warning(hour, callback):
    W,H=400,230
    img,draw=make_base(W,H,77)
    cx(draw,W,"LATE NIGHT WARNING",22,get_font(13,True),(255,160,60))
    draw.line([(30,48),(W-30,48)],fill=(0,60,45),width=1)
    cx(draw,W,f"Sir, it's after {hour}:00 right now.",64,get_font(11),(240,220,180))
    cx(draw,W,"Sleep matters for memory & focus.",88,get_font(10),(160,140,100))
    cx(draw,W,"IOE prep needs a rested brain.",108,get_font(10),(120,100,70))
    draw.line([(30,130),(W-30,130)],fill=(0,60,45),width=1)

    pw=Popup(W,H)
    pw._img=ImageTk.PhotoImage(img)
    pw.canvas.create_image(0,0,image=pw._img,anchor="nw")
    pw.btn("YES, LET'S GO",  30,144,170,36,CYAN, lambda:[pw.root.destroy(),callback(True)])
    pw.btn("NO, I'LL SLEEP",210,144,170,36,RED,  lambda:[pw.root.destroy(),callback(False)])
    tk.Label(pw.root,text="You can always override this.",
        font=("Courier New",8),bg=BG,fg="#305040").place(x=0,y=202,width=W)
    pw._fade(); pw.root.wait_window()

# ══════════════════════════════════════════
#  POPUP: MIN SESSION WARNING
# ══════════════════════════════════════════
def show_min_session_warning(elapsed_min, callback):
    W,H=400,215
    img,draw=make_base(W,H,33)
    cx(draw,W,"STOPPING EARLY?",22,get_font(13,True),(255,200,60))
    draw.line([(30,46),(W-30,46)],fill=(0,60,45),width=1)
    cx(draw,W,f"That was only {elapsed_min} minutes, sir.",62,get_font(11),(240,230,200))
    cx(draw,W,"Minimum effective session is 20 min.",84,get_font(10),(160,150,100))
    bar_w=W-80; fill_w=min(bar_w,int((elapsed_min/20)*bar_w))
    draw.rectangle([40,108,40+bar_w,122],fill=(15,25,18),outline=(0,60,45),width=1)
    col=(255,80,40) if elapsed_min<10 else (255,160,40) if elapsed_min<20 else (0,255,100)
    draw.rectangle([40,108,40+fill_w,122],fill=col)
    cx(draw,W,f"{elapsed_min} / 20 minutes",128,get_font(8),(80,100,70))
    draw.line([(30,146),(W-30,146)],fill=(0,60,45),width=1)

    pw=Popup(W,H)
    pw._img=ImageTk.PhotoImage(img)
    pw.canvas.create_image(0,0,image=pw._img,anchor="nw")
    pw.btn("KEEP GOING",  30,158,170,36,CYAN, lambda:[pw.root.destroy(),callback(False)])
    pw.btn("STOP ANYWAY",210,158,170,36,RED,  lambda:[pw.root.destroy(),callback(True)])
    pw._fade(); pw.root.wait_window()

# ══════════════════════════════════════════
#  POPUP: SESSION RATING
# ══════════════════════════════════════════
def show_rating_popup(minutes, subject, callback):
    W,H=360,250
    img,draw=make_base(W,H,55)
    cx(draw,W,"SESSION COMPLETE",18,get_font(13,True),(0,255,231))
    draw.line([(30,42),(W-30,42)],fill=(0,60,45),width=1)
    cx(draw,W,f"{minutes} minutes  ·  {subject}",56,get_font(10),(0,180,130))
    cx(draw,W,"How was your session, sir?",78,get_font(10),(200,220,200))
    draw.line([(30,100),(W-30,100)],fill=(0,60,45),width=1)

    pw=Popup(W,H)
    pw._img=ImageTk.PhotoImage(img)
    pw.canvas.create_image(0,0,image=pw._img,anchor="nw")
    moods=[("💪  CRUSHED IT","crushed",CYAN),("😐  IT WAS OKAY","okay",GOLD),("😴  STRUGGLED","struggled",RED)]
    for i,(label,val,col) in enumerate(moods):
        pw.btn(label,30,112+i*42,W-60,34,col,lambda v=val:[pw.root.destroy(),callback(v)])
    pw._fade(); pw.root.wait_window()

# ══════════════════════════════════════════
#  POPUP: IOE COUNTDOWN
# ══════════════════════════════════════════
def show_ioe_countdown():
    days=(IOE_DATE-date.today()).days
    data=load_streak()
    total_hrs=round(data.get("total_minutes",0)/60,1)
    W,H=400,200
    img,draw=make_base(W,H,11)
    cx(draw,W,"IOE ENTRANCE COUNTDOWN",18,get_font(12,True),(0,255,231))
    draw.line([(30,40),(W-30,40)],fill=(0,60,45),width=1)
    dcol=(0,255,231) if days>60 else (255,200,60) if days>30 else (255,80,80)
    cx(draw,W,str(days),46,get_font(52,True),dcol)
    cx(draw,W,"DAYS REMAINING",112,get_font(10),(0,140,100))
    draw.line([(30,134),(W-30,134)],fill=(0,60,45),width=1)
    cx(draw,W,f"Every session brings you closer, sir.",148,get_font(9),(80,140,100))
    cx(draw,W,f"Studied: {total_hrs}h total  ·  Streak: {data.get('streak',0)} days",164,get_font(9),(60,100,75))

    pw=Popup(W,H)
    pw._img=ImageTk.PhotoImage(img)
    pw.canvas.create_image(0,0,image=pw._img,anchor="nw")
    pw.btn("CLOSE",(W-80)//2,H-36,80,28,GREY,pw.root.destroy)
    pw._fade(); pw.root.wait_window()

# ══════════════════════════════════════════
#  STREAK / DATA
# ══════════════════════════════════════════
def load_streak():
    try:
        if os.path.exists(STREAK_FILE):
            with open(STREAK_FILE) as f: return json.load(f)
    except: pass
    return {"streak":0,"last_date":"","total_sessions":0,"total_minutes":0,"sessions":[],"topic_progress":{}}

def save_streak(data):
    with open(STREAK_FILE,"w") as f: json.dump(data,f,indent=2)

def update_streak(minutes, mood, subject):
    data=load_streak()
    today=str(date.today())
    yesterday=str(date.fromordinal(date.today().toordinal()-1))
    last=data.get("last_date","")
    if last==today: pass
    elif last==yesterday: data["streak"]=data.get("streak",0)+1
    else: data["streak"]=1
    data["last_date"]=today
    data["total_sessions"]=data.get("total_sessions",0)+1
    data["total_minutes"]=data.get("total_minutes",0)+minutes
    data["last_subject"]=subject
    data.setdefault("sessions",[]).append({
        "date":today,"minutes":minutes,"mood":mood,
        "subject":subject,"time":datetime.now().strftime("%H:%M")
    })
    save_streak(data)
    threading.Thread(target=lambda: export_to_excel(data),daemon=True).start()
    return data

# ══════════════════════════════════════════
#  EXCEL EXPORT
# ══════════════════════════════════════════
def export_to_excel(data):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        BX="050608";CY="00FFE7";GR="00C878";GO="FFD700";RE="FF4444";WH="E8F0EE";DM="0D1810"

        def sc(ws,r,c,val,bold=False,sz=10,col=WH,bg=BX,align="center"):
            cell=ws.cell(r,c,value=val)
            cell.font=Font(name="Courier New",bold=bold,size=sz,color=col)
            cell.alignment=Alignment(horizontal=align,vertical="center")
            cell.fill=PatternFill("solid",start_color=bg)

        wb=Workbook()
        ws=wb.active; ws.title="📅 All Sessions"
        ws.sheet_view.showGridLines=False; ws.sheet_properties.tabColor=CY
        for w,c in zip([2,14,10,16,14,12,22,2],range(1,9)):
            ws.column_dimensions[get_column_letter(c)].width=w
        for r in range(1,500):
            ws.row_dimensions[r].height=20
            for c in range(1,9): ws.cell(r,c).fill=PatternFill("solid",start_color=BX)

        ws.merge_cells("B2:G2"); ws["B2"]="KRISH BARAL — STUDY LOG"
        ws["B2"].font=Font(name="Courier New",bold=True,size=16,color=CY)
        ws["B2"].alignment=Alignment(horizontal="center",vertical="center")
        ws["B2"].fill=PatternFill("solid",start_color=BX)
        ws.row_dimensions[2].height=34

        for i,(h,col) in enumerate(zip(["DATE","MINUTES","SUBJECT","MOOD","POMODOROS","TIME"],[CY,GR,"64C8FF",GO,GR,WH])):
            c=ws.cell(4,i+2,value=h)
            c.font=Font(name="Courier New",bold=True,size=9,color=col)
            c.alignment=Alignment(horizontal="center",vertical="center")
            c.fill=PatternFill("solid",start_color=DM)
        ws.row_dimensions[4].height=22

        mc={"crushed":GR,"okay":GO,"struggled":RE}
        sessions=data.get("sessions",[])
        for i,s in enumerate(sessions):
            r=5+i; rb="0A100C" if i%2==0 else BX
            for c in range(1,9): ws.cell(r,c).fill=PatternFill("solid",start_color=rb)
            ws.row_dimensions[r].height=20
            sc(ws,r,2,s.get("date",""),sz=9,col=WH,bg=rb)
            sc(ws,r,3,s.get("minutes",0),sz=11,bold=True,col=CY,bg=rb)
            sc(ws,r,4,s.get("subject",""),sz=9,col="64C8FF",bg=rb)
            sc(ws,r,5,s.get("mood","").upper(),sz=9,col=mc.get(s.get("mood",""),WH),bg=rb)
            sc(ws,r,6,max(0,s.get("minutes",0)//25),sz=10,bold=True,col=GR,bg=rb)
            sc(ws,r,7,s.get("time",""),sz=9,col=WH,bg=rb)

        if sessions:
            tr=5+len(sessions)+1
            ws.row_dimensions[tr].height=24
            for c in range(1,9): ws.cell(tr,c).fill=PatternFill("solid",start_color=DM)
            sc(ws,tr,2,"TOTAL",bold=True,sz=10,col=CY,bg=DM)
            ws.cell(tr,3,value=f"=SUM(C5:C{4+len(sessions)})")
            ws.cell(tr,3).font=Font(name="Courier New",bold=True,size=12,color=CY)
            ws.cell(tr,3).alignment=Alignment(horizontal="center",vertical="center")
            ws.cell(tr,3).fill=PatternFill("solid",start_color=DM)

        ws2=wb.create_sheet("📊 Stats")
        ws2.sheet_view.showGridLines=False; ws2.sheet_properties.tabColor=GR
        for r in range(1,40):
            ws2.row_dimensions[r].height=22
            for c in range(1,8): ws2.cell(r,c).fill=PatternFill("solid",start_color=BX)
        for w,c in zip([2,22,16,16,16,2],range(1,7)):
            ws2.column_dimensions[get_column_letter(c)].width=w

        ws2.merge_cells("B2:E2"); ws2["B2"]="STUDY STATS"
        ws2["B2"].font=Font(name="Courier New",bold=True,size=16,color=CY)
        ws2["B2"].alignment=Alignment(horizontal="center",vertical="center")
        ws2["B2"].fill=PatternFill("solid",start_color=BX)
        ws2.row_dimensions[2].height=34

        stats=[("STREAK",f"{data.get('streak',0)} days",GO),
               ("SESSIONS",str(data.get("total_sessions",0)),CY),
               ("TOTAL MINS",str(data.get("total_minutes",0)),GR),
               ("TOTAL HOURS",f"{round(data.get('total_minutes',0)/60,1)}h","64C8FF")]
        for i,(label,val,col) in enumerate(stats):
            r=4+i*3; ws2.row_dimensions[r].height=16; ws2.row_dimensions[r+1].height=28
            sc(ws2,r,2,label,bold=True,sz=8,col="606060",bg=DM)
            sc(ws2,r+1,2,val,bold=True,sz=14,col=col,bg=DM)

        ws2.row_dimensions[18].height=24
        ws2.merge_cells("B18:E18"); ws2["B18"]="SUBJECT BREAKDOWN"
        ws2["B18"].font=Font(name="Courier New",bold=True,size=11,color=CY)
        ws2["B18"].alignment=Alignment(horizontal="center",vertical="center")
        ws2["B18"].fill=PatternFill("solid",start_color=DM)

        subj={}
        for s in sessions: subj[s.get("subject","?")]=subj.get(s.get("subject","?"),0)+s.get("minutes",0)
        for i,(sn,sm) in enumerate(sorted(subj.items(),key=lambda x:-x[1])):
            r=20+i*2; col=[CY,"64C8FF",GR,GO][i%4]
            sc(ws2,r,2,sn,bold=True,sz=10,col=col,bg=BX)
            sc(ws2,r,3,f"{sm} min",bold=True,sz=10,col=col,bg=BX)
            sc(ws2,r,4,f"{round(sm/60,1)}h",sz=9,col="506050",bg=BX)
            ws2.row_dimensions[r].height=24

        # Topic Progress Sheet
        ws4=wb.create_sheet("📚 Topic Progress")
        ws4.sheet_view.showGridLines=False; ws4.sheet_properties.tabColor="00FFE7"
        for r in range(1,60):
            ws4.row_dimensions[r].height=20
            for c in range(1,7): ws4.cell(r,c).fill=PatternFill("solid",start_color=BX)
        for w,c in zip([2,20,30,14,2],range(1,6)):
            ws4.column_dimensions[get_column_letter(c)].width=w

        ws4.merge_cells("B2:D2"); ws4["B2"]="IOE TOPIC PROGRESS — KRISH"
        ws4["B2"].font=Font(name="Courier New",bold=True,size=14,color=CY)
        ws4["B2"].alignment=Alignment(horizontal="center",vertical="center")
        ws4["B2"].fill=PatternFill("solid",start_color=BX)
        ws4.row_dimensions[2].height=30

        topic_progress = data.get("topic_progress", {})
        row = 4
        subj_colors = {"MATHEMATICS": CY, "PHYSICS": "64C8FF", "CHEMISTRY": GR, "ENGLISH": GO}
        for subj_name, topics in CURRICULUM.items():
            tidx = topic_progress.get(subj_name, 0)
            scol = subj_colors.get(subj_name, WH)
            ws4.merge_cells(f"B{row}:D{row}")
            ws4.cell(row, 2, value=f"── {subj_name} ──  ({tidx}/{len(topics)} done)")
            ws4.cell(row, 2).font=Font(name="Courier New",bold=True,size=10,color=scol)
            ws4.cell(row, 2).fill=PatternFill("solid",start_color=DM)
            ws4.cell(row, 2).alignment=Alignment(horizontal="left",vertical="center")
            ws4.row_dimensions[row].height=22
            row += 1
            for i, t in enumerate(topics):
                done = i < tidx
                status = "✓ DONE" if done else ("→ CURRENT" if i == tidx else "○ UPCOMING")
                sc_col = GR if done else (scol if i == tidx else "304030")
                ws4.cell(row, 2, value=t["topic"])
                ws4.cell(row, 2).font=Font(name="Courier New",size=9,color=sc_col)
                ws4.cell(row, 2).fill=PatternFill("solid",start_color="0A100C" if done else BX)
                ws4.cell(row, 2).alignment=Alignment(horizontal="left",vertical="center")
                ws4.cell(row, 3, value=t["detail"])
                ws4.cell(row, 3).font=Font(name="Courier New",size=8,color="405040")
                ws4.cell(row, 3).fill=PatternFill("solid",start_color="0A100C" if done else BX)
                ws4.cell(row, 3).alignment=Alignment(horizontal="left",vertical="center")
                ws4.cell(row, 4, value=status)
                ws4.cell(row, 4).font=Font(name="Courier New",bold=True,size=8,color=sc_col)
                ws4.cell(row, 4).fill=PatternFill("solid",start_color="0A100C" if done else BX)
                ws4.cell(row, 4).alignment=Alignment(horizontal="center",vertical="center")
                row += 1
            row += 1

        ws3=wb.create_sheet("🎯 IOE Countdown")
        ws3.sheet_view.showGridLines=False; ws3.sheet_properties.tabColor="FF4444"
        for r in range(1,20):
            ws3.row_dimensions[r].height=28
            for c in range(1,7): ws3.cell(r,c).fill=PatternFill("solid",start_color=BX)
        for w,c in zip([2,22,18,18,2],range(1,6)):
            ws3.column_dimensions[get_column_letter(c)].width=w

        days=(IOE_DATE-date.today()).days
        ws3.merge_cells("B2:D2"); ws3["B2"]="IOE ENTRANCE EXAM"
        ws3["B2"].font=Font(name="Courier New",bold=True,size=18,color=CY)
        ws3["B2"].alignment=Alignment(horizontal="center",vertical="center")
        ws3["B2"].fill=PatternFill("solid",start_color=BX)
        ws3.row_dimensions[2].height=40

        dcol=CY if days>60 else GO if days>30 else RE
        ws3.merge_cells("B4:D4"); ws3["B4"]=f"{days} DAYS REMAINING"
        ws3["B4"].font=Font(name="Courier New",bold=True,size=28,color=dcol)
        ws3["B4"].alignment=Alignment(horizontal="center",vertical="center")
        ws3["B4"].fill=PatternFill("solid",start_color=BX)
        ws3.row_dimensions[4].height=52

        for i,(label,val,col) in enumerate([
            ("Exam Date",str(IOE_DATE),GO),
            ("Days Left",str(days),dcol),
            ("Hours Studied",f"{round(data.get('total_minutes',0)/60,1)}h",GR),
            ("Sessions Done",str(data.get("total_sessions",0)),CY),
            ("Current Streak",f"{data.get('streak',0)} days",GO),
        ]):
            r=7+i*2
            sc(ws3,r,2,label,bold=True,sz=9,col="606060",bg=DM)
            sc(ws3,r,3,val,bold=True,sz=12,col=col,bg=DM)
            ws3.row_dimensions[r].height=22

        wb.save(EXCEL_FILE)
    except Exception as e: print(f"Excel: {e}")

# ══════════════════════════════════════════
#  WEBSITE BLOCKING
# ══════════════════════════════════════════
BLOCKED=["instagram.com","www.instagram.com","facebook.com","www.facebook.com",
         "twitter.com","www.twitter.com","tiktok.com","www.tiktok.com","reddit.com","www.reddit.com"]

def _run_as_admin(ps_script):
    import tempfile
    tmp=tempfile.NamedTemporaryFile(mode='w',suffix='.ps1',delete=False)
    tmp.write(ps_script); tmp.close()
    subprocess.run([
        "powershell","-ExecutionPolicy","Bypass","-WindowStyle","Hidden","-Command",
        f'Start-Process powershell -ArgumentList \'-ExecutionPolicy Bypass -WindowStyle Hidden -File "{tmp.name}"\' -Verb RunAs -Wait'
    ],capture_output=True,timeout=15)
    try: os.remove(tmp.name)
    except: pass

def block_websites():
    try:
        lines="\n".join([f"127.0.0.1 {s}" for s in BLOCKED])
        ps=f'''$h="C:\\Windows\\System32\\drivers\\etc\\hosts"
$c=Get-Content $h
@"{lines}"@ -split "`n" | ForEach-Object {{
    if ($c -notcontains $_) {{ Add-Content $h $_ -Encoding ASCII }}
}}
ipconfig /flushdns | Out-Null'''
        _run_as_admin(ps)
    except: pass

def unblock_websites():
    try:
        ps='''$h="C:\\Windows\\System32\\drivers\\etc\\hosts"
$c = Get-Content $h | Where-Object { $_ -notmatch "instagram|facebook|twitter|tiktok|reddit" }
$c | Set-Content $h -Encoding ASCII
ipconfig /flushdns | Out-Null'''
        _run_as_admin(ps)
    except: pass

def close_brave():
    try: subprocess.run(["powershell","-Command","Get-Process brave -ErrorAction SilentlyContinue|Stop-Process -Force"],capture_output=True)
    except: pass

def close_edge():
    try: subprocess.run(["powershell","-Command","Get-Process msedge -ErrorAction SilentlyContinue|Stop-Process -Force"],capture_output=True)
    except: pass

# ══════════════════════════════════════════
#  TRAY
# ══════════════════════════════════════════
def make_tray_icon(active=False):
    img=Image.new("RGB",(64,64),(20,20,20))
    draw=ImageDraw.Draw(img)
    draw.ellipse([8,8,56,56],fill=(0,200,140) if active else (80,80,80))
    if active: draw.ellipse([20,20,44,44],fill=(0,120,80))
    return img

def update_tray():
    if not tray_icon: return
    tray_icon.icon=make_tray_icon(study_active)
    tray_icon.title="📚 Study Mode: ACTIVE" if study_active else "Study Mode — Press F9"

# ══════════════════════════════════════════
#  POMODORO
# ══════════════════════════════════════════
def pomodoro_loop():
    session=0
    while study_active and running:
        session+=1
        for _ in range(POMODORO_WORK):
            if not study_active or not running: return
            time.sleep(1)
        if not study_active: return
        play_voice("break")
        notification.notify(title="⏱ Break Time!",message=f"Session {session} done! 5 min break.",timeout=8)
        for _ in range(POMODORO_BREAK):
            if not study_active or not running: return
            time.sleep(1)
        if not study_active: return
        play_voice("resume")
        notification.notify(title="💪 Back to Work!",message=f"Session {session+1}. Let's go Krish.",timeout=5)

# ══════════════════════════════════════════
#  ACTIVATE / DEACTIVATE
# ══════════════════════════════════════════
def do_activate_final(subject):
    """Actually starts the study session after plan popup."""
    global study_active, study_start_time, current_subject
    current_subject=subject; study_active=True; study_start_time=time.time()
    subject_url=SUBJECT_URLS.get(subject,SUBJECT_URLS['General'])

    # Open resource for current topic
    topic_data, idx, total = get_current_topic(subject)
    if topic_data:
        resource_url = topic_data.get("resource", subject_url)
    else:
        resource_url = subject_url

    close_edge()
    subprocess.Popen([BRAVE, LOFI_URL]); time.sleep(2)
    subprocess.Popen([BRAVE, resource_url])

    threading.Thread(target=lambda: play_voice(f"start_{random.randint(0,3)}"),daemon=True).start()
    play_sound("start")
    threading.Thread(target=pomodoro_loop,daemon=True).start()
    run_on_main_thread(lambda: study_timer.start(subject))
    update_tray()

    topic_name = topic_data["topic"] if topic_data else subject
    notification.notify(title=f"📚 {subject} — {topic_name}",message="F9 to stop. Let's go, Krish!",timeout=4)

def do_activate(subject):
    """Show today's plan first, then start."""
    run_on_main_thread(lambda: show_today_plan(subject, lambda: do_activate_final(subject)))

def activate_study_mode():
    hour=datetime.now().hour
    if hour>=LATE_NIGHT_H:
        def after_warn(proceed):
            if proceed:
                threading.Thread(target=lambda: show_subject_selector(do_activate),daemon=True).start()
        run_on_main_thread(lambda: show_late_night_warning(hour,after_warn))
    else:
        threading.Thread(target=lambda: show_subject_selector(do_activate),daemon=True).start()

def do_stop():
    global study_active, study_start_time
    study_active=False
    minutes=max(1,int((time.time()-study_start_time)/60))
    study_start_time=None
    subj=current_subject or "General"
    close_brave(); study_timer.close(); update_tray()

    def after_rating(mood):
        data=update_streak(minutes,mood,subj)
        msgs={
            "crushed": f"That's {data['streak']} days straight. Outstanding, sir.",
            "okay":    f"Consistency beats intensity. {data['streak']} day streak.",
            "struggled":f"Hard days build the foundation. {data['streak']} day streak.",
        }
        def play_end():
            if mood=="crushed":   play_sound("party");  time.sleep(1.8); play_voice("end_crushed")
            elif mood=="okay":    play_sound("chime");  time.sleep(0.8); play_voice("end_okay")
            else:                 play_sound("gentle"); time.sleep(0.6); play_voice("end_struggled")
        threading.Thread(target=play_end,daemon=True).start()
        notification.notify(title=f"✅ {minutes} min — {subj}",message=msgs.get(mood,""),timeout=10)

    run_on_main_thread(lambda: show_rating_popup(minutes,subj,after_rating))

def deactivate_study_mode():
    if not study_active: return
    elapsed=int((time.time()-study_start_time)/60)
    if elapsed<(MIN_SESSION//60):
        def after_warn(stop):
            if stop: do_stop()
        run_on_main_thread(lambda: show_min_session_warning(elapsed,after_warn))
    else:
        do_stop()

# ══════════════════════════════════════════
#  F9
# ══════════════════════════════════════════
# Updated popup and Pomodoro behavior overrides.
class PomodoroState:
    def __init__(self, work_secs, break_secs):
        self.work_secs = work_secs
        self.break_secs = break_secs
        self.lock = threading.Lock()
        self.stop_event = threading.Event()
        self.running = False
        self.phase = "idle"
        self.session = 0
        self.phase_started_at = 0.0
        self.paused = False
        self.pause_started_at = None
        self.paused_accumulated = 0.0
        self.focus_accumulated = 0

    def start(self):
        with self.lock:
            self.stop_event = threading.Event()
            self.running = True
            self.phase = "work"
            self.session = 1
            self.phase_started_at = time.monotonic()
            self.paused = False
            self.pause_started_at = None
            self.paused_accumulated = 0.0
            self.focus_accumulated = 0

    def stop(self):
        with self.lock:
            self.running = False
            self.paused = False
            self.pause_started_at = None
            self.stop_event.set()

    def toggle_pause(self):
        with self.lock:
            if not self.running:
                return False
            now = time.monotonic()
            if self.paused:
                if self.pause_started_at is not None:
                    self.paused_accumulated += now - self.pause_started_at
                self.pause_started_at = None
                self.paused = False
            else:
                self.paused = True
                self.pause_started_at = now
            return self.paused

    def _phase_elapsed_locked(self, now=None):
        if not self.running or self.phase == "idle":
            return 0
        now = now or time.monotonic()
        ref = self.pause_started_at if (self.paused and self.pause_started_at is not None) else now
        return max(0, int(ref - self.phase_started_at - self.paused_accumulated))

    def snapshot(self):
        with self.lock:
            phase = self.phase if self.phase != "idle" else "work"
            phase_total = self.break_secs if phase == "break" else self.work_secs
            elapsed = self._phase_elapsed_locked()
            remaining = max(0, phase_total - elapsed)
            focus_seconds = self.focus_accumulated + (elapsed if phase == "work" and self.running else 0)
            progress = 0.0 if phase_total <= 0 else min(1.0, elapsed / phase_total)
            return {
                "running": self.running,
                "phase": phase,
                "session": max(1, self.session),
                "paused": self.paused,
                "elapsed": elapsed,
                "remaining": remaining,
                "phase_total": phase_total,
                "progress": progress,
                "focus_seconds": focus_seconds,
            }

    def focus_seconds(self):
        return self.snapshot()["focus_seconds"]

    def advance_if_needed(self):
        with self.lock:
            if not self.running or self.paused or self.phase == "idle":
                return None
            now = time.monotonic()
            elapsed = self._phase_elapsed_locked(now)
            phase_total = self.break_secs if self.phase == "break" else self.work_secs
            if elapsed < phase_total:
                return None

            if self.phase == "work":
                finished_session = self.session
                self.focus_accumulated += self.work_secs
                self.phase = "break"
                next_session = self.session
            else:
                finished_session = self.session
                self.phase = "work"
                self.session += 1
                next_session = self.session

            self.phase_started_at = now
            self.paused_accumulated = 0.0
            self.pause_started_at = None
            self.paused = False
            return {
                "new_phase": self.phase,
                "finished_session": finished_session,
                "next_session": next_session,
            }


pomodoro_state = PomodoroState(POMODORO_WORK, POMODORO_BREAK)

def get_effective_study_seconds():
    focus_seconds = pomodoro_state.focus_seconds()
    if focus_seconds > 0:
        return focus_seconds
    if study_start_time:
        return max(0, int(time.time() - study_start_time))
    return 0

POPUP_BG = "#0b1114"
POPUP_PANEL = "#111b20"
POPUP_TEXT = "#f3fbff"
POPUP_MUTED = "#b8c7ce"
POPUP_DIM = "#7f949c"
BUTTON_BG = "#132127"
BUTTON_ACTIVE = "#1a2b32"

def make_readable_popup(title, width, height, accent, title_color=None):
    root = tk.Toplevel(_tk_root)
    root.overrideredirect(True)
    root.attributes("-topmost", True)
    root.configure(bg=accent)
    sw = root.winfo_screenwidth()
    sh = root.winfo_screenheight()
    root.geometry(f"{width}x{height}+{(sw-width)//2}+{(sh-height)//2}")

    panel = tk.Frame(root, bg=POPUP_BG)
    panel.place(x=3, y=3, width=width - 6, height=height - 6)

    tk.Label(
        panel,
        text=title,
        font=("Consolas", 22, "bold"),
        fg=title_color or accent,
        bg=POPUP_BG,
    ).pack(pady=(18, 8))

    tk.Frame(panel, bg="#163038", height=2).pack(fill="x", padx=28)
    root.bind("<Escape>", lambda event: root.destroy())
    root.grab_set()
    root.focus_force()
    return root, panel

def popup_section(parent, title):
    section = tk.Frame(parent, bg=POPUP_PANEL)
    section.pack(fill="x", padx=28, pady=(0, 16))
    tk.Label(
        section,
        text=title,
        font=("Consolas", 12, "bold"),
        fg="#4fd7e8",
        bg=POPUP_PANEL,
        anchor="w",
    ).pack(fill="x", padx=18, pady=(14, 6))
    return section

def popup_button(parent, text, fg, command):
    return tk.Button(
        parent,
        text=text,
        font=("Segoe UI", 14, "bold"),
        bg=BUTTON_BG,
        fg=fg,
        activebackground=BUTTON_ACTIVE,
        activeforeground=fg,
        relief="flat",
        bd=0,
        cursor="hand2",
        command=command,
        highlightbackground=fg,
        highlightthickness=2,
        padx=12,
        pady=12,
    )

def show_subject_selector(callback):
    root, panel = make_readable_popup("SELECT YOUR SUBJECT", 760, 520, CYAN)
    tk.Label(
        panel,
        text="Choose what you want to study today.",
        font=("Segoe UI", 16),
        fg=POPUP_MUTED,
        bg=POPUP_BG,
    ).pack(pady=(16, 14))

    grid = tk.Frame(panel, bg=POPUP_BG)
    grid.pack(fill="both", expand=True, padx=28, pady=(4, 18))
    subjects = [
        ("MATHEMATICS", "Calculus, Algebra, Trigonometry", CYAN),
        ("PHYSICS", "Mechanics, Waves, Electricity", "#50b4ff"),
        ("CHEMISTRY", "Organic, Inorganic, Physical", "#8cff72"),
        ("ENGLISH", "Grammar, Reading, Writing", "#ffbe50"),
    ]
    last = load_streak().get("last_subject", "")

    def pick_subject(name):
        root.destroy()
        callback(name)

    for idx, (name, subtitle, accent) in enumerate(subjects):
        card = tk.Frame(grid, bg=POPUP_PANEL, highlightbackground=accent, highlightthickness=3 if name == last else 2)
        card.grid(row=idx // 2, column=idx % 2, sticky="nsew", padx=10, pady=10)
        grid.grid_columnconfigure(idx % 2, weight=1)
        grid.grid_rowconfigure(idx // 2, weight=1)

        if name == last:
            tk.Label(card, text="LAST SESSION", font=("Consolas", 10, "bold"), fg=accent, bg=POPUP_PANEL).pack(anchor="e", padx=16, pady=(12, 0))
        else:
            tk.Frame(card, bg=POPUP_PANEL, height=18).pack()

        tk.Label(card, text=name, font=("Consolas", 20, "bold"), fg=accent, bg=POPUP_PANEL).pack(pady=(6, 8))
        tk.Label(card, text=subtitle, font=("Segoe UI", 13), fg=POPUP_TEXT, bg=POPUP_PANEL).pack(pady=(0, 8))

        topic_data, _, _ = get_current_topic(name)
        next_topic = topic_data["topic"] if topic_data else "No topic available"
        tk.Label(
            card,
            text=f"Next topic: {next_topic}",
            font=("Segoe UI", 12),
            fg=POPUP_MUTED,
            bg=POPUP_PANEL,
            wraplength=280,
            justify="center",
        ).pack(padx=18, pady=(0, 16))

        popup_button(card, f"Study {name.title()}", accent, lambda subj=name: pick_subject(subj)).pack(fill="x", padx=18, pady=(0, 18))

    tk.Label(
        panel,
        text="Press ESC to skip and open the default subject.",
        font=("Segoe UI", 12),
        fg=POPUP_DIM,
        bg=POPUP_BG,
    ).pack(pady=(0, 18))

    root.bind("<Escape>", lambda event: (root.destroy(), callback("General")))
    root.wait_window()

def show_today_plan(subject, on_start):
    topic_data, idx, total = get_current_topic(subject)
    if not topic_data:
        on_start()
        return

    topic = topic_data["topic"]
    detail = topic_data["detail"]
    resource = topic_data["resource"]
    week = topic_data["week"]
    pct_done = int((idx / total) * 100)
    subject_colors = {
        "MATHEMATICS": CYAN,
        "PHYSICS": "#50b4ff",
        "CHEMISTRY": "#8cff72",
        "ENGLISH": "#ffbe50",
    }
    accent = subject_colors.get(subject, CYAN)
    root, panel = make_readable_popup("TODAY'S STUDY PLAN", 940, 720, accent)

    badge_row = tk.Frame(panel, bg=POPUP_BG)
    badge_row.pack(fill="x", padx=28, pady=(18, 18))
    for text, fg, bg in [
        (subject, "#071014", accent),
        (f"WEEK {week}", accent, POPUP_PANEL),
        (f"{idx+1} / {total} topics", "#96ffbd", POPUP_PANEL),
    ]:
        label = tk.Label(
            badge_row,
            text=text,
            font=("Consolas", 15, "bold"),
            fg=fg,
            bg=bg,
            padx=18,
            pady=10,
            highlightbackground=accent if bg == POPUP_PANEL else bg,
            highlightthickness=2,
        )
        label.pack(side="left", padx=(0, 14))

    tk.Label(
        panel,
        text="TOPIC",
        font=("Consolas", 12, "bold"),
        fg=POPUP_DIM,
        bg=POPUP_BG,
    ).pack(anchor="w", padx=30)
    tk.Label(
        panel,
        text=topic,
        font=("Consolas", 28, "bold"),
        fg=accent,
        bg=POPUP_BG,
        wraplength=840,
        justify="center",
    ).pack(fill="x", padx=28, pady=(6, 20))

    detail_section = popup_section(panel, "WHAT TO STUDY")
    tk.Label(
        detail_section,
        text=detail,
        font=("Segoe UI", 18),
        fg=POPUP_TEXT,
        bg=POPUP_PANEL,
        justify="left",
        wraplength=820,
        anchor="w",
    ).pack(fill="x", padx=18, pady=(0, 18))

    tips_section = popup_section(panel, "HOW TO STUDY THIS")
    tips_text = "\n".join([
        "1. Watch one clear lesson on this topic for 15 to 20 minutes.",
        "2. Write the key rule, formula, or idea by hand before solving anything.",
        "3. Solve at least 10 practice questions while the topic is fresh.",
        "4. Reopen the resource only for the exact step that is confusing you.",
    ])
    tk.Label(
        tips_section,
        text=tips_text,
        font=("Segoe UI", 17),
        fg=POPUP_TEXT,
        bg=POPUP_PANEL,
        justify="left",
        wraplength=820,
        anchor="w",
    ).pack(fill="x", padx=18, pady=(0, 18))

    progress_frame = popup_section(panel, "SYLLABUS PROGRESS")
    tk.Label(
        progress_frame,
        text=f"{pct_done}% complete",
        font=("Segoe UI", 16, "bold"),
        fg=POPUP_TEXT,
        bg=POPUP_PANEL,
        anchor="w",
    ).pack(fill="x", padx=18, pady=(0, 10))
    bar = tk.Frame(progress_frame, bg="#193128", height=22)
    bar.pack(fill="x", padx=18, pady=(0, 18))
    fill = tk.Frame(bar, bg="#3ae38f" if pct_done < 70 else "#ffd166" if pct_done < 90 else "#ff7a7a")
    fill.place(relheight=1, relwidth=max(0.0, pct_done / 100))

    tk.Label(
        panel,
        text="Open the resource first if you want a quick refresher, then start the session.",
        font=("Segoe UI", 13),
        fg=POPUP_MUTED,
        bg=POPUP_BG,
    ).pack(padx=28, pady=(2, 14))

    def open_resource():
        try:
            subprocess.Popen([BRAVE, resource])
        except:
            pass

    def start_session(start_fn):
        root.destroy()
        threading.Thread(target=start_fn, daemon=True).start()

    def mark_done_start():
        save_topic_progress(subject, min(idx + 1, total - 1))
        start_session(on_start)

    button_row = tk.Frame(panel, bg=POPUP_BG)
    button_row.pack(fill="x", padx=28, pady=(0, 24))
    popup_button(button_row, "OPEN RESOURCE", accent, open_resource).pack(side="left", expand=True, fill="x", padx=(0, 10))
    popup_button(button_row, "START SESSION", CYAN, lambda: start_session(on_start)).pack(side="left", expand=True, fill="x", padx=10)
    popup_button(button_row, "DONE AND NEXT", GREEN, mark_done_start).pack(side="left", expand=True, fill="x", padx=(10, 0))
    root.wait_window()

def show_late_night_warning(hour, callback):
    root, panel = make_readable_popup("LATE NIGHT WARNING", 700, 360, "#ffbe50", "#ffbe50")
    tk.Label(
        panel,
        text=f"It is already after {hour}:00.",
        font=("Segoe UI", 22, "bold"),
        fg=POPUP_TEXT,
        bg=POPUP_BG,
    ).pack(pady=(32, 8))
    tk.Label(
        panel,
        text="Sleep helps memory, focus, and retention.\nIf you still want to study, you can continue.",
        font=("Segoe UI", 17),
        fg=POPUP_MUTED,
        bg=POPUP_BG,
        justify="center",
    ).pack(pady=(0, 24))

    button_row = tk.Frame(panel, bg=POPUP_BG)
    button_row.pack(fill="x", padx=28, pady=(0, 18))
    popup_button(button_row, "YES, LET'S GO", CYAN, lambda: (root.destroy(), callback(True))).pack(side="left", expand=True, fill="x", padx=(0, 10))
    popup_button(button_row, "NO, I'LL SLEEP", RED, lambda: (root.destroy(), callback(False))).pack(side="left", expand=True, fill="x", padx=(10, 0))

    tk.Label(
        panel,
        text="You can override this warning any time.",
        font=("Segoe UI", 12),
        fg=POPUP_DIM,
        bg=POPUP_BG,
    ).pack(pady=(0, 18))
    root.wait_window()

def show_min_session_warning(elapsed_min, callback):
    root, panel = make_readable_popup("STOPPING EARLY?", 700, 360, "#ffd166", "#ffd166")
    tk.Label(
        panel,
        text=f"You have only studied for {elapsed_min} minutes.",
        font=("Segoe UI", 22, "bold"),
        fg=POPUP_TEXT,
        bg=POPUP_BG,
    ).pack(pady=(30, 8))
    tk.Label(
        panel,
        text="A longer session usually helps you retain more. Do you want to keep going or stop anyway?",
        font=("Segoe UI", 17),
        fg=POPUP_MUTED,
        bg=POPUP_BG,
        wraplength=620,
        justify="center",
    ).pack(pady=(0, 18))

    progress = min(1.0, elapsed_min / max(1, MIN_SESSION // 60))
    bar = tk.Frame(panel, bg="#193128", height=24)
    bar.pack(fill="x", padx=34, pady=(0, 20))
    tk.Frame(bar, bg="#3ae38f" if progress >= 1 else "#ffd166" if progress >= 0.5 else "#ff7a7a").place(relheight=1, relwidth=progress)

    button_row = tk.Frame(panel, bg=POPUP_BG)
    button_row.pack(fill="x", padx=28, pady=(0, 22))
    popup_button(button_row, "KEEP GOING", CYAN, lambda: (root.destroy(), callback(False))).pack(side="left", expand=True, fill="x", padx=(0, 10))
    popup_button(button_row, "STOP ANYWAY", RED, lambda: (root.destroy(), callback(True))).pack(side="left", expand=True, fill="x", padx=(10, 0))
    root.wait_window()

def show_rating_popup(minutes, subject, callback):
    root, panel = make_readable_popup("SESSION COMPLETE", 660, 420, CYAN)
    tk.Label(
        panel,
        text=f"{minutes} minutes in {subject}",
        font=("Segoe UI", 20, "bold"),
        fg=POPUP_TEXT,
        bg=POPUP_BG,
    ).pack(pady=(28, 8))
    tk.Label(
        panel,
        text="How did this session feel?",
        font=("Segoe UI", 17),
        fg=POPUP_MUTED,
        bg=POPUP_BG,
    ).pack(pady=(0, 22))

    choices = tk.Frame(panel, bg=POPUP_BG)
    choices.pack(fill="x", padx=28, pady=(0, 24))
    popup_button(choices, "CRUSHED IT", CYAN, lambda: (root.destroy(), callback("crushed"))).pack(fill="x", pady=(0, 12))
    popup_button(choices, "IT WAS OKAY", GOLD, lambda: (root.destroy(), callback("okay"))).pack(fill="x", pady=12)
    popup_button(choices, "STRUGGLED", RED, lambda: (root.destroy(), callback("struggled"))).pack(fill="x", pady=(12, 0))
    root.wait_window()

def show_ioe_countdown():
    days = (IOE_DATE - date.today()).days
    data = load_streak()
    total_hrs = round(data.get("total_minutes", 0) / 60, 1)
    accent = CYAN if days > 60 else GOLD if days > 30 else RED
    root, panel = make_readable_popup("IOE ENTRANCE COUNTDOWN", 640, 360, accent, accent)
    tk.Label(
        panel,
        text=str(days),
        font=("Consolas", 52, "bold"),
        fg=accent,
        bg=POPUP_BG,
    ).pack(pady=(26, 0))
    tk.Label(
        panel,
        text="days remaining",
        font=("Segoe UI", 18),
        fg=POPUP_TEXT,
        bg=POPUP_BG,
    ).pack(pady=(0, 18))
    tk.Label(
        panel,
        text=f"Studied so far: {total_hrs} hours\nCurrent streak: {data.get('streak', 0)} days",
        font=("Segoe UI", 17),
        fg=POPUP_MUTED,
        bg=POPUP_BG,
        justify="center",
    ).pack(pady=(0, 22))
    popup_button(panel, "CLOSE", accent, root.destroy).pack(fill="x", padx=28, pady=(0, 24))

class StudyTimer:
    W, H = 320, 80

    def __init__(self):
        self.root = None
        self.canvas = None
        self._tk_img = None
        self._running = False
        self.subject = "General"
        self._pause_rect = (0, 0, 0, 0)
        self._stop_rect = (0, 0, 0, 0)

    def start(self, subject):
        self.close()
        self.subject = subject
        self._running = True
        self._run_window()

    def _run_window(self):
        self.root = tk.Toplevel(_tk_root)
        self.root.overrideredirect(True)
        self.root.attributes("-topmost", True)
        self.root.attributes("-alpha", 0.93)
        self.root.configure(bg=BG)
        sw = self.root.winfo_screenwidth()
        self.root.geometry(f"{self.W}x{self.H}+{sw-self.W-12}+8")
        self.canvas = tk.Canvas(self.root, width=self.W, height=self.H, bg=BG, highlightthickness=0, cursor="hand2")
        self.canvas.pack()
        self._drag_x = 0
        self._drag_y = 0
        self._dragging = False
        self.canvas.bind("<ButtonPress-1>", self._drag_start)
        self.canvas.bind("<B1-Motion>", self._drag_move)
        self.canvas.bind("<ButtonRelease-1>", self._on_release)
        self._update()

    def _drag_start(self, event):
        self._drag_x = event.x_root - self.root.winfo_x()
        self._drag_y = event.y_root - self.root.winfo_y()
        self._dragging = False
        self._press_x = event.x
        self._press_y = event.y

    def _drag_move(self, event):
        if abs(event.x - self._press_x) > 4 or abs(event.y - self._press_y) > 4:
            self._dragging = True
        if self._dragging:
            self.root.geometry(f"+{event.x_root-self._drag_x}+{event.y_root-self._drag_y}")

    def _on_release(self, event):
        if self._dragging:
            self._dragging = False
            return
        x, y = event.x, event.y
        px1, py1, px2, py2 = self._pause_rect
        sx1, sy1, sx2, sy2 = self._stop_rect
        if px1 <= x <= px2 and py1 <= y <= py2:
            pomodoro_state.toggle_pause()
        elif sx1 <= x <= sx2 and sy1 <= y <= sy2:
            self._end_session()

    def _end_session(self):
        self._running = False
        try:
            if self.root:
                self.root.destroy()
                self.root = None
        except:
            pass
        threading.Thread(target=deactivate_study_mode, daemon=True).start()

    def close(self):
        self._running = False
        try:
            if self.root:
                self.root.destroy()
                self.root = None
        except:
            pass

    def _render(self, state):
        W, H = self.W, self.H
        subject_color = {
            "MATHEMATICS": (0, 255, 231),
            "PHYSICS": (80, 180, 255),
            "CHEMISTRY": (160, 255, 80),
            "ENGLISH": (255, 190, 80),
        }.get(self.subject, (0, 255, 231))
        paused = state["paused"]
        phase = state["phase"]
        phase_remaining = state["remaining"]
        phase_progress = state["progress"]
        focus_seconds = state["focus_seconds"]
        session = state["session"]
        phase_name = "BREAK" if phase == "break" else "FOCUS"
        phase_color = (255, 170, 70) if phase == "break" else subject_color
        border_color = (255, 80, 80) if paused else phase_color

        img = Image.new("RGB", (W, H), (5, 6, 8))
        draw = ImageDraw.Draw(img)
        for x in range(0, W, 20):
            draw.line([(x, 0), (x, H)], fill=(9, 13, 10), width=1)
        for y in range(0, H, 20):
            draw.line([(0, y), (W, y)], fill=(9, 13, 10), width=1)
        for i in range(4, 0, -1):
            draw.rounded_rectangle([i, i, W - i, H - i], radius=10, outline=border_color)
        draw.rounded_rectangle([0, 0, W - 1, H - 1], radius=10, outline=border_color, width=1)

        draw.rounded_rectangle([0, 0, 5, H - 1], radius=4, fill=subject_color)
        draw.text((12, 7), self.subject[:4], fill=subject_color, font=get_font(8, True))

        status_color = (255, 80, 80) if paused else ((255, 170, 70) if phase == "break" else (0, 220, 100))
        draw.ellipse([12, H - 19, 20, H - 11], fill=status_color)
        draw.text((24, H - 20), "PAUSED" if paused else phase_name, fill=status_color, font=get_font(7, True))

        topic_data, _, _ = get_current_topic(self.subject)
        if topic_data:
            draw.text((12, 20), topic_data["topic"][:22], fill=(0, 130, 90), font=get_font(7))

        timer = format_clock(phase_remaining)
        timer_font = get_font(18, True)
        bb = draw.textbbox((0, 0), timer, font=timer_font)
        tx = max(10, ((W - 72) - (bb[2] - bb[0])) // 2 + 8)
        draw.text((tx, 30), timer, fill=border_color, font=timer_font)

        bx1, by1, bx2, by2 = 12, H - 9, W - 12, H - 4
        draw.rounded_rectangle([bx1, by1, bx2, by2], radius=2, fill=(18, 35, 22))
        if phase_progress > 0:
            draw.rounded_rectangle([bx1, by1, bx1 + int((bx2 - bx1) * phase_progress), by2], radius=2, fill=phase_color)

        draw.text((120, 8), f"POMO {session}", fill=(0, 100, 75), font=get_font(7, True))
        draw.text((W - 150, H - 20), f"TOTAL {format_clock(focus_seconds, include_hours=True)}", fill=(0, 100, 75), font=get_font(7))

        PBX1, PBY1, PBX2, PBY2 = W - 68, 28, W - 36, 54
        pause_color = (255, 160, 40) if not paused else (0, 200, 100)
        draw.rounded_rectangle([PBX1, PBY1, PBX2, PBY2], radius=5, fill=(10, 18, 12), outline=pause_color, width=2)
        pause_text = "||" if not paused else ">"
        bb3 = draw.textbbox((0, 0), pause_text, font=get_font(10, True))
        draw.text((PBX1 + (PBX2 - PBX1 - (bb3[2] - bb3[0])) // 2, PBY1 + 4), pause_text, fill=pause_color, font=get_font(10, True))
        self._pause_rect = (PBX1, PBY1, PBX2, PBY2)

        SBX1, SBY1, SBX2, SBY2 = W - 32, 28, W - 4, 54
        draw.rounded_rectangle([SBX1, SBY1, SBX2, SBY2], radius=5, fill=(10, 18, 12), outline=(220, 50, 50), width=2)
        bb4 = draw.textbbox((0, 0), "[]", font=get_font(10, True))
        draw.text((SBX1 + (SBX2 - SBX1 - (bb4[2] - bb4[0])) // 2, SBY1 + 4), "[]", fill=(220, 50, 50), font=get_font(10, True))
        self._stop_rect = (SBX1, SBY1, SBX2, SBY2)
        return img

    def _update(self):
        if not self._running or not self.root:
            return
        try:
            self.canvas.delete("all")
            img = self._render(pomodoro_state.snapshot())
            self._tk_img = ImageTk.PhotoImage(img)
            self.canvas.create_image(0, 0, image=self._tk_img, anchor="nw")
            self.root.after(1000, self._update)
        except:
            pass


study_timer = StudyTimer()

def pomodoro_loop():
    while study_active and running and pomodoro_state.running:
        if pomodoro_state.stop_event.wait(1):
            return
        transition = pomodoro_state.advance_if_needed()
        if not transition:
            continue
        if transition["new_phase"] == "break":
            play_voice("break")
            notification.notify(
                title="Break Time",
                message=f"Pomodoro {transition['finished_session']} complete. Take 5 minutes.",
                timeout=8
            )
        else:
            play_voice("resume")
            notification.notify(
                title="Back to Work",
                message=f"Break finished. Pomodoro {transition['next_session']} starts now.",
                timeout=5
            )

def do_activate_final(subject):
    global study_active, study_start_time, current_subject
    current_subject = subject
    study_active = True
    study_start_time = time.time()
    subject_url = SUBJECT_URLS.get(subject, SUBJECT_URLS["General"])
    topic_data, _, _ = get_current_topic(subject)
    resource_url = topic_data.get("resource", subject_url) if topic_data else subject_url

    close_edge()
    subprocess.Popen([BRAVE, LOFI_URL])
    time.sleep(2)
    subprocess.Popen([BRAVE, resource_url])

    threading.Thread(target=lambda: play_voice(f"start_{random.randint(0,3)}"), daemon=True).start()
    play_sound("start")
    pomodoro_state.start()
    threading.Thread(target=pomodoro_loop, daemon=True).start()
    run_on_main_thread(lambda: study_timer.start(subject))
    update_tray()

    topic_name = topic_data["topic"] if topic_data else subject
    notification.notify(title=f"{subject} - {topic_name}", message="F9 to stop. Study mode is active.", timeout=4)

def do_activate(subject):
    run_on_main_thread(lambda: show_today_plan(subject, lambda: do_activate_final(subject)))

def activate_study_mode():
    hour = datetime.now().hour
    if hour >= LATE_NIGHT_H:
        def after_warn(proceed):
            if proceed:
                run_on_main_thread(lambda: show_subject_selector(do_activate))
        run_on_main_thread(lambda: show_late_night_warning(hour, after_warn))
    else:
        run_on_main_thread(lambda: show_subject_selector(do_activate))

def do_stop():
    global study_active, study_start_time
    study_active = False
    minutes = max(1, int(get_effective_study_seconds() / 60))
    study_start_time = None
    subj = current_subject or "General"
    pomodoro_state.stop()
    close_brave()
    run_on_main_thread(study_timer.close)
    update_tray()

    def after_rating(mood):
        data = update_streak(minutes, mood, subj)
        msgs = {
            "crushed": f"That's {data['streak']} days straight. Outstanding, sir.",
            "okay": f"Consistency beats intensity. {data['streak']} day streak.",
            "struggled": f"Hard days build the foundation. {data['streak']} day streak.",
        }

        def play_end():
            if mood == "crushed":
                play_sound("party")
                time.sleep(1.8)
                play_voice("end_crushed")
            elif mood == "okay":
                play_sound("chime")
                time.sleep(0.8)
                play_voice("end_okay")
            else:
                play_sound("gentle")
                time.sleep(0.6)
                play_voice("end_struggled")

        threading.Thread(target=play_end, daemon=True).start()
        notification.notify(title=f"{minutes} min - {subj}", message=msgs.get(mood, ""), timeout=10)

    run_on_main_thread(lambda: show_rating_popup(minutes, subj, after_rating))

def deactivate_study_mode():
    if not study_active:
        return
    elapsed = int(get_effective_study_seconds() / 60)
    if elapsed < (MIN_SESSION // 60):
        def after_warn(stop):
            if stop:
                do_stop()
        run_on_main_thread(lambda: show_min_session_warning(elapsed, after_warn))
    else:
        do_stop()

keyboard.add_hotkey('f9', lambda: deactivate_study_mode() if study_active else activate_study_mode())

# ══════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════
if __name__=="__main__":
    import sys

    # When the legacy file is launched directly, hand off to the updated app.
    sys.modules.setdefault("study_mode", sys.modules[__name__])
    from study_mode_krish import main as krish_main

    krish_main()
